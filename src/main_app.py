from collections import defaultdict
import math
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import unicodedata
import pandas as pd
from datetime import datetime, time
import customtkinter as ctk
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.animation import FuncAnimation
from matplotlib.colors import LinearSegmentedColormap
import os
import glob
import re
from openpyxl.styles import Border, Side
import requests
import sys
import tkinter.messagebox as messagebox
from packaging import version
from update_checker import AutoUpdater
from typing import List, Optional
from base_app import BaseApp
from temp_handler import temp_handler
import logging
import threading

checking_updates = False

# Configurar logging
logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)   

def check_for_updates():
    global checking_updates
    if checking_updates:
        return
    
    checking_updates = True
    try:
        updater = AutoUpdater()
        needs_update, latest_version = updater.check_for_updates()
        
        if needs_update:
            respuesta = messagebox.askyesno(
                "Actualización Disponible",
                f"Hay una nueva versión disponible: v{latest_version}\n¿Deseas actualizar ahora?",
                icon='info'
            )
            
            if respuesta:
                progress = tk.Toplevel()
                progress.title("Actualizando")
                progress.geometry("300x100")
                
                label = ttk.Label(progress, text="Descargando actualización...")
                label.pack(pady=10)
                
                pb = ttk.Progressbar(progress, mode='indeterminate')
                pb.pack(padx=20, fill='x')
                pb.start()
                
                def update_task():
                    try:
                        update_file = updater.download_update(latest_version)
                        if update_file:
                            updater.apply_update(update_file)
                        else:
                            progress.destroy()
                            messagebox.showerror(
                                "Error",
                                "No se pudo descargar la actualización."
                            )
                    except Exception as e:
                        progress.destroy()
                        messagebox.showerror(
                            "Error",
                            f"Error durante la actualización: {str(e)}"
                        )
                
                thread = threading.Thread(target=update_task)
                thread.daemon = True
                thread.start()
    except Exception as e:
        print(f"Error al verificar actualizaciones: {e}")
    finally:
        checking_updates = False

# Configure CustomTkinter appearance
ctk.set_appearance_mode("dark")  # Modes: "System", "Dark", "Light"
ctk.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue"

class SimplePivotTable:
    def __init__(self, parent, data, title="Análisis de Variación"):
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.state('zoomed')
        
        # Pre-procesar los datos
        self.data = data.copy()
        
        # Configurar estilo
        self.style = ttk.Style()
        self.configure_styles()
        
        # Frame principal con padding
        main_frame = ttk.Frame(self.window, style='PivotMain.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # Header mejorado
        self.create_header(main_frame)
        
        # Barra de herramientas
        self.create_toolbar(main_frame)
        
        # Frame para la tabla con borde y sombra
        table_container = ttk.Frame(main_frame, style='TableContainer.TFrame')
        table_container.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Crear tabla mejorada
        self.create_table(table_container)

    def configure_styles(self):
        # Estilo principal
        self.style.configure('PivotMain.TFrame',
            background='white'
        )
        
        # Estilo para el contenedor de la tabla
        self.style.configure('TableContainer.TFrame',
            background='white',
            borderwidth=1,
            relief='solid'
        )
        
        # Estilo para el header
        self.style.configure('Header.TLabel',
            font=('Segoe UI', 20, 'bold'),
            background='white',
            foreground='#1e293b'
        )
        
        # Estilo para la barra de herramientas
        self.style.configure('Toolbar.TFrame',
            background='#f8fafc'
        )
        
        # Estilo para los botones de la barra de herramientas
        self.style.configure('Toolbar.TButton',
            font=('Segoe UI', 10),
            padding=5
        )

    def create_header(self, parent):
        header_frame = ttk.Frame(parent, style='PivotMain.TFrame')
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Título principal
        title_label = ttk.Label(
            header_frame,
            text="Análisis de Variación",
            style='Header.TLabel'
        )
        title_label.pack(side=tk.LEFT)
        
        # Frame para las acciones del header
        actions_frame = ttk.Frame(header_frame, style='PivotMain.TFrame')
        actions_frame.pack(side=tk.RIGHT)
        
        # Botones de acción
        for text, icon in [("Exportar", "📊"), ("Filtrar", "🔍"), ("Actualizar", "🔄")]:
            btn = ttk.Button(
                actions_frame,
                text=f"{icon} {text}",
                style='Toolbar.TButton'
            )
            btn.pack(side=tk.LEFT, padx=5)

    def create_toolbar(self, parent):
        toolbar_frame = ttk.Frame(parent, style='Toolbar.TFrame')
        toolbar_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Campo de búsqueda
        search_var = tk.StringVar()
        search_entry = ttk.Entry(
            toolbar_frame,
            textvariable=search_var,
            font=('Segoe UI', 10)
        )
        search_entry.pack(side=tk.LEFT, padx=5)
        
        # Filtros rápidos
        filter_label = ttk.Label(
            toolbar_frame,
            text="Filtros:",
            background='#f8fafc',
            font=('Segoe UI', 10)
        )
        filter_label.pack(side=tk.LEFT, padx=(10, 5))
        
        for text in ["Todos", "Con variación", "Sin variación"]:
            btn = ttk.Button(
                toolbar_frame,
                text=text,
                style='Toolbar.TButton'
            )
            btn.pack(side=tk.LEFT, padx=2)

    def create_table(self, parent):
        # Frame con scrollbars
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        # Crear Treeview con las columnas en el orden especificado
        self.table = ttk.Treeview(
            table_frame,
            columns=["Categoría", "Subcategoría", "Segmento", "PLU", "Artículo"] + 
                    [str(c) for c in sorted(self.data['Centro'].unique())],
            show='headings',
            style='Pivot.Treeview'
        )
        
        # Configurar estilo del Treeview (solo fuente y altura de fila)
        self.style.configure('Pivot.Treeview',
            font=('Segoe UI', 10),
            rowheight=30
        )
        self.style.configure('Pivot.Treeview.Heading',
            font=('Segoe UI', 10, 'bold')
        )
        
        # Configurar columnas con ancho fijo para mejor alineación
        column_configs = {
            "Categoría": {"width": 150, "anchor": 'w'},
            "Subcategoría": {"width": 150, "anchor": 'w'},
            "Segmento": {"width": 150, "anchor": 'w'},
            "PLU": {"width": 100, "anchor": 'center'},
            "Artículo": {"width": 200, "anchor": 'w'}
        }

        # Configurar encabezados y columnas con filtros y ordenamiento
        for col, config in column_configs.items():
            self.table.heading(
                col,
                text=col,
                command=lambda c=col: self.sort_column(c)
            )
            self.table.column(col, width=config["width"], anchor=config["anchor"])
            
            # Agregar binding para mostrar filtro al hacer click derecho
            self.table.heading(col, command=lambda c=col: self.sort_column(c))
            self.table.bind(f'<Button-3>', lambda e, c=col: self.show_column_filter(c, e))
        
        # Configurar columnas de centros
        for centro in sorted(self.data['Centro'].unique()):
            self.table.heading(str(centro), text=str(centro))
            self.table.column(str(centro), width=60, anchor='center')
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.table.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.table.xview)
        self.table.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        # Layout
        self.table.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        
        # Cargar datos iniciales
        self.load_data()

    def load_data(self):
        # Limpiar tabla existente
        for item in self.table.get_children():
            self.table.delete(item)
        
        try:
            # Procesar datos por PLU incluyendo todas las columnas
            grouped = self.data.groupby([
                'Categoria', 'Subcategoria', 'Segmento', 
                'PLU_SAP', 'Articulo', 'Centro'
            ]).size().unstack(fill_value=0)
            
            # Insertar datos
            for idx, row in enumerate(grouped.iterrows()):
                # Convertir el índice multicolumna a lista de valores
                values = [str(x) for x in row[0]]
                # Agregar los valores de los centros
                values.extend(['1' if x > 0 else '0' for x in row[1]])
                
                # Insertar fila sin tags de colores
                self.table.insert("", tk.END, values=values)
                
        except Exception as e:
            print(f"Error en load_data: {str(e)}")
            messagebox.showerror("Error", f"Error al cargar los datos: {str(e)}")

    def create_column_filters(self):
        """Crear filtros para cada columna."""
        for column in self.table["columns"]:
            values = set()
            for item in self.table.get_children():
                value = self.table.set(item, column)
                values.add(value)
            self.filters[column] = {
                'values': sorted(list(values)),
                'selected': set(values)  # Inicialmente todos seleccionados
            }

    def show_column_filter(self, column, event):
        """Mostrar menú de filtro para una columna."""
        filter_menu = tk.Toplevel(self.window)
        filter_menu.overrideredirect(True)
        filter_menu.transient(self.window)
        
        # Posicionar menú debajo del encabezado
        x = event.x_root
        y = event.y_root
        filter_menu.geometry(f"+{x}+{y}")
        
        # Frame principal del filtro
        main_frame = ttk.Frame(filter_menu, style='Card.TFrame')
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Barra de búsqueda del filtro
        search_var = tk.StringVar()
        search_entry = ttk.Entry(
            main_frame,
            textvariable=search_var,
            font=('Segoe UI', 10)
        )
        search_entry.pack(fill=tk.X, padx=5, pady=5)
        
        # Lista de valores con scrollbar
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        values_list = tk.Listbox(
            list_frame,
            selectmode=tk.MULTIPLE,
            yscrollcommand=scrollbar.set,
            height=10
        )
        values_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=values_list.yview)
        
        # Poblar lista de valores
        for value in self.filters[column]['values']:
            values_list.insert(tk.END, value)
            if value in self.filters[column]['selected']:
                values_list.selection_set(tk.END)
        
        # Botones de acción
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(
            btn_frame,
            text="Aplicar",
            command=lambda: self.apply_filter(column, values_list.curselection(), filter_menu)
        ).pack(side=tk.RIGHT, padx=2)
        
        ttk.Button(
            btn_frame,
            text="Cancelar",
            command=filter_menu.destroy
        ).pack(side=tk.RIGHT, padx=2)

    def apply_filter(self, column, selections, menu):
        """Aplicar filtro seleccionado."""
        values = [self.filters[column]['values'][i] for i in selections]
        self.filters[column]['selected'] = set(values)
        self.filter_table()
        menu.destroy()

    def filter_table(self):
        """Aplicar todos los filtros activos."""
        # Mostrar todas las filas
        for item in self.table.get_children():
            self.table.detach(item)
        
        # Aplicar filtros
        for item in self.table.get_children(""):
            show = True
            for column in self.filters:
                value = self.table.set(item, column)
                if value not in self.filters[column]['selected']:
                    show = False
                    break
            if show:
                self.table.reattach(item, "", "end")

    def sort_column(self, column):
        """Ordenar tabla por columna."""
        # Obtener datos actuales
        data = [
            (self.table.set(child, column), child)
            for child in self.table.get_children('')
        ]
        
        # Ordenar
        data.sort(reverse=self.table.heading(column, "text").startswith("↓"))
        
        # Actualizar encabezado
        for col in self.table["columns"]:
            self.table.heading(col, text=col)  # Reset otros encabezados
        self.table.heading(
            column,
            text=f"{column} {'↑' if data[0][0] <= data[-1][0] else '↓'}"
        )
        
        # Reordenar items
        for i, item in enumerate(data):
            self.table.move(item[1], '', i)

    def search_table(self, *args):
        """Buscar en todas las columnas."""
        search_term = self.search_var.get().lower()
        
        # Mostrar todas las filas primero
        for item in self.table.get_children():
            self.table.detach(item)
        
        # Filtrar por término de búsqueda
        for item in self.table.get_children(""):
            show = False
            for column in self.table["columns"]:
                value = str(self.table.set(item, column)).lower()
                if search_term in value:
                    show = True
                    break
            if show:
                self.table.reattach(item, "", "end")

    def manage_columns(self):
        """Gestionar visibilidad de columnas."""
        columns_window = tk.Toplevel(self.window)
        columns_window.title("Gestionar Columnas")
        columns_window.geometry("300x400")
        
        # Frame principal
        main_frame = ttk.Frame(columns_window, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Lista de columnas con checkboxes
        columns_frame = ttk.Frame(main_frame)
        columns_frame.pack(fill=tk.BOTH, expand=True)
        
        # Variables para checkboxes
        checkboxes = {}
        for column in self.table["columns"]:
            var = tk.BooleanVar(value=not self.table.column(column, 'width') == 0)
            checkboxes[column] = var
            
            ttk.Checkbutton(
                columns_frame,
                text=column,
                variable=var,
                command=lambda c=column, v=var: self.toggle_column(c, v.get())
            ).pack(anchor='w', pady=2)
        
        # Botones
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(
            btn_frame,
            text="Seleccionar Todo",
            command=lambda: self.select_all_columns(checkboxes, True)
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            btn_frame,
            text="Deseleccionar Todo",
            command=lambda: self.select_all_columns(checkboxes, False)
        ).pack(side=tk.LEFT)

    def toggle_column(self, column, show):
        """Mostrar/ocultar columna."""
        width = self.table.column(column, 'width')
        if show:
            self.table.column(column, width=width if width > 0 else 100)
        else:
            self.table.column(column, width=0)

    def select_all_columns(self, checkboxes, select):
        """Seleccionar/deseleccionar todas las columnas."""
        for column, var in checkboxes.items():
            var.set(select)
            self.toggle_column(column, select)

    def preprocess_data(self, df):
        """Pre-procesar datos para optimizar el rendimiento."""
        # Hacer una copia eficiente
        df = df.copy()
        
        # No establecer Centro como índice, solo optimizar tipos
        categorical_columns = ['Centro', 'Categoria', 'PLU_SAP', 'Articulo', 'Subcategoria', 'Segmento']
        for col in categorical_columns:
            if col in df.columns:
                df[col] = df[col].astype(str)
                if df[col].nunique() < len(df) * 0.5:  # Si hay menos de 50% valores únicos
                    df[col] = df[col].astype('category')
        
        return df

    def setup_ui(self):
        """Configurar UI optimizada."""
        # Frame principal usando grid para mejor rendimiento
        main_frame = ttk.Frame(self.window)
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)
        
        # Panel izquierdo (campos)
        left_frame = ttk.Frame(main_frame)
        left_frame.grid(row=0, column=0, sticky="ns", padx=5, pady=5)
        
        # Lista de campos (usar Listbox con altura fija)
        fields_frame = ttk.LabelFrame(left_frame, text="Campos")
        fields_frame.pack(fill=tk.X, pady=5)
        
        self.available_fields = tk.Listbox(
            fields_frame,
            height=6,
            exportselection=False,
            selectmode=tk.EXTENDED
        )
        self.available_fields.pack(fill=tk.X, padx=5, pady=5)
        
        # Cargar campos disponibles
        self.load_available_fields()
        
        # Areas de pivot
        self.create_pivot_areas(left_frame)
        
        # Panel derecho (tabla)
        table_frame = ttk.Frame(main_frame)
        table_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)
        main_frame.grid_columnconfigure(1, weight=1)
        
        # Crear tabla optimizada
        self.create_table(table_frame)

    def load_available_fields(self):
        """Cargar campos disponibles en el listbox con optimizaciones."""
        try:
            # Limpiar lista actual
            self.available_fields.delete(0, tk.END)
            
            # Definir orden preferido de campos
            preferred_fields = [
                'Categoria', 'Subcategoria', 'Segmento',
                'PLU_SAP', 'Articulo', 'Centro'
            ]
            
            # Obtener todas las columnas del DataFrame
            all_columns = set(self.data.columns)
            
            # Primero insertar campos preferidos si están disponibles
            for field in preferred_fields:
                if field in all_columns:
                    self.available_fields.insert(tk.END, field)
                    all_columns.remove(field)
            
            # Luego insertar el resto de campos ordenados alfabéticamente
            remaining_fields = sorted(list(all_columns))
            for field in remaining_fields:
                self.available_fields.insert(tk.END, field)
                
        except Exception as e:
            print(f"Error al cargar campos: {str(e)}")
            # Intentar cargar campos de manera básica como fallback
            for col in sorted(self.data.columns):
                self.available_fields.insert(tk.END, col)

    def load_default_config(self):
        """Cargar configuración predeterminada optimizada."""
        # Limpiar áreas
        for area in ['filas', 'columnas', 'valores']:
            if hasattr(self, f'{area}_list'):
                getattr(self, f'{area}_list').delete(0, tk.END)
        
        # Configuración básica
        default_config = {
            'columnas': ['Centro'],
            'filas': ['Categoria', 'PLU_SAP', 'Articulo'],
            'valores': ['Centro']
        }
        
        # Aplicar configuración verificando disponibilidad
        for area, fields in default_config.items():
            listbox = getattr(self, f'{area}_list')
            for field in fields:
                if field in self.data.columns:
                    listbox.insert(tk.END, field)
        
        # Actualizar tabla
        self.update_pivot_table()

    def update_pivot_table(self):
        """Actualizar tabla pivote con rendimiento optimizado."""
        try:
            # Obtener configuración
            rows = list(self.filas_list.get(0, tk.END))
            cols = list(self.columnas_list.get(0, tk.END))
            
            if not (rows or cols):
                return
            
            # Limpiar tabla
            self.table.delete(*self.table.get_children())
            
            # Crear agregación eficiente
            group_cols = rows + cols if cols else rows
            if not group_cols:
                return
            
            # Reset index si existe
            if self.data.index.name:
                self.data = self.data.reset_index()
            
            # Usar groupby para mejor rendimiento
            grouped = self.data.groupby(group_cols).size().reset_index(name='count')
            
            # Si hay columnas de pivote
            if cols:
                # Crear pivot eficiente
                pivot = pd.pivot_table(
                    grouped,
                    index=rows,
                    columns=cols,
                    values='count',
                    fill_value=0,
                    aggfunc='sum'
                )
                
                # Configurar columnas
                self.table["columns"] = list(pivot.columns)
                for col in pivot.columns:
                    self.table.heading(col, text=str(col))
                    self.table.column(col, width=100)
                
                # Insertar datos en lotes
                batch_size = 1000
                data_rows = []
                
                for idx in pivot.index:
                    if isinstance(idx, tuple):
                        values = list(idx) + list(pivot.loc[idx])
                    else:
                        values = [idx] + list(pivot.loc[idx])
                    data_rows.append(values)
                    
                    if len(data_rows) >= batch_size:
                        for row in data_rows:
                            self.table.insert("", tk.END, values=row)
                        data_rows = []
                        self.window.update_idletasks()
                
                # Insertar filas restantes
                for row in data_rows:
                    self.table.insert("", tk.END, values=row)
            else:
                # Mostrar resultados agrupados
                self.table["columns"] = group_cols + ["Cantidad"]
                for col in self.table["columns"]:
                    self.table.heading(col, text=str(col))
                    self.table.column(col, width=100)
                
                for _, row in grouped.iterrows():
                    self.table.insert("", tk.END, values=list(row))
            
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def create_pivot_areas(self, parent):
        """Crear áreas de pivot optimizadas."""
        # Crear frames para cada área
        areas = ['Filas', 'Columnas', 'Valores']
        
        for area in areas:
            frame = ttk.LabelFrame(parent, text=area)
            frame.pack(fill=tk.X, pady=2)
            
            listbox = tk.Listbox(
                frame,
                height=4,
                selectmode=tk.EXTENDED,
                exportselection=False
            )
            listbox.pack(fill=tk.X, padx=5, pady=5)
            
            # Guardar referencia
            setattr(self, f'{area.lower()}_list', listbox)
            
            # Configurar drag & drop
            listbox.bind('<ButtonPress-1>', self.on_drag_start)
            listbox.bind('<B1-Motion>', self.on_drag_motion)
            listbox.bind('<ButtonRelease-1>', self.on_drag_release)

    def on_drag_start(self, event):
        """Inicio de drag & drop optimizado."""
        widget = event.widget
        if widget.curselection():
            self.drag_data = {
                'widget': widget,
                'index': widget.curselection()[0],
                'value': widget.get(widget.curselection()[0])
            }
            widget.configure(cursor="hand2")

    def on_drag_motion(self, event):
        """Movimiento de drag & drop optimizado."""
        if hasattr(self, 'drag_data'):
            event.widget.configure(cursor="hand2")

    def on_drag_release(self, event):
        """Finalización de drag & drop optimizada."""
        if hasattr(self, 'drag_data'):
            target = event.widget
            if target != self.drag_data['widget']:
                value = self.drag_data['value']
                target.insert(tk.END, value)
                self.drag_data['widget'].delete(self.drag_data['index'])
                self.update_pivot_table()
            
            # Limpiar
            if hasattr(self, 'drag_data'):
                del self.drag_data
            event.widget.configure(cursor="")

    def get_pivot_data(self):
        """
        Retorna el DataFrame con los datos procesados de la tabla pivot.
        
        Returns:
            pandas.DataFrame: Datos procesados de la tabla pivot
        """
        try:
            # Obtener datos de la tabla
            data = []
            columns = [col["text"] for col in self.table["columns"]]
            
            # Recolectar datos de cada fila
            for item in self.table.get_children():
                row_data = []
                for col in range(len(columns)):
                    value = self.table.set(item, col)
                    row_data.append(value)
                data.append(row_data)
            
            # Crear DataFrame
            return pd.DataFrame(data, columns=columns)
            
        except Exception as e:
            print(f"Error al obtener datos de la tabla: {str(e)}")
            return pd.DataFrame()  # Retornar DataFrame vacío en caso de error

class ColumnValidator:
    def __init__(self):
        self.COLUMN_VARIATIONS = {
            'CENTRO': {
                'valid': [
                    'CENTRO', 'COD_CENTRO', 'CODIGO_CENTRO', 'CENTRO_ID',
                    'ID_CENTRO', 'NUM_CENTRO', 'CENTRO_CODIGO'
                ],
                'excluded': [
                    'NOMBRE_CENTRO', 'NOMBRE CENTRO', 'DESC_CENTRO',
                    'DESCRIPCION_CENTRO', 'CENTRO_NOMBRE'
                ],
                'required_type': None,
                'required': True
            },
            'PLU_SAP': {
                'valid': [
                    'PLU_SAP', 'PLU', 'CODIGO_PLU', 'COD_PLU', 'SKU',
                    'CODIGO_PRODUCTO', 'COD_PRODUCTO'
                ],
                'excluded': [
                    'DESCRIPCION_PLU', 'NOMBRE_PLU', 'PLU_DESCRIPCION'
                ],
                'required_type': None,
                'required': True
            },
            'CATEGORIA': {
                'valid': ['CATEGORIA', 'CATEGORIAS'],
                'excluded': [],
                'required_type': None,
                'required': False
            },
            'SUBCATEGORIA': {
                'valid': ['SUBCATEGORIA', 'SUBCATEGORIAS'],
                'excluded': [],
                'required_type': None,
                'required': False
            },
            'SEGMENTO': {
                'valid': ['SEGMENTO', 'SEGMENTOS'],
                'excluded': [],
                'required_type': None,
                'required': False
            },
            'ARTICULO': {
                'valid': ['ARTICULO', 'ARTICULOS'],
                'excluded': [],
                'required_type': None,
                'required': False
            },
            'DISTRITO': {
                'valid': ['Distrito'],  # Solo la columna exacta
                'excluded': [],
                'required_type': None,
                'required': False,
                'exact_match': True    # Requerir coincidencia exacta
            },
            'REGION': {
                'valid': ['Región'],   # Solo la columna exacta
                'excluded': [],
                'required_type': None,
                'required': False,
                'exact_match': True    # Requerir coincidencia exacta
            }
        }

    def normalize_column_name(self, column_name):
        """
        Normaliza el nombre de una columna para comparaciones consistentes.
        """
        return str(column_name).upper().replace(' ', '_').strip()

    def find_column(self, df, column_type, raise_error=True):
        """
        Encuentra una columna específica en el DataFrame.

        Args:
            df: DataFrame donde buscar
            column_type: Tipo de columna a buscar ('CENTRO', 'PLU_SAP', etc.)
            raise_error: Si es True, lanza error cuando no encuentra la columna

        Returns:
            str: Nombre de la columna encontrada o None si no se encuentra
        """

        config = self.COLUMN_VARIATIONS[column_type]
        exact_match = config.get('exact_match', False)

        if exact_match:
            # Buscar coincidencia exacta
            for valid_name in config['valid']:
                if valid_name in df.columns:
                    return valid_name
        else:
            if column_type not in self.COLUMN_VARIATIONS:
                raise ValueError(f"Tipo de columna no soportado: {column_type}")

            valid_variations = self.COLUMN_VARIATIONS[column_type]['valid']
            excluded_variations = self.COLUMN_VARIATIONS[column_type]['excluded']
            required_type = self.COLUMN_VARIATIONS[column_type]['required_type']

            # Normalizar nombres de columnas del DataFrame
            df_columns = {col: self.normalize_column_name(col) for col in df.columns}

            # Excluir columnas no deseadas
            excluded_cols = []
            for col, normalized_name in df_columns.items():
                for excluded in excluded_variations:
                    if self.normalize_column_name(excluded) in normalized_name:
                        excluded_cols.append(col)

            # Buscar coincidencias exactas primero
            exact_matches = []
            for col, normalized_name in df_columns.items():
                if col not in excluded_cols:
                    for valid_var in valid_variations:
                        if normalized_name == self.normalize_column_name(valid_var):
                            exact_matches.append(col)

            if len(exact_matches) == 1:
                found_column = exact_matches[0]
                if self._validate_column_type(df, found_column, required_type):
                    return found_column

            # Si no hay coincidencias exactas, buscar coincidencias parciales
            partial_matches = []
            for col, normalized_name in df_columns.items():
                if col not in excluded_cols:
                    if column_type in normalized_name:
                        if self._validate_column_type(df, col, required_type):
                            partial_matches.append(col)

            if len(partial_matches) == 1:
                return partial_matches[0]
            elif len(partial_matches) > 1:
                if raise_error:
                    raise ValueError(
                        f"Se encontraron múltiples columnas que podrían ser '{column_type}': "
                        f"{', '.join(partial_matches)}.\nPor favor, verifique el archivo "
                        f"y asegúrese de que la columna '{column_type}' esté claramente identificada."
                    )
            elif raise_error:
                raise ValueError(
                    f"No se encontró la columna '{column_type}' en el archivo.\n"
                    f"Columnas disponibles: {', '.join(df.columns)}"
                )

            return None

    def _validate_column_type(self, df, column, required_type):
        """
        Valida el tipo de datos de una columna.
        """
        if required_type == 'numeric':
            return df[column].astype(str).str.match(r'^\d+$').all()
        return True

    def validate_required_columns(self, df, required_columns):
        """
        Valida múltiples columnas requeridas en el DataFrame.

        Args:
            df: DataFrame a validar
            required_columns: Lista de tipos de columnas requeridas

        Returns:
            dict: Diccionario con los nombres de las columnas encontradas
        """
        found_columns = {}
        for col_type in required_columns:
            found_columns[col_type] = self.find_column(df, col_type)
        return found_columns

class WidgetRecycler:
    def __init__(self):
        self.widget_pool = {}
    
    def get_widget(self, widget_type, parent):
        if widget_type in self.widget_pool and self.widget_pool[widget_type]:
            widget = self.widget_pool[widget_type].pop()
            widget.pack_forget()  # Desacoplar del padre anterior
            return widget
        return widget_type(parent)
    
    def recycle_widget(self, widget, widget_type):
        if widget_type not in self.widget_pool:
            self.widget_pool[widget_type] = []
        self.widget_pool[widget_type].append(widget)

class CustomGroupingDialog:
    def __init__(self, parent):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Criterios de Agrupación")
        self.dialog.grab_set()
        
        # Centrar el diálogo
        window_width = 600
        window_height = 400
        screen_width = self.dialog.winfo_screenwidth()
        screen_height = self.dialog.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Frame principal con padding y fondo blanco
        main_frame = ctk.CTkFrame(
            self.dialog,
            fg_color="white",
            corner_radius=10
        )
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Título
        ttk.Label(
            main_frame,
            text="Seleccione los criterios de agrupación",
            font=('Segoe UI', 16, 'bold'),
            foreground='#1e293b',
            background='white'
        ).pack(pady=(0, 20))
        
        # Frame para los niveles de agrupación
        self.levels_frame = ttk.Frame(main_frame, style="Card.TFrame")
        self.levels_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Lista de niveles de agrupación
        self.grouping_levels = []
        
        # Botones de control de niveles
        control_frame = ttk.Frame(main_frame, style="Card.TFrame")
        control_frame.pack(fill=tk.X, pady=10)
        
        ctk.CTkButton(
            control_frame,
            text="Agregar nivel",
            command=self.add_level,
            width=120
        ).pack(side=tk.LEFT, padx=5)
        
        ctk.CTkButton(
            control_frame,
            text="Eliminar nivel",
            command=self.remove_level,
            width=120,
            fg_color="#ef4444",
            hover_color="#dc2626"
        ).pack(side=tk.LEFT, padx=5)
        
        # Frame para botones de acción
        button_frame = ttk.Frame(main_frame, style="Card.TFrame")
        button_frame.pack(fill=tk.X, pady=(20, 0))
        
        ctk.CTkButton(
            button_frame,
            text="Cancelar",
            command=self.dialog.destroy,
            width=100,
            fg_color="#f1f5f9",
            hover_color="#e2e8f0",
            text_color="#64748b"
        ).pack(side=tk.RIGHT, padx=5)
        
        ctk.CTkButton(
            button_frame,
            text="Aceptar",
            command=self.accept,
            width=100
        ).pack(side=tk.RIGHT, padx=5)
        
        # Agregar primer nivel por defecto
        self.add_level()
        
        # Variable para almacenar el resultado
        self.result = None
        
    def add_level(self):
        """Agrega un nuevo nivel de agrupación."""
        level_frame = ttk.Frame(self.levels_frame, style="Card.TFrame")
        level_frame.pack(fill=tk.X, pady=2)
        
        # Criterios disponibles
        criteria = [
            "Distrito",
            "Región",
            "Departamento",
            "Ciudad",
            "Clúster",
            "Modulación"
        ]
        
        # Combobox para selección de criterio
        criteria_var = tk.StringVar()
        combo = ttk.Combobox(
            level_frame,
            values=criteria,
            textvariable=criteria_var,
            state="readonly",
            width=30
        )
        combo.pack(side=tk.LEFT, padx=5)
        
        # Guardar referencia al nivel
        self.grouping_levels.append({
            'frame': level_frame,
            'combo': combo,
            'variable': criteria_var
        })
        
        # Seleccionar primer criterio por defecto
        if criteria:
            combo.set(criteria[0])
    
    def remove_level(self):
        """Elimina el último nivel de agrupación."""
        if len(self.grouping_levels) > 1:  # Mantener al menos un nivel
            level = self.grouping_levels.pop()
            level['frame'].destroy()
    
    def accept(self):
        """Procesa la selección y cierra el diálogo."""
        # Recolectar criterios seleccionados
        self.result = [
            level['variable'].get()
            for level in self.grouping_levels
        ]
        self.dialog.destroy()
    
    def show(self):
        """Muestra el diálogo y retorna los criterios seleccionados."""
        self.dialog.wait_window()
        return self.result

class CustomGroupingAnalysis(BaseApp):
    def __init__(self, parent):
        self.parent = parent

    def load_master_data(self):
        """Carga datos maestros desde archivos externos."""
        try:
            # Buscar archivo db_maestrospdv en la misma carpeta
            db_file = self.get_data_path("db_maestrospdv.xlsx")
            
            if not os.path.exists(db_file):
                raise FileNotFoundError("No se encontró el archivo db_maestrospdv.xlsx")
            
            # Leer datos maestros
            self.master_data = pd.read_excel(db_file)
            
            # Normalizar nombres de columnas
            column_mapping = {
                'Centro': 'Centro',
                'Distrito': 'Distrito',
                'Región': 'Region',
                'Departamento': 'Departamento',
                'Ciudad': 'Ciudad',
                'Formato Procura': 'Formato_Procura'
            }
            
            self.master_data = self.master_data.rename(columns=column_mapping)
            
            # Convertir columnas a string y limpiar
            for col in self.master_data.columns:
                if col != 'Centro':  # Mantener Centro como está
                    self.master_data[col] = self.master_data[col].astype(str).str.strip()
            
            return True
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar datos maestros: {str(e)}")
            return False
    
    def load_modulation_data(self, category):
        """Carga datos de modulación para una categoría específica."""
        try:
            maestro_file = self.get_data_path("Form.Maestro.Neg.xlsx")
            
            if not os.path.exists(maestro_file):
                raise FileNotFoundError("No se encontró el archivo Form.Maestro.Neg.xlsx")
            
            # Leer la hoja de MOBILIARIOS
            df = pd.read_excel(maestro_file, sheet_name='MOBILIARIOS')
            
            # Normalizar la categoría y convertir columnas a string
            category = category.upper().strip()
            df['CENTRO'] = df['CENTRO'].astype(str).str.strip()
            df['CATEGORIA'] = df['CATEGORIA'].astype(str).str.upper().str.strip()
            
            # Filtrar por categoría
            df_filtered = df[df['CATEGORIA'] == category]
            
            # Crear diccionario de modulación por centro
            modulation_data = {}
            for _, row in df_filtered.iterrows():
                centro = str(row['CENTRO']).strip()
                modulos = str(row['NÚMERO DE MÓDULOS'])
                
                if pd.isna(modulos) or modulos == '':
                    modulos = 'Sin datos'
                    
                modulation_data[centro] = modulos
            
            return modulation_data
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar datos de modulación: {str(e)}")
            return None
    
    def perform_custom_grouping(self, centers, criteria):
        """Realiza la agrupación personalizada según los criterios seleccionados."""
        try:
            # Verificar si se necesitan datos de modulación
            needs_modulation = 'Modulación' in criteria
            modulation_data = None
            
            if needs_modulation:
                # Pedir categoría al usuario
                category = self.parent.get_category_input()
                if not category:
                    return None
                    
                # Cargar datos de modulación
                modulation_data = self.load_modulation_data(category)
                if not modulation_data:
                    return None
            
            # Cargar datos maestros si no están cargados
            if not hasattr(self, 'master_data'):
                if not self.load_master_data():
                    return None
            
            # Filtrar master_data para los centros proporcionados
            centers_str = [str(c).strip() for c in centers]
            filtered_data = self.master_data[
                self.master_data['Centro'].astype(str).str.strip().isin(centers_str)
            ].copy()
            
            # Agregar datos de modulación si es necesario
            if needs_modulation:
                filtered_data['Modulacion'] = filtered_data['Centro'].map(
                    lambda x: modulation_data.get(str(x).strip(), 'No encontrado')
                )
            
            # Mapear criterios a nombres de columnas
            criteria_mapping = {
                'Distrito': 'Distrito',
                'Región': 'Region',
                'Departamento': 'Departamento',
                'Ciudad': 'Ciudad',
                'Clúster': 'Formato_Procura',
                'Modulación': 'Modulacion'
            }
            
            # Convertir criterios a nombres de columnas
            grouping_columns = [criteria_mapping[c] for c in criteria]
            
            # Realizar la agrupación
            groups = []
            for _, group in filtered_data.groupby(grouping_columns):
                group_centers = group['Centro'].tolist()
                if len(group_centers) > 0:
                    groups.append({
                        'centers': group_centers,
                        'criteria_values': {
                            criteria[i]: group[criteria_mapping[criteria[i]]].iloc[0]
                            for i in range(len(criteria))
                        }
                    })
            
            return groups
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al realizar la agrupación: {str(e)}")
            return None

class CustomGroupResults:
    def __init__(self, parent, groups, criteria, app_reference):
        self.parent = parent
        self.groups = groups
        self.criteria = criteria
        self.app = app_reference
        self.app.custom_groups = groups 

        # Frame principal para resultados con scroll
        self.main_frame = ttk.Frame(parent, style="Card.TFrame")
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Crear canvas con scrollbar
        self.canvas = tk.Canvas(self.main_frame, bg='white', highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.main_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas, style="Card.TFrame")

        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Pack scrollbar and canvas
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Create window in canvas
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # Bind scroll events
        self.scrollable_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)
        self.bind_mousewheel()

        # Mostrar el contenido
        self.display_content()

    def bind_mousewheel(self):
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def on_frame_configure(self, event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_canvas_configure(self, event):
        # Actualizar el ancho de la ventana del canvas
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def display_content(self):
        # Mostrar estadísticas
        total_centers = sum(len(group['centers']) for group in self.groups)
        total_groups = len(self.groups)
        avg_centers = total_centers / total_groups if total_groups > 0 else 0

        # Frame para estadísticas
        stats_frame = ctk.CTkFrame(
            self.scrollable_frame,
            fg_color="white",
            corner_radius=10
        )
        stats_frame.pack(fill=tk.X, pady=(0, 20))

        # Crear las tres métricas
        self.create_stats_cards(stats_frame, total_centers, total_groups, avg_centers)

        # Separador
        separator = ttk.Frame(self.scrollable_frame, height=2, style="BlackSeparator.TFrame")
        separator.pack(fill=tk.X, pady=(0, 15))

        # Mostrar grupos
        for i, group in enumerate(self.groups, 1):
            self.create_group_frame(i, group)

    def create_stats_cards(self, parent, total_centers, total_groups, avg_centers):
        stats_container = ttk.Frame(parent, style="Card.TFrame")
        stats_container.pack(fill=tk.X, padx=20, pady=10)

        # Configurar tres columnas iguales
        for i in range(3):
            stats_container.columnconfigure(i, weight=1)

        # Crear cada tarjeta de estadística
        stats = [
            ("Total de Centros", total_centers, "#2563eb"),
            ("Grupos Generados", total_groups, "#16a34a"),
            ("Promedio de Centros por Grupo", f"{avg_centers:.1f}", "#7c3aed")
        ]

        for i, (title, value, color) in enumerate(stats):
            self.create_stat_card(stats_container, title, value, color, i)

    def create_stat_card(self, parent, title, value, color, column):
        frame = ctk.CTkFrame(
            parent,
            fg_color="white",
            corner_radius=8
        )
        frame.grid(row=0, column=column, padx=5, sticky='ew')

        ctk.CTkLabel(
            frame,
            text=str(value),
            font=("Segoe UI", 24, "bold"),
            text_color=color
        ).pack(pady=(10, 5))

        ctk.CTkLabel(
            frame,
            text=title,
            font=("Segoe UI", 12),
            text_color="#64748b"
        ).pack(pady=(0, 10))
       
    def create_data_table_with_chart(self, parent, title, group_data):
        """Crear tabla de datos con gráfico usando el nuevo diseño."""
        # Crear contenedor principal
        container = ctk.CTkFrame(
            parent,
            fg_color="white",
            corner_radius=10,
            border_width=1,
            border_color="#E5E7EB"
        )
        container.pack(fill=tk.X, padx=5, pady=5)
        
        # Header frame
        header_frame = ctk.CTkFrame(
            container,
            fg_color="#F9FAFB",
            corner_radius=0,
            height=50
        )
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)

        # Frame izquierdo para título
        title_frame = ctk.CTkFrame(
            header_frame,
            fg_color="transparent"
        )
        title_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Título
        ctk.CTkLabel(
            title_frame,
            text=title,
            font=("Segoe UI", 13, "bold"),
            text_color="#111827"
        ).pack(side=tk.LEFT, padx=15, pady=15)

        # Botón de análisis
        analyze_button = ctk.CTkButton(
            header_frame,
            text="Ver Análisis de Variación",
            command=lambda centers=group_data['centers']: self.app.show_portfolio_variation('Personalizado', centers),
            width=180,
            height=32,
            fg_color="#2563eb",
            hover_color="#1d4ed8"
        )
        analyze_button.pack(side=tk.RIGHT, padx=15)
        
        # Frame para el contenido principal
        content_frame = ctk.CTkFrame(
            container,
            fg_color="white",
            corner_radius=0
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        return container   

    def create_stats_frame(self):
        """Crea el frame de estadísticas."""
        frame = ttk.Frame(self.upper_frame, style="Card.TFrame")
        
        # Calcular estadísticas
        total_centers = sum(len(g['centers']) for g in self.groups)
        total_groups = len(self.groups)
        avg_centers = total_centers / total_groups if total_groups > 0 else 0
        
        # Crear tarjetas de estadísticas
        stats = [
            ("Total de Centros", total_centers, "#2563eb"),
            ("Grupos Generados", total_groups, "#16a34a"),
            ("Promedio de Centros por Grupo", f"{avg_centers:.1f}", "#8b5cf6")
        ]
        
        for i, (title, value, color) in enumerate(stats):
            stat_card = ctk.CTkFrame(
                frame,
                fg_color="white",
                corner_radius=8,
                border_width=1,
                border_color="#e2e8f0"
            )
            stat_card.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=5)
            
            # Valor
            ttk.Label(
                stat_card,
                text=str(value),
                font=('Segoe UI', 24, 'bold'),
                foreground=color,
                background='white'
            ).pack(pady=(10, 5))
            
            # Título
            ttk.Label(
                stat_card,
                text=title,
                font=('Segoe UI', 12),
                foreground='#64748b',
                background='white'
            ).pack(pady=(0, 10))
            
        return frame
    
    def bind_scroll_events(self):
        """Configura los eventos de scroll para un funcionamiento más suave."""
        def on_mousewheel(event):
            if event.delta:
                # Para Windows
                self.canvas.yview_scroll(int(-1 * (event.delta/120)), "units")
            else:
                # Para Linux
                if event.num == 4:
                    self.canvas.yview_scroll(-1, "units")
                elif event.num == 5:
                    self.canvas.yview_scroll(1, "units")

        # Binding para Windows
        self.canvas.bind_all("<MouseWheel>", on_mousewheel)
        
        # Binding para Linux
        self.canvas.bind_all("<Button-4>", on_mousewheel)
        self.canvas.bind_all("<Button-5>", on_mousewheel)
        
        def unbind_mouse_wheel():
            self.canvas.unbind_all("<MouseWheel>")
            self.canvas.unbind_all("<Button-4>")
            self.canvas.unbind_all("<Button-5>")
        
        def rebind_mouse_wheel(event):
            self.canvas.bind_all("<MouseWheel>", on_mousewheel)
            self.canvas.bind_all("<Button-4>", on_mousewheel)
            self.canvas.bind_all("<Button-5>", on_mousewheel)
        
        # Desactivar/reactivar scroll cuando el mouse sale/entra del área
        self.canvas.bind('<Enter>', rebind_mouse_wheel)
        self.canvas.bind('<Leave>', lambda e: unbind_mouse_wheel())

    def create_group_name(self, group):
        """
        Crea el nombre del grupo basado en los criterios de agrupación y sus valores.
        
        Args:
            group (dict): Diccionario con los datos del grupo incluyendo criteria_values
            
        Returns:
            tuple: (nombre_grupo, nombre_completo) donde nombre_grupo es la versión corta y 
                nombre_completo incluye el número de centros
        """
        if 'criteria_values' not in group:
            return "Grupo sin criterios", "Grupo sin criterios"

        # Función para normalizar texto (eliminar tildes y espacios extras)
        def normalize_text(text):
            """Elimina tildes y normaliza espacios y símbolos"""
            text = str(text).strip()
            # Reemplazar variaciones de separadores
            text = text.replace(' / ', ' ').replace('/', ' ')
            # Eliminar tildes
            text = ''.join(c for c in unicodedata.normalize('NFD', text)
                        if unicodedata.category(c) != 'Mn')
            # Normalizar espacios
            text = ' '.join(text.split())
            return text

        # Procesar cada criterio y su valor
        name_parts = []
        for criterion, value in group['criteria_values'].items():
            # Normalizar el valor
            value = normalize_text(value)
            
            if criterion == "Clúster":
                # Para Clúster, si el valor es "-", dejar un espacio
                if value != "-":
                    name_parts.append(str(value))
            elif criterion == "Modulación":
                # Para Modulación, agregar "M" al final
                if value not in ["Sin datos", "No encontrado"]:
                    name_parts.append(f"{value}M")
            else:
                name_parts.append(str(value))

        # Unir con guiones evitando guiones duplicados o innecesarios
        group_name = ""
        for i, part in enumerate(name_parts):
            if i > 0:
                # Si la parte anterior o actual es solo un guión, no agregar guión adicional
                if part != "-" and name_parts[i-1] != "-":
                    group_name += "-"
            group_name += part

        # Crear versión corta (para el título) y versión larga (con número de centros)
        if len(group_name) > 50:  # Si el nombre es muy largo
            short_name = group_name[:47] + "..."
        else:
            short_name = group_name

        full_name = f"{short_name} - {len(group['centers'])} centros"

        return short_name, full_name

    def create_group_frame(self, group_num, group):
        # Frame principal
        group_frame = ctk.CTkFrame(
            self.scrollable_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        group_frame.pack(fill=tk.X, pady=2, padx=5)

        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)

        # Header frame
        header_frame = ttk.Frame(group_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))

        # Canvas para el ícono de expandir/colapsar
        icon_size = 24
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg='white',
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(10, 5), pady=10)

        def draw_icon(expanded):
            icon_canvas.delete("all")
            icon_canvas.create_oval(
                0, 0, icon_size, icon_size,
                fill="#2563eb",
                outline=""
            )
            if expanded:
                icon_canvas.create_line(
                    8, 10, 12, 14, 16, 10,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
            else:
                icon_canvas.create_line(
                    10, 8, 14, 12, 10, 16,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )

        # Frame para el título y contenido
        content_container = ttk.Frame(header_frame, style="Card.TFrame")
        content_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))

        # Obtener el nombre del grupo
        short_name, full_name = self.create_group_name(group)

        # Título
        title_label = ttk.Label(
            content_container,
            text=full_name,
            font=('Segoe UI', 14, 'bold'),
            foreground='#2563eb',
            background='white',
            wraplength=600
        )
        title_label.pack(anchor='w')

        # Botón de análisis
        analyze_button = ctk.CTkButton(
            header_frame,
            text="Ver Análisis",
            command=lambda: self.app.show_portfolio_variation('Personalizado', group['centers']),
            width=100,
            height=32,
            fg_color="#2563eb",
            hover_color="#1d4ed8"
        )
        analyze_button.pack(side=tk.RIGHT, padx=15)

        # Frame para el contenido expandible
        detail_frame = ttk.Frame(group_frame, style="ContentCard.TFrame")
        
        def get_group_plus_count():
            all_centers_plus = []
            for center in group['centers']:
                if center in self.app.unique_portfolios:
                    # Si el centro tiene portafolio único
                    all_centers_plus.append(len(self.app.unique_portfolios[center]))
                else:
                    # Si el centro está en un grupo idéntico, buscar su grupo
                    for centers, plus in self.app.identical_portfolios.items():
                        if center in centers:
                            all_centers_plus.append(len(plus))
                            break
            return sum(all_centers_plus)

        total_plus = get_group_plus_count()

        # Contenido adicional cuando se expande
        ttk.Label(
            detail_frame,
            text=f"Total de PLUs: {total_plus:,}",  # Agregado formato de número con comas
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        ).pack(anchor='w', padx=15, pady=(10, 5))

        ttk.Label(
            detail_frame,
            text=f"Centros: {', '.join(sorted(map(str, group['centers'])))}",
            wraplength=800,
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        ).pack(anchor='w', padx=15, pady=(0, 10))

        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                detail_frame.pack(fill=tk.X)
            else:
                detail_frame.pack_forget()
            draw_icon(is_expanded.get())

        # Hacer que el header y el ícono sean clickeables
        header_frame.bind("<Button-1>", lambda e: toggle_content())
        icon_canvas.bind("<Button-1>", lambda e: toggle_content())

        # Efectos hover
        def on_enter(e):
            icon_canvas.configure(cursor="hand2")
            header_frame.configure(cursor="hand2")
            
        def on_leave(e):
            icon_canvas.configure(cursor="")
            header_frame.configure(cursor="")

        header_frame.bind("<Enter>", on_enter)
        header_frame.bind("<Leave>", on_leave)
        icon_canvas.bind("<Enter>", on_enter)
        icon_canvas.bind("<Leave>", on_leave)

        # Dibujar ícono inicial
        draw_icon(False)

        return group_frame

    def create_group_card(self, parent, group_num, group):
        """Crea una tarjeta expandible para un grupo."""
        # Frame principal para la tarjeta
        group_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Crear el frame CustomTkinter para el contenido
        content_frame = ctk.CTkFrame(
            group_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Frame para el header
        header_frame = ttk.Frame(content_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))
        
        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)
        
        # Canvas para el ícono
        icon_size = 24
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg='white',
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(10, 5), pady=10)
        
        def draw_icon(expanded):
            icon_canvas.delete("all")
            icon_canvas.create_oval(
                0, 0, icon_size, icon_size,
                fill="#2563eb",
                outline=""
            )
            if expanded:
                icon_canvas.create_line(
                    8, 10, 12, 14, 16, 10,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
            else:
                icon_canvas.create_line(
                    10, 8, 14, 12, 10, 16,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
        
        # Contenedor para el texto y botón
        text_container = ttk.Frame(header_frame, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))
        
        # Título
        title_label = ttk.Label(
            text_container,
            text=f"Grupo {group_num} - {len(group['centers'])} centros",
            font=('Segoe UI', 14, 'bold'),
            foreground='#2563eb',
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Botón de análisis
        analyze_button = ctk.CTkButton(
            header_frame,
            text="Ver Análisis",
            command=lambda: self.app.show_portfolio_variation('Personalizado', group['centers']),
            width=100,
            height=32,
            fg_color="#2563eb",
            hover_color="#1d4ed8"
        )
        analyze_button.pack(side=tk.RIGHT, padx=15)
        
        # Frame para el contenido expandible
        content_detail_frame = ttk.Frame(content_frame, style="ContentCard.TFrame")
        
        # Mostrar criterios si existen
        if 'criteria_values' in group:
            for criterion, value in group['criteria_values'].items():
                ttk.Label(
                    content_detail_frame,
                    text=f"{criterion}: {value}",
                    font=('Segoe UI', 11),
                    foreground='#4b5563',
                    background='#f9fafb'
                ).pack(anchor='w', padx=15, pady=2)
        
        # Mostrar centros
        ttk.Label(
            content_detail_frame,
            text=f"Centros: {', '.join(sorted(map(str, group['centers'])))}",
            wraplength=800,
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        ).pack(anchor='w', padx=15, pady=(5, 10))
        
        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                content_detail_frame.pack(fill=tk.X)
            else:
                content_detail_frame.pack_forget()
            draw_icon(is_expanded.get())
        
        # Hacer que todo el header sea clickeable
        header_frame.bind("<Button-1>", lambda e: toggle_content())
        icon_canvas.bind("<Button-1>", lambda e: toggle_content())
        
        # Efectos hover
        def on_enter(e):
            icon_canvas.configure(cursor="hand2")
            header_frame.configure(cursor="hand2")
            
        def on_leave(e):
            icon_canvas.configure(cursor="")
            header_frame.configure(cursor="")
        
        header_frame.bind("<Enter>", on_enter)
        header_frame.bind("<Leave>", on_leave)
        icon_canvas.bind("<Enter>", on_enter)
        icon_canvas.bind("<Leave>", on_leave)
        
        # Dibujar ícono inicial
        draw_icon(False)
        
        return group_frame

    def display_groups(self):
        """Muestra los grupos generados."""
        # Mostrar estadísticas primero
        total_centers = sum(len(group['centers']) for group in self.groups)
        total_groups = len(self.groups)
        avg_centers = total_centers / total_groups if total_groups > 0 else 0
        
        # Crear tarjetas de estadísticas
        stats_frame = ctk.CTkFrame(self.results_frame, fg_color="white", corner_radius=10)
        stats_frame.pack(fill=tk.X, pady=10)
        
        # Contenedor para las estadísticas
        metrics_frame = ttk.Frame(stats_frame, style="Card.TFrame")
        metrics_frame.pack(fill=tk.X, padx=20, pady=10)
        
        # Configurar tres columnas iguales
        for i in range(3):
            metrics_frame.columnconfigure(i, weight=1)
        
        # Crear las métricas
        self.create_metric(metrics_frame, "Total de Centros", total_centers, "#2563eb", 0)
        self.create_metric(metrics_frame, "Grupos Generados", total_groups, "#16a34a", 1)
        self.create_metric(metrics_frame, "Promedio de Centros por Grupo", f"{avg_centers:.1f}", "#7c3aed", 2)
        
        # Separador
        separator = ttk.Frame(self.results_frame, height=2, style="BlackSeparator.TFrame")
        separator.pack(fill=tk.X, pady=15)
        
        # Mostrar grupos
        for i, group in enumerate(self.groups, 1):
            group_card = self.create_group_card(self.results_frame, i, group)
            group_card.pack(fill=tk.X, padx=5, pady=2)

    def create_metric(self, parent, title, value, color, column):
        """Crea una métrica individual."""
        frame = ctk.CTkFrame(
            parent,
            fg_color="white",
            corner_radius=8
        )
        frame.grid(row=0, column=column, padx=5, sticky='ew')
        
        # Valor
        ctk.CTkLabel(
            frame,
            text=str(value),
            font=("Segoe UI", 24, "bold"),
            text_color=color
        ).pack(pady=(10, 5))
        
        # Título
        ctk.CTkLabel(
            frame,
            text=title,
            font=("Segoe UI", 12),
            text_color="#64748b"
        ).pack(pady=(0, 10))

class ModernPortfolioAnalyzerApp(BaseApp):
    def __init__(self, root):
        self.root = root
        self.root.title("Analizador de Portafolios")
        self.current_plu_limit = 10

        # Agregar el validador de columnas
        self.column_validator = ColumnValidator()
        
        # Agregar estos atributos
        self.identical_portfolios = {}
        self.unique_portfolios = {}
        self.recommendations = []
        self.non_compatible = []

        # Agregar variable para mensajes de carga
        self.loading_message = None
        
        # Configurar pantalla completa
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        self.root.geometry(f"{screen_width}x{screen_height}+0+0")
        self.root.state('zoomed')
        
        # Variables de estado
        self.file_path_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Esperando archivo...")
        self.total_centers = tk.StringVar()
        self.unique_centers = tk.StringVar()
        self.identical_groups = tk.StringVar()
        self.initial_masters = tk.StringVar()
        self.final_masters = tk.StringVar()

        self.title_font = Font(name='Segoe UI', size=14, bold=True)
        self.subtitle_font = Font(name='Segoe UI', size=12, bold=True)
        self.base_font = Font(name='Segoe UI', size=11)
        self.header_font = Font(name='Segoe UI', size=11, bold=True)
        
        # Alineaciones
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        # Colores
        self.header_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        self.warning_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
        self.success_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
        self.alternate_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        # Configurar el tema y estilo
        self.root.configure(bg='#f8fafc')
        self.style = ttk.Style()
        self.configure_styles()
        
        self.create_widgets()
        self.setup_shortcuts()

    ICONS = {
        'variation_analysis': """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24">
            <!-- Icono de tabla con gráfico -->
            <path d="M4 6c0-1.1.9-2 2-2h12a2 2 0 012 2v12a2 2 0 01-2 2H6a2 2 0 01-2-2V6z" fill="none" stroke="currentColor" stroke-width="1.5"/>
            <!-- Líneas horizontales de la tabla -->
            <path d="M4 9h16M4 14h16" stroke="currentColor" stroke-width="1.5" stroke-linecap="round"/>
            <!-- Gráfico de tendencia -->
            <path d="M7 17l3-4 3 2 4-6" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"/>
        </svg>"""
    }

    def setup_shortcuts(self):
        """Configura los atajos de teclado para la aplicación."""
        # Atajos para acciones principales
        self.root.bind('<Control-o>', lambda e: self.browse_file())
        self.root.bind('<Control-a>', lambda e: self.analyze_portfolios())
        self.root.bind('<Control-s>', lambda e: self.export_to_excel())
        self.root.bind('<Control-i>', lambda e: self.show_reports_window())
        
        # Atajos para navegación entre pestañas
        self.root.bind('<Control-Key-1>', lambda e: self.select_tab(0))
        self.root.bind('<Control-Key-2>', lambda e: self.select_tab(1))
        self.root.bind('<Control-Key-3>', lambda e: self.select_tab(2))
        self.root.bind('<Control-Key-4>', lambda e: self.select_tab(3))
        self.root.bind('<Control-Key-5>', lambda e: self.select_tab(4))
        
        # Atajos para la ventana de informes
        def setup_report_shortcuts(window):
            # Navegación con flechas
            window.bind('<Left>', lambda e: self.change_report(-1) if hasattr(self, 'current_report') else None)
            window.bind('<Right>', lambda e: self.change_report(1) if hasattr(self, 'current_report') else None)
            # Análisis de modulación
            window.bind('<Control-m>', lambda e: self.show_modulation_analysis(
                self.current_group_num, self.current_centers) if hasattr(self, 'current_group_num') else None)
            window.bind('<Control-M>', lambda e: self.show_modulation_analysis(
                self.current_group_num, self.current_centers) if hasattr(self, 'current_group_num') else None)
        
        # Vincular la configuración de atajos a las nuevas ventanas de informes
        self.root.bind('<Configure>', lambda e: setup_report_shortcuts(e.widget) 
                    if isinstance(e.widget, tk.Toplevel) and 'Informes' in e.widget.title() else None)

    def add_custom_grouping_tab(self):
        """Agrega la pestaña de agrupación personalizada."""
        if not hasattr(self, 'custom_grouping_frame'):
            # Crear frame para la pestaña
            self.custom_grouping_frame = ttk.Frame(self.notebook, style="Card.TFrame")
            self.notebook.add(self.custom_grouping_frame, text="Agrupación Personalizada")
        
        # Limpiar el frame existente
        for widget in self.custom_grouping_frame.winfo_children():
            widget.destroy()
        
        # Crear el frame principal
        main_frame = ttk.Frame(self.custom_grouping_frame, style="Card.TFrame")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Preguntar al usuario si desea realizar la agrupación personalizada
        if messagebox.askyesno(
            "Agrupación Personalizada",
            "¿Desea realizar una agrupación personalizada de los centros?"
        ):
            # Mostrar diálogo de selección de criterios
            dialog = CustomGroupingDialog(self.root)
            criteria = dialog.show()
            
            if criteria:
                # Obtener todos los centros
                all_centers = set()
                for centers in self.identical_portfolios.keys():
                    all_centers.update(centers)
                all_centers.update(self.unique_portfolios.keys())
                
                # Realizar la agrupación personalizada
                analyzer = CustomGroupingAnalysis(self)
                groups = analyzer.perform_custom_grouping(all_centers, criteria)
                
                if groups:
                    # Mostrar resultados
                    CustomGroupResults(main_frame, groups, criteria, self)
                else:
                    self.show_no_results_message(main_frame)
        else:
            self.show_no_results_message(main_frame)

    def show_no_results_message(self, parent):
        """Muestra un mensaje cuando no hay resultados que mostrar."""
        message_frame = ctk.CTkFrame(
            parent,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        message_frame.pack(expand=True, padx=20, pady=20)
        
        # Ícono de información
        icon_size = 48
        icon_canvas = tk.Canvas(
            message_frame,
            width=icon_size,
            height=icon_size,
            bg="white",
            highlightthickness=0
        )
        icon_canvas.pack(pady=(20, 10))
        
        # Dibujar ícono de información
        icon_canvas.create_oval(
            4, 4,
            icon_size-4, icon_size-4,
            outline="#2563eb",
            width=2
        )
        icon_canvas.create_text(
            icon_size/2,
            icon_size/2,
            text="i",
            font=("Segoe UI", 24, "bold"),
            fill="#2563eb"
        )
        
        # Mensaje
        ttk.Label(
            message_frame,
            text="No hay resultados para mostrar",
            font=("Segoe UI", 16, "bold"),
            foreground="#1e293b",
            background="white"
        ).pack(pady=(0, 10))
        
        ttk.Label(
            message_frame,
            text="Seleccione 'Analizar' y elija los criterios de agrupación para ver los resultados",
            font=("Segoe UI", 12),
            foreground="#64748b",
            background="white"
        ).pack(pady=(0, 20))

    def select_tab(self, tab_index):
        """
        Selecciona una pestaña específica del notebook.
        
        Args:
            tab_index: Índice de la pestaña a seleccionar (0-4)
        """
        try:
            if hasattr(self, 'notebook') and self.notebook.index('end') > 0:
                if 0 <= tab_index < self.notebook.index('end'):
                    self.notebook.select(tab_index)
        except Exception as e:
            print(f"Error al cambiar de pestaña: {str(e)}")

    def change_report(self, direction):
        """
        Cambia el informe actual en la dirección especificada.
        
        Args:
            direction: -1 para anterior, 1 para siguiente
        """
        try:
            if hasattr(self, 'current_report') and hasattr(self, 'total_reports'):
                new_value = (self.current_report.get() + direction) % self.total_reports
                self.current_report.set(new_value)
                # Actualizar interfaz
                if hasattr(self, 'update_report_view'):
                    self.update_report_view()
        except Exception as e:
            print(f"Error al cambiar de informe: {str(e)}")

    def get_plu_limit(self):
        """
        Crear un diálogo personalizado con diseño moderno y mejorado para obtener el límite de PLUs.
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("Configurar Análisis")
        dialog.grab_set()
        
        # Dimensiones y posicionamiento mejorados
        window_width = 480
        window_height = 320
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Frame principal con gradiente
        main_frame = ctk.CTkFrame(
            dialog,
            fg_color="#ffffff",
            corner_radius=15,
            border_width=1,
            border_color="#e2e8f0"
        )
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header con icono
        header_frame = ctk.CTkFrame(
            main_frame,
            fg_color="#f8fafc",
            corner_radius=10,
            height=60
        )
        header_frame.pack(fill=tk.X, padx=15, pady=(15, 0))
        header_frame.pack_propagate(False)
        
        # Canvas para el icono
        icon_size = 32
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg="#f8fafc",
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(15, 10), pady=10)
        
        # Dibujar icono de configuración
        def draw_gear_icon():
            # Color azul moderno
            color = "#2563eb"
            
            # Círculo exterior
            icon_canvas.create_oval(4, 4, 28, 28, outline=color, width=2)
            
            # Dientes del engranaje
            for i in range(8):
                angle = i * (360/8)
                rad = math.radians(angle)
                x1 = 16 + 12 * math.cos(rad)
                y1 = 16 + 12 * math.sin(rad)
                x2 = 16 + 14 * math.cos(rad)
                y2 = 16 + 14 * math.sin(rad)
                icon_canvas.create_line(x1, y1, x2, y2, fill=color, width=2)
            
            # Círculo interior
            icon_canvas.create_oval(10, 10, 22, 22, fill=color)
        
        draw_gear_icon()
        
        # Título en el header
        header_label = ctk.CTkLabel(
            header_frame,
            text="Configuración del Análisis",
            font=("Segoe UI", 16, "bold"),
            text_color="#1e293b"
        )
        header_label.pack(side=tk.LEFT, padx=5)
        
        # Contenedor principal con efecto de profundidad
        content_frame = ctk.CTkFrame(
            main_frame,
            fg_color="#f8fafc",
            corner_radius=10
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Frame para la descripción con icono de información
        info_frame = ctk.CTkFrame(
            content_frame,
            fg_color="transparent"
        )
        info_frame.pack(fill=tk.X, padx=20, pady=(20, 0))
        
        # Canvas para icono de información
        info_icon_size = 24
        info_canvas = tk.Canvas(
            info_frame,
            width=info_icon_size,
            height=info_icon_size,
            bg="#f8fafc",
            highlightthickness=0
        )
        info_canvas.pack(side=tk.LEFT, padx=(0, 10))
        
        # Dibujar icono de información
        info_canvas.create_oval(2, 2, 22, 22, outline="#2563eb", width=2)
        info_canvas.create_text(12, 12, text="i", fill="#2563eb", font=("Segoe UI", 12, "bold"))
        
        description_label = ctk.CTkLabel(
            info_frame,
            text="Configure el número máximo de PLUs diferentes permitidos\npara agrupar los centros en el análisis",
            font=("Segoe UI", 11),
            text_color="#64748b",
            justify="left"
        )
        description_label.pack(side=tk.LEFT)
        
        # Frame para el input con diseño mejorado
        input_frame = ctk.CTkFrame(
            content_frame,
            fg_color="#ffffff",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        input_frame.pack(fill=tk.X, padx=20, pady=20)
        
        input_label = ctk.CTkLabel(
            input_frame,
            text="Máximo PLUs diferentes:",
            font=("Segoe UI", 12),
            text_color="#1e293b"
        )
        input_label.pack(side=tk.LEFT, padx=15, pady=15)
        
        # Variable para almacenar el valor
        plu_limit = tk.StringVar(value="10")
        
        # Entry con estilo mejorado
        entry = ctk.CTkEntry(
            input_frame,
            textvariable=plu_limit,
            width=100,
            height=35,
            font=("Segoe UI", 12),
            border_color="#e2e8f0",
            fg_color="#ffffff",
            text_color="#000000"  # Color negro para el texto
        )
        entry.pack(side=tk.LEFT, padx=(0, 15))
        
        # Variable para almacenar el resultado
        result = [None]
        
        def validate_and_accept():
            try:
                value = int(plu_limit.get())
                if value <= 0:
                    raise ValueError("El valor debe ser mayor que 0")
                result[0] = value
                dialog.destroy()
            except ValueError:
                # Frame de error con animación de shake
                error_frame = ctk.CTkFrame(
                    content_frame,
                    fg_color="#fef2f2",
                    corner_radius=8,
                    height=40
                )
                error_frame.pack(fill=tk.X, padx=20, pady=(0, 20))
                error_frame.pack_propagate(False)
                
                # Icono de error
                error_icon = tk.Canvas(
                    error_frame,
                    width=16,
                    height=16,
                    bg="#fef2f2",
                    highlightthickness=0
                )
                error_icon.pack(side=tk.LEFT, padx=(15, 5))
                
                # Dibujar X en rojo
                error_icon.create_line(4, 4, 12, 12, fill="#dc2626", width=2)
                error_icon.create_line(12, 4, 4, 12, fill="#dc2626", width=2)
                
                error_label = ctk.CTkLabel(
                    error_frame,
                    text="Por favor, ingrese un número entero positivo válido",
                    font=("Segoe UI", 11),
                    text_color="#dc2626"
                )
                error_label.pack(side=tk.LEFT, padx=5)
                
                # Animación de shake
                original_x = error_frame.winfo_x()
                for i in range(5):
                    error_frame.place(x=original_x + (10 if i % 2 == 0 else -10))
                    error_frame.update()
                    time.sleep(0.05)
                error_frame.place(x=original_x)
        
        def on_cancel():
            dialog.destroy()
        
        # Frame para botones con separador
        separator = ctk.CTkFrame(
            main_frame,
            fg_color="#e2e8f0",
            height=1
        )
        separator.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        button_frame = ctk.CTkFrame(
            main_frame,
            fg_color="transparent"
        )
        button_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        # Botones con diseño mejorado
        cancel_button = ctk.CTkButton(
            button_frame,
            text="Cancelar",
            command=on_cancel,
            width=120,
            height=40,
            font=("Segoe UI", 12),
            fg_color="#f1f5f9",
            hover_color="#e2e8f0",
            text_color="#64748b",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        cancel_button.pack(side=tk.RIGHT, padx=5)
        
        accept_button = ctk.CTkButton(
            button_frame,
            text="Aceptar",
            command=validate_and_accept,
            width=120,
            height=40,
            font=("Segoe UI", 12),
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            text_color="#ffffff",
            corner_radius=8
        )
        accept_button.pack(side=tk.RIGHT, padx=5)
        
        # Bind teclas y focus
        dialog.bind('<Return>', lambda e: validate_and_accept())
        dialog.bind('<Escape>', lambda e: on_cancel())
        entry.focus_set()
        
        # Esperar hasta que se cierre el diálogo
        dialog.wait_window()
        
        return result[0]

    def create_svg_icon(self, canvas, svg_code, color="#FFFFFF", size=24):
        """
        Renderiza un ícono SVG en un canvas de Tkinter.
        
        Args:
            canvas: Canvas de Tkinter donde se dibujará el ícono
            svg_code: Código SVG del ícono
            color: Color para el ícono (por defecto blanco)
            size: Tamaño del ícono (por defecto 24x24)
        """
        # Limpiar el canvas
        canvas.delete("all")
        
        # Escalar el canvas al tamaño deseado
        canvas.configure(width=size, height=size)
        
        # Reemplazar el color en el código SVG
        svg_code = svg_code.replace("currentColor", color)
        
        try:
            # Crear línea base horizontal
            canvas.create_line(
                size * 0.125, size * 0.75,  # x1, y1
                size * 0.875, size * 0.75,  # x2, y2
                fill=color,
                width=2,
                capstyle=tk.ROUND
            )
            
            # Crear línea vertical (eje Y)
            canvas.create_line(
                size * 0.125, size * 0.75,  # x1, y1
                size * 0.125, size * 0.25,  # x2, y2
                fill=color,
                width=2,
                capstyle=tk.ROUND
            )
            
            # Crear barras de variación
            # Barra 1
            canvas.create_rectangle(
                size * 0.25, size * 0.333,   # x1, y1
                size * 0.375, size * 0.75,   # x2, y2
                fill=color,
                stipple="gray50"
            )
            
            # Barra 2
            canvas.create_rectangle(
                size * 0.458, size * 0.5,    # x1, y1
                size * 0.583, size * 0.75,   # x2, y2
                fill=color,
                stipple="gray25"
            )
            
            # Barra 3
            canvas.create_rectangle(
                size * 0.667, size * 0.25,   # x1, y1
                size * 0.792, size * 0.75,   # x2, y2
                fill=color
            )
            
            # Crear línea de tendencia
            canvas.create_line(
                size * 0.3125, size * 0.583,  # x1, y1
                size * 0.520, size * 0.458,   # x2, y2
                size * 0.729, size * 0.333,   # x3, y3
                fill=color,
                width=1.5,
                dash=(2, 2)
            )
            
            # Crear puntos de marcador
            point_radius = size * 0.042  # Radio proporcional al tamaño
            for x, y in [
                (size * 0.3125, size * 0.583),
                (size * 0.520, size * 0.458),
                (size * 0.729, size * 0.333)
            ]:
                canvas.create_oval(
                    x - point_radius, y - point_radius,
                    x + point_radius, y + point_radius,
                    fill=color,
                    outline=color
                )
                
        except Exception as e:
            print(f"Error al dibujar el ícono: {str(e)}")
            # Dibujar un ícono de fallback simple
            canvas.create_rectangle(2, 2, size-2, size-2, outline=color)
            canvas.create_line(2, 2, size-2, size-2, fill=color)
            canvas.create_line(2, size-2, size-2, 2, fill=color)

    def animate_stat(self, target_label, target_value, duration=1000):
        steps = 50  # Número de pasos en la animación
        increment = target_value / steps  # Incremento en cada paso
        delay = duration // steps  # Tiempo entre cada paso (milisegundos)

        def update_label(current_value):
            if current_value < target_value:
                target_label.set(f"{int(current_value)}")
                self.root.after(delay, update_label, current_value + increment)
            else:
                target_label.set(f"{target_value}")

        update_label(0)  # Iniciar la animación desde 0

    def show_loading_spinner(self, message):
        """Mostrar spinner de carga optimizado."""
        loading_window = tk.Toplevel(self.root)
        loading_window.transient(self.root)
        loading_window.overrideredirect(True)
        loading_window.attributes('-topmost', True)
        loading_window.configure(bg='white')
        
        # Configurar ventana
        window_width = 300
        window_height = 120
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        loading_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Agregar un borde suave
        loading_window.configure(borderwidth=1, relief='solid')
        
        # Crear un marco con borde redondeado
        loading_frame = ctk.CTkFrame(
            loading_window,
            fg_color="white",
            corner_radius=10,
            border_width=0
        )
        loading_frame.pack(expand=True, fill='both', padx=2, pady=2)
        
        # Canvas para el spinner más grande
        canvas_size = 50
        canvas = tk.Canvas(
            loading_frame,
            width=canvas_size,
            height=canvas_size,
            bg='white',
            highlightthickness=0
        )
        canvas.pack(pady=(15, 10))
        
        def create_arc(start, extent):
            padding = 5
            canvas.create_arc(
                padding, padding,
                canvas_size - padding, canvas_size - padding,
                start=start,
                extent=extent,
                fill='#2563eb',
                width=0
            )
        
        # Mensajes con estilo personalizado
        self.loading_message = tk.StringVar(value=message)
        message_label = ttk.Label(
            loading_frame,
            textvariable=self.loading_message,
            font=('Segoe UI', 12),
            foreground='#64748b',
            background='white'
        )
        message_label.pack(pady=(0, 15))
        
        # Animación del spinner más suave
        rotation = [0]
        def update_spinner():
            if loading_window.winfo_exists():
                canvas.delete("all")
                create_arc(rotation[0], 280)  # Arco más corto para mejor efecto
                rotation[0] = (rotation[0] + 8) % 360  # Velocidad ajustada
                loading_window.after(16, update_spinner)  # ~60 FPS
        
        # Iniciar animación
        update_spinner()
        
        # Centrar la ventana
        loading_window.update_idletasks()
        loading_window.grab_set()  # Hacer la ventana modal
        
        return loading_window

    def update_loading_message(self, loading_window, message):
        """Actualizar mensaje de carga de manera eficiente."""
        if loading_window and loading_window.winfo_exists():
            self.loading_message.set(message)
            loading_window.update_idletasks()

    def create_widgets(self):
            # Contenedor principal
            main_frame = ttk.Frame(self.root, style="Card.TFrame")
            main_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=20)
            
            # Header con título
            header_frame = ttk.Frame(main_frame, style="Card.TFrame")
            header_frame.pack(fill=tk.X, pady=(0, 15))
            
            title_label = ttk.Label(
                header_frame,
                text="Análisis de Portafolios",
                style="Title.TLabel"
            )
            title_label.pack(side=tk.LEFT)
            
            # Línea divisoria horizontal
            separator = ttk.Frame(main_frame, height=4, style="BlackSeparator.TFrame")
            separator.pack(fill=tk.X, pady=(0, 15))
            
            # Panel de control
            control_frame = ttk.Frame(main_frame, style="Card.TFrame")
            control_frame.pack(fill=tk.X, pady=(0, 15))
            
            self.file_entry = ctk.CTkEntry(
                control_frame,
                textvariable=self.file_path_var,
                placeholder_text="Seleccione un archivo Excel...",
                height=35,
                font=("Segoe UI", 11)
            )
            self.file_entry.pack(side=tk.LEFT, padx=(0, 15), fill=tk.X, expand=True)
            
            select_button = ctk.CTkButton(
                control_frame,
                text="Seleccionar Excel",
                command=self.browse_file,
                height=35,
                width=150
            )
            select_button.pack(side=tk.LEFT, padx=5)
            
            analyze_button = ctk.CTkButton(
                control_frame,
                text="Analizar",
                command=self.analyze_portfolios,
                height=35,
                width=150
            )
            analyze_button.pack(side=tk.LEFT, padx=5)
            
            export_button = ctk.CTkButton(
                control_frame,
                text="Exportar",
                width=150,
                height=35,
                command=self.export_to_excel
            )
            export_button.pack(side=tk.LEFT, padx=5)

            self.report_button = ctk.CTkButton(
                control_frame,
                text="Generar Informes",
                width=150,
                height=35,
                command=self.show_reports_window,
                state="disabled"
            )
            self.report_button.pack(side=tk.LEFT, padx=5)
            
            # Frame para estadísticas
            stats_frame = ttk.Frame(main_frame, style="Card.TFrame")
            stats_frame.pack(fill=tk.X, pady=(0, 15))
            
            # Configurar 5 columnas iguales para las estadísticas
            for i in range(5):
                stats_frame.columnconfigure(i, weight=1)
            
            # Crear las 5 tarjetas de estadísticas
            self.create_stat_card(stats_frame, "Total de Centros", self.total_centers, '#2563eb', 0)
            self.create_stat_card(stats_frame, "Centros Únicos", self.unique_centers, '#16a34a', 1)
            self.create_stat_card(stats_frame, "Grupos Idénticos", self.identical_groups, '#9333ea', 2)
            self.create_stat_card(stats_frame, "Planogramas\nMásteres Iniciales", self.initial_masters, '#f97316', 3)
            self.create_stat_card(stats_frame, "Planogramas\nMásteres Finales", self.final_masters, '#ef4444', 4)
            
            # Línea divisoria horizontal negra y más gruesa
            separator = ttk.Frame(main_frame, height=4, style="BlackSeparator.TFrame")
            separator.pack(fill=tk.X, pady=(0, 15))
            
            # Frame para el contenido principal (gráfico y grupos)
            content_frame = ttk.Frame(main_frame, style="Card.TFrame")
            content_frame.pack(fill=tk.BOTH, expand=True)
            
            content_frame.pack_propagate(False)
            
            content_frame.grid_rowconfigure(0, weight=1)
            content_frame.grid_columnconfigure(0, weight=3)
            content_frame.grid_columnconfigure(1, weight=7)
            
            # Frame izquierdo para el gráfico
            chart_frame = ttk.Frame(content_frame, style="Card.TFrame")
            chart_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 10))
            
            # Título del gráfico
            ttk.Label(
                chart_frame,
                text="Distribución de Centros",
                font=('Segoe UI', 14, 'bold'),
                foreground='#1e293b',
                background='white'
            ).pack(pady=(0, 10))
            
            # Crear el gráfico
            self.distribution_chart = self.create_distribution_chart(chart_frame)
            
            # Frame derecho para los grupos
            groups_frame = ttk.Frame(content_frame, style="Card.TFrame")
            groups_frame.grid(row=0, column=1, sticky='nsew')
            groups_frame.grid_propagate(False)
            
            groups_frame.grid_rowconfigure(0, weight=1)
            groups_frame.grid_columnconfigure(0, weight=1)
            
            # Crear el notebook con las pestañas
            self.create_results_tabs(groups_frame)

    def get_category_input(self):
        """
        Crear un diálogo personalizado y moderno para seleccionar la categoría.
        """
        # Lista de categorías predefinidas
        categories = [
            "ACEITE",
            "ADEREZOS Y VINAGRES",
            "AFEITADA Y DEPILACION",
            "AGUAS LINEAL",
            "ALIMENTACIÓN DEL BEBÉ",
            "ALIMENTOS CONGELADOS",
            "ALMACEN",
            "AMBIENTADORES",
            "AREPAS REFRIGERADAS",
            "ARROZ LINEAL",
            "ASEO DEL BEBE",
            "AVENAS",
            "BEBIDAS LACTEAS",
            "BEBIDAS VEGETALES",
            "BLANQUEADORES CON CLORO",
            "BLANQUEADORES SIN CLORO",
            "CAFE",
            "CALDOS SOPAS Y CREMAS",
            "CARNES ENLATADAS",
            "CARNES FRIAS (EMBUTIDOS)",
            "CEREALES LISTOS",
            "CERVEZAS LINEAL",
            "CHOCOLATE DE MESA",
            "COLORACIÓN",
            "COMIDAS ESPECIALES",
            "CONDIMENTOS",
            "CONFITERIA",
            "CREMAS Y PAÑITOS HÚMEDOS",
            "CUIDADO CORPORAL Y CUIDADO FACIAL",
            "CUIDADO DE LA NALGUITA",
            "CUIDADO DE LOS PISOS",
            "CUIDADO DEL CABELLO",
            "CUIDADO ORAL",
            "DESECHABLES",
            "DESODORANTES Y ANTITRANSPIRANTES",
            "DETERGENTES EN POLVO",
            "DETERGENTES LIQUIDOS",
            "DULCES Y CONSERVAS",
            "ESPONJILLAS PAÑOS",
            "GALLETAS",
            "GASEOSAS LINEAL",
            "GRANOLAS",
            "GUANTES Y BOLSAS DE BASURA",
            "HARINAS",
            "INSECTICIDAS Y VELAS/JARDINERÍA",
            "ISOTONICOS LINEAL",
            "JABONES DE CUERPO",
            "JABONES DE TOCADOR",
            "JABONES EN BARRA",
            "JUGOS",
            "JUGOS REFRIGERADOS",
            "LAVAPLATOS",
            "LECHE EN POLVO",
            "LECHES LIQUIDAS",
            "LICORES",
            "LIMPIADORES DE SUPERFICIES Y DESINFECCIÓN Y BAÑOS",
            "MALTAS",
            "MANI",
            "MANTEQUILLAS MARGARINAS Y ESPARCIBLES",
            "MASCOTAS",
            "MODIFICADORES DE LECHE",
            "OTROS LICORES",
            "PAPEL HIGIENICO",
            "PAPEL LIMPIADOR DE COCINA",
            "PASABOCAS",
            "PASTAS ALIMENTICIAS",
            "POSTRES REFRIGERADOS",
            "PROTECCIÓN SANITARIA",
            "QUESOS REFRIGERADOS",
            "REFRESCANTES",
            "REFRESCOS EN POLVO",
            "REFRIGERADOS Y CONGELADOS",
            "REPOSTERIA",
            "RONES Y AGUARDIENTES",
            "SALSAS",
            "SALUDABLES",
            "SAZONADORAS",
            "SERVILLETAS",
            "SUAVIZANTES",
            "TE CALIENTE",
            "TE FRIO",
            "VEGETALES ENVASADOS",
            "VINOS",
            "WHISKIES"
        ]

        dialog = tk.Toplevel(self.root)
        dialog.title("Análisis de Modulación")
        dialog.grab_set()  # Hacer el diálogo modal
        
        # Centrar el diálogo
        window_width = 600
        window_height = 500
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Frame principal con fondo blanco y bordes redondeados
        main_frame = ctk.CTkFrame(
            dialog,
            fg_color="white",
            corner_radius=10
        )
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Header con icono
        header_frame = ctk.CTkFrame(
            main_frame,
            fg_color="#f8fafc",
            corner_radius=10,
            height=60
        )
        header_frame.pack(fill=tk.X, padx=15, pady=(15, 0))
        header_frame.pack_propagate(False)
        
        # Canvas para el icono
        icon_size = 32
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg="#f8fafc",
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(15, 10), pady=10)
        
        # Dibujar icono de categoría
        def draw_category_icon():
            color = "#2563eb"
            # Dibujar folder
            icon_canvas.create_polygon(
                4, 8, 12, 8, 14, 12, 28, 12, 28, 26, 4, 26,
                fill="#2563eb", outline=color, width=1.5
            )
            # Dibujar líneas de categoría
            for y in [16, 20, 24]:
                icon_canvas.create_line(8, y, 24, y, fill="white", width=1.5)
        
        draw_category_icon()
        
        # Título en el header
        header_label = ctk.CTkLabel(
            header_frame,
            text="Selección de Categoría",
            font=("Segoe UI", 16, "bold"),
            text_color="#1e293b"
        )
        header_label.pack(side=tk.LEFT, padx=5)
        
        # Frame de búsqueda
        search_frame = ctk.CTkFrame(
            main_frame,
            fg_color="#f8fafc",
            corner_radius=10,
            height=50
        )
        search_frame.pack(fill=tk.X, padx=15, pady=15)
        
        # Variable para búsqueda
        search_var = tk.StringVar()
        filtered_categories = tk.StringVar(value=categories)
        
        def on_search(*args):
            search_term = search_var.get().lower()
            filtered = [cat for cat in categories if search_term in cat.lower()]
            filtered_categories.set(filtered)
            
            # Actualizar el contador
            count_label.configure(text=f"{len(filtered)} categorías encontradas")
        
        search_var.trace('w', on_search)
        
        # Campo de búsqueda con icono
        search_icon = tk.Canvas(
            search_frame,
            width=16,
            height=16,
            bg="#f8fafc",
            highlightthickness=0
        )
        search_icon.pack(side=tk.LEFT, padx=(15, 5), pady=15)
        
        # Dibujar lupa
        search_icon.create_oval(2, 2, 12, 12, outline="#64748b", width=1.5)
        search_icon.create_line(10, 10, 14, 14, fill="#64748b", width=1.5)
        
        search_entry = ctk.CTkEntry(
            search_frame,
            textvariable=search_var,
            placeholder_text="Buscar categoría...",
            font=("Segoe UI", 12),
            width=200,
            height=35
        )
        search_entry.pack(side=tk.LEFT, padx=5)
        
        # Contador de resultados
        count_label = ctk.CTkLabel(
            search_frame,
            text=f"{len(categories)} categorías encontradas",
            font=("Segoe UI", 11),
            text_color="#64748b"
        )
        count_label.pack(side=tk.RIGHT, padx=15)
        
        # Frame para la lista de categorías
        list_frame = ctk.CTkFrame(
            main_frame,
            fg_color="#ffffff",
            corner_radius=10,
            border_width=1,
            border_color="#e2e8f0"
        )
        list_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        
        # Lista de categorías con scrollbar
        listbox = tk.Listbox(
            list_frame,
            listvariable=filtered_categories,
            font=("Segoe UI", 11),
            selectmode=tk.SINGLE,
            activestyle='none',
            bg="white",
            fg="#1e293b",
            selectbackground="#2563eb",
            selectforeground="white",
            highlightthickness=0,
            bd=0
        )
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(2, 0), pady=2)
        
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        listbox.configure(yscrollcommand=scrollbar.set)
        
        # Frame para botones con separador
        separator = ctk.CTkFrame(
            main_frame,
            fg_color="#e2e8f0",
            height=1
        )
        separator.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        button_frame = ctk.CTkFrame(
            main_frame,
            fg_color="transparent"
        )
        button_frame.pack(fill=tk.X, padx=15)
        
        # Variable para almacenar el resultado
        result = [None]
        
        def on_accept():
            selection = listbox.curselection()
            if not selection:
                # Mostrar error si no hay selección
                error_frame = ctk.CTkFrame(
                    main_frame,
                    fg_color="#fef2f2",
                    corner_radius=8,
                    height=40
                )
                error_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
                error_frame.pack_propagate(False)
                
                error_icon = tk.Canvas(
                    error_frame,
                    width=16,
                    height=16,
                    bg="#fef2f2",
                    highlightthickness=0
                )
                error_icon.pack(side=tk.LEFT, padx=(15, 5))
                error_icon.create_line(4, 4, 12, 12, fill="#dc2626", width=2)
                error_icon.create_line(12, 4, 4, 12, fill="#dc2626", width=2)
                
                error_label = ctk.CTkLabel(
                    error_frame,
                    text="Por favor, seleccione una categoría",
                    font=("Segoe UI", 11),
                    text_color="#dc2626"
                )
                error_label.pack(side=tk.LEFT, padx=5)
                
                # Animar error y destruir después de 2 segundos
                dialog.after(2000, error_frame.destroy)
                return
                
            result[0] = listbox.get(selection[0])
            dialog.destroy()
        
        def on_cancel():
            dialog.destroy()
        
        # Botones con diseño mejorado
        cancel_button = ctk.CTkButton(
            button_frame,
            text="Cancelar",
            command=on_cancel,
            width=120,
            height=40,
            font=("Segoe UI", 12),
            fg_color="#f1f5f9",
            hover_color="#e2e8f0",
            text_color="#64748b",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        cancel_button.pack(side=tk.RIGHT, padx=5)
        
        accept_button = ctk.CTkButton(
            button_frame,
            text="Aceptar",
            command=on_accept,
            width=120,
            height=40,
            font=("Segoe UI", 12),
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            text_color="#ffffff",
            corner_radius=8
        )
        accept_button.pack(side=tk.RIGHT, padx=5)
        
        # Doble click para seleccionar
        listbox.bind('<Double-Button-1>', lambda e: on_accept())
        
        # Bind teclas
        dialog.bind('<Return>', lambda e: on_accept())
        dialog.bind('<Escape>', lambda e: on_cancel())
        
        # Focus en búsqueda
        search_entry.focus_set()
        
        # Esperar hasta que se cierre el diálogo
        dialog.wait_window()
        
        return result[0]

    def get_modulation_data(self, centers, category):
        """
        Obtiene los datos de modulación para un conjunto de centros y una categoría específica.
        Incluye centros con datos vacíos y no encontrados.
        
        Args:
            centers (list): Lista de centros a analizar
            category (str): Categoría a buscar
            
        Returns:
            dict: Diccionario con la estructura {num_modulos: {'count': n, 'centers': [...]}}
        """
        try:
            maestro_file = self.get_data_path("Form.Maestro.Neg.xlsx")
            
            if not os.path.exists(maestro_file):
                raise FileNotFoundError("No se encontró el archivo Form.Maestro.Neg.xlsx")
            
            # Convertir centros a set para búsqueda más eficiente
            centers = set(str(c).strip() for c in centers)
            
            # Leer la hoja de MOBILIARIOS
            df = pd.read_excel(maestro_file, sheet_name='MOBILIARIOS')
            
            # Normalizar la categoría y convertir columnas a string
            category = category.upper().strip()
            df['CENTRO'] = df['CENTRO'].astype(str).str.strip()
            df['CATEGORIA'] = df['CATEGORIA'].astype(str).str.upper().str.strip()
            
            # Filtrar por categoría
            df_filtered = df[df['CATEGORIA'] == category]
            
            # Identificar diferentes tipos de centros
            centers_in_excel = set(df_filtered['CENTRO'].unique())
            centers_with_empty_data = set()
            centers_with_data = set()
            
            # Analizar cada centro en el Excel
            for center in centers:
                center_data = df_filtered[df_filtered['CENTRO'] == center]
                if not center_data.empty:
                    # Verificar si tiene valor en NÚMERO DE MÓDULOS
                    if pd.isna(center_data['NÚMERO DE MÓDULOS'].iloc[0]) or center_data['NÚMERO DE MÓDULOS'].iloc[0] == '':
                        centers_with_empty_data.add(center)
                    else:
                        centers_with_data.add(center)
            
            # Identificar centros no encontrados
            centers_not_found = centers - centers_in_excel
            
            # Inicializar resultado
            result = {}
            
            # Agregar datos por número de módulos
            if not df_filtered.empty:
                for num_modulos, group in df_filtered[df_filtered['CENTRO'].isin(centers_with_data)].groupby('NÚMERO DE MÓDULOS'):
                    result[str(num_modulos)] = {
                        'count': len(group),
                        'centers': sorted(group['CENTRO'].tolist())
                    }
            
            # Agregar centros con datos vacíos
            if centers_with_empty_data:
                result['Datos vacíos'] = {
                    'count': len(centers_with_empty_data),
                    'centers': sorted(list(centers_with_empty_data))
                }
                
            # Agregar centros no encontrados
            if centers_not_found:
                result['No encontrados'] = {
                    'count': len(centers_not_found),
                    'centers': sorted(list(centers_not_found))
                }
                
            # Imprimir información de diagnóstico
            print(f"\nDiagnóstico de modulación:")
            print(f"Total centros en grupo: {len(centers)}")
            print(f"Centros con datos: {len(centers_with_data)}")
            print(f"Centros con datos vacíos: {len(centers_with_empty_data)}")
            print(f"Centros no encontrados: {len(centers_not_found)}")
            
            return result
            
        except Exception as e:
            print(f"Error al leer datos de modulación: {str(e)}")
            return None

    def create_modulation_section(self, parent, category, modulation_data):
        """
        Crea una sección visual para mostrar el análisis de modulación.
        
        Args:
            parent: Frame padre donde se creará la sección
            category (str): Categoría analizada
            modulation_data (dict): Datos de modulación procesados
            
        Returns:
            Frame: Frame conteniendo la tabla de modulación
        """
        # Crear frame principal con borde y esquinas redondeadas
        section_frame = ctk.CTkFrame(
            parent,
            fg_color="white",
            corner_radius=10,
            border_width=1,
            border_color="#E5E7EB"
        )
        
        # Header
        header_frame = ctk.CTkFrame(
            section_frame,
            fg_color="#F9FAFB",
            corner_radius=0,
            height=50
        )
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # Título con icono
        ctk.CTkLabel(
            header_frame,
            text=f"Análisis de Modulación - {category}",
            font=("Segoe UI", 13, "bold"),
            text_color="#111827"
        ).pack(side=tk.LEFT, padx=15, pady=15)
        
        # Frame para la tabla
        table_frame = ctk.CTkFrame(
            section_frame,
            fg_color="white",
            corner_radius=0
        )
        table_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Headers de la tabla
        headers = ["No. Módulos", "No. Centros", "Centros"]
        for i, header in enumerate(headers):
            header_label = ctk.CTkLabel(
                table_frame,
                text=header,
                font=("Segoe UI", 12, "bold"),
                text_color="#374151"
            )
            header_label.grid(row=0, column=i, padx=10, pady=(0, 10), sticky="w")
        
        # Línea separadora
        separator = ctk.CTkFrame(
            table_frame,
            height=1,
            fg_color="#E5E7EB"
        )
        separator.grid(row=1, column=0, columnspan=3, sticky="ew", padx=10, pady=5)
        
        if modulation_data:
            # Ordenar las claves para que los números aparezcan primero
            sorted_keys = sorted(
                modulation_data.keys(),
                key=lambda x: (
                    # Primero los números
                    float('inf') if not str(x).replace('.', '').isdigit() else float(x),
                    # Luego el orden específico para las categorías especiales
                    0 if x == 'Datos vacíos' else 1 if x == 'No encontrados' else 2,
                    # Finalmente el texto
                    str(x)
                )
            )
            
            row_idx = 2  # Empezar después del header y el separador
            for key in sorted_keys:
                data = modulation_data[key]
                
                # Color de texto según el tipo de dato
                text_color = "#4B5563"  # Color normal para números
                if key == "Datos vacíos":
                    text_color = "#EAB308"  # Amarillo para datos vacíos
                elif key == "No encontrados":
                    text_color = "#EF4444"  # Rojo para no encontrados
                
                # Número de módulos
                ctk.CTkLabel(
                    table_frame,
                    text=str(key),
                    font=("Segoe UI", 11),
                    text_color=text_color
                ).grid(row=row_idx, column=0, padx=10, pady=5, sticky="w")
                
                # Número de centros
                ctk.CTkLabel(
                    table_frame,
                    text=str(data['count']),
                    font=("Segoe UI", 11),
                    text_color=text_color
                ).grid(row=row_idx, column=1, padx=10, pady=5, sticky="w")
                
                # Lista de centros
                ctk.CTkLabel(
                    table_frame,
                    text=", ".join(data['centers']),
                    font=("Segoe UI", 11),
                    text_color=text_color
                ).grid(row=row_idx, column=2, padx=10, pady=5, sticky="w")
                
                row_idx += 1
        else:
            # Mensaje si no hay datos
            ctk.CTkLabel(
                table_frame,
                text="No se encontraron datos de modulación para esta categoría",
                font=("Segoe UI", 11),
                text_color="#EF4444"
            ).grid(row=2, column=0, columnspan=3, padx=10, pady=20)
        
        return section_frame

    def show_reports_window(self):
        """Mostrar la ventana de reportes."""
        # Crear una nueva ventana para reportes
        reports_window = tk.Toplevel(self.root)
        reports_window.title("Informes de Análisis")
        
        # Configurar pantalla completa
        screen_width = reports_window.winfo_screenwidth()
        screen_height = reports_window.winfo_screenheight()
        reports_window.geometry(f"{screen_width}x{screen_height}+0+0")
        reports_window.state('zoomed')
        
        # Variables de control
        self.current_report = tk.IntVar(value=0)
        self.total_reports = 2

        # Agregar atajos específicos para esta ventana
        def on_left(event):
            change_report(-1)
            
        def on_right(event):
            change_report(1)
            
        def change_report(direction):
            new_value = (self.current_report.get() + direction) % self.total_reports
            self.current_report.set(new_value)
            title_label.configure(text=update_title())
            subtitle_label.configure(text=update_subtitle())
            show_current_report()
            prev_btn.configure(state="normal" if new_value > 0 else "disabled")
            next_btn.configure(state="normal" if new_value < self.total_reports - 1 else "disabled")
        
        # Vincular teclas de flecha
        reports_window.bind('<Left>', on_left)
        reports_window.bind('<Right>', on_right)
        
        # Vincular Ctrl+M para análisis de modulación
        def on_ctrl_m(event):
            if hasattr(self, 'current_group_num') and hasattr(self, 'current_centers'):
                self.show_modulation_analysis(self.current_group_num, self.current_centers)
        
        reports_window.bind('<Control-m>', on_ctrl_m)
        reports_window.bind('<Control-M>', on_ctrl_m) 
        
        # Frame principal con fondo blanco
        main_frame = ttk.Frame(reports_window, style="Card.TFrame")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Frame superior para título y navegación
        header_frame = ttk.Frame(main_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(20, 0))
        
        def update_title():
            if self.current_report.get() == 0:
                return "Análisis de grupos"
            else:
                return "Análisis de grupos"
        
        def update_subtitle():
            if self.current_report.get() == 0:
                return "Distribución general de centros"
            else:
                return "Análisis de distribución por grupo"
        
        # Contenedor para el título y subtítulo
        title_container = ttk.Frame(header_frame, style="Card.TFrame")
        title_container.pack(fill=tk.X, padx=40)
        
        # Frame izquierdo para título y subtítulo
        text_container = ttk.Frame(title_container, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.Y)
        
        # Título con fuente en negrita
        title_label = ttk.Label(
            text_container,
            text=update_title(),
            font=('Segoe UI', 28, 'bold'),
            foreground='#1e293b',
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Subtítulo
        subtitle_label = ttk.Label(
            text_container,
            text=update_subtitle(),
            font=('Segoe UI', 14),
            foreground='#6B7280',
            background='white'
        )
        subtitle_label.pack(anchor='w')
        
        # Frame para botones de navegación
        nav_frame = ttk.Frame(title_container, style="Card.TFrame")
        nav_frame.pack(side=tk.RIGHT)
        
        def change_report(direction):
            new_value = (self.current_report.get() + direction) % self.total_reports
            self.current_report.set(new_value)
            title_label.configure(text=update_title())
            subtitle_label.configure(text=update_subtitle())
            show_current_report()
            prev_btn.configure(state="normal" if new_value > 0 else "disabled")
            next_btn.configure(state="normal" if new_value < self.total_reports - 1 else "disabled")
        
        # Botones de navegación
        prev_btn = ctk.CTkButton(
            nav_frame,
            text="←",
            width=40,
            height=40,
            command=lambda: change_report(-1),
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            state="disabled"
        )
        prev_btn.pack(side=tk.LEFT, padx=5)
        
        next_btn = ctk.CTkButton(
            nav_frame,
            text="→",
            width=40,
            height=40,
            command=lambda: change_report(1),
            fg_color="#2563eb",
            hover_color="#1d4ed8"
        )
        next_btn.pack(side=tk.LEFT, padx=5)
        
        # Línea divisoria más gruesa
        separator = ttk.Frame(main_frame, height=4, style="BlackSeparator.TFrame")
        separator.pack(fill=tk.X, pady=(10, 15))  # Reducido el padding superior
        
        # Frame para contenido
        self.reports_frame = ttk.Frame(main_frame, style="Card.TFrame")
        self.reports_frame.pack(fill=tk.BOTH, expand=True)
        
        def show_current_report():
            # Limpiar el frame actual
            for widget in self.reports_frame.winfo_children():
                widget.destroy()
            
            if self.current_report.get() == 0:
                self.generate_general_report()
            else:
                self.generate_group_reports()
        
        # Mostrar el primer informe
        show_current_report()

    def generate_general_report(self):
        """Generar el informe general de distribución con diseño mejorado."""
        try:
            # Cargar datos geográficos
            geo_data = self.load_geographic_data()
            if geo_data is None:
                return
            
            # Obtener todos los centros analizados
            all_centers = set()
            for centers in self.identical_portfolios.keys():
                all_centers.update(centers)
            all_centers.update(self.unique_portfolios.keys())
            
            # Filtrar datos para los centros analizados
            centers_data = geo_data[geo_data['Centro'].isin(all_centers)]
            
            # Crear el contenedor principal con altura fija
            main_container = ctk.CTkFrame(
                self.reports_frame,
                fg_color="#F9FAFB",
                corner_radius=0
            )
            main_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=15) 

            def create_bar_chart_icon(parent, size=16):
                """Crear ícono de barras estadísticas."""
                icon = ctk.CTkCanvas(
                    parent,
                    width=size,
                    height=size,
                    bg='#F9FAFB',
                    highlightthickness=0
                )
                
                bar_width = 3
                spacing = 2
                
                x1 = 2
                icon.create_rectangle(
                    x1, size-6,
                    x1+bar_width, size-2,
                    fill='#9CA3AF', outline=''
                )
                
                x2 = x1 + bar_width + spacing
                icon.create_rectangle(
                    x2, 2,
                    x2+bar_width, size-2,
                    fill='#9CA3AF', outline=''
                )
                
                x3 = x2 + bar_width + spacing
                icon.create_rectangle(
                    x3, size-8,
                    x3+bar_width, size-2,
                    fill='#9CA3AF', outline=''
                )
                
                return icon

            def show_tooltip(event, text):
                """Mostrar tooltip con información."""
                tooltip = tk.Toplevel()
                tooltip.wm_overrideredirect(True)
                tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
                
                frame = ttk.Frame(tooltip, style="Tooltip.TFrame", padding=5)
                frame.pack(fill='both', expand=True)
                
                label = ttk.Label(
                    frame,
                    text=text,
                    style="Tooltip.TLabel",
                    background='#1e293b',
                    foreground='white',
                    font=('Segoe UI', 10)
                )
                label.pack()
                
                return tooltip

            def hide_tooltip(tooltip):
                """Ocultar tooltip."""
                if tooltip:
                    tooltip.destroy()

            def normalize_text(text):
                """Elimina tildes, normaliza espacios y símbolos"""
                text = str(text).upper().strip()
                # Reemplazar variaciones de separadores
                text = text.replace(' / ', ' ').replace('/', ' ')
                # Eliminar tildes
                text = ''.join(c for c in unicodedata.normalize('NFD', text)
                            if unicodedata.category(c) != 'Mn')
                # Normalizar espacios
                text = ' '.join(text.split())
                return text

            def create_data_table_with_chart(parent, title, data, district=None, base_color=None, width_ratio=0.6):
                """Crear tabla de datos con gráfico integrado."""
                # Frame contenedor principal
                container = ctk.CTkFrame(
                    parent,
                    fg_color="white",
                    corner_radius=10,
                    border_width=1,
                    border_color="#E5E7EB"
                )
                
                # Header
                header_frame = ctk.CTkFrame(
                    container,
                    fg_color="#F9FAFB",
                    corner_radius=0,
                    height=40
                )
                header_frame.pack(fill=tk.X)
                header_frame.pack_propagate(False)
                
                bar_icon = create_bar_chart_icon(header_frame)
                bar_icon.pack(side=tk.LEFT, padx=(15, 8), pady=10)
                
                ctk.CTkLabel(
                    header_frame,
                    text=title,
                    font=("Segoe UI", 12, "bold"),
                    text_color="#111827"
                ).pack(side=tk.LEFT, padx=5, pady=10)
                
                # Frame para contenido
                content_frame = ctk.CTkFrame(
                    container,
                    fg_color="white",
                    corner_radius=0
                )
                content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10)) 

                # Tabla
                table_frame = ctk.CTkFrame(
                    content_frame,
                    fg_color="white",
                    corner_radius=0
                )
                table_frame.pack(fill=tk.BOTH, expand=True)
                
                # Definir colores para cada distrito y región
                district_colors = {
                    'COSTA': '#4299e1',    # Azul
                    'INTERIOR': '#10b981'  # Verde
                }
                
                region_colors = {
                    'BARRANQUILLA NORTE CARTAGENA': '#2563EB',  # Azul royal
                    'BARRANQUILLA CENTROSUR': '#7C3AED',        # Violeta
                    'SANTA MARTA VALLEDUPAR': '#EC4899',        # Rosa intenso
                    'SINCELEJO MONTERIA': '#F59E0B',            # Ámbar
                    'CENTRO II': '#10B981',                     # Esmeralda
                    'OCCIDENTE': '#0EA5E9',                     # Celeste
                    'CAFETERO': '#94A3B8',                      # Gris
                    'CENTRO I': '#8B5CF6'                       # Púrpura
                }
                
                if district:
                    district_data = centers_data[centers_data['Distrito'] == district]
                    data = district_data['Region'].value_counts()
                    percentages = (data / len(district_data)) * 100
                    color_dict = region_colors
                else:
                    percentages = (data / len(centers_data)) * 100
                    color_dict = district_colors

                # Crear filas de datos
                for i, (name, count) in enumerate(data.items()):
                    percentage = percentages[name]
                    normalized_name = normalize_text(name)
                    
                    try:
                        row_color = color_dict[normalized_name]
                    except KeyError:
                        print(f"No se encontró color para: {name} (normalizado: {normalized_name})")
                        row_color = '#bfdbfe' if district == 'COSTA' else '#a7f3d0'
                    
                    # Frame para cada fila
                    row_frame = ctk.CTkFrame(
                        table_frame,
                        fg_color="white",
                        corner_radius=0,
                        height=50
                    )
                    row_frame.pack(fill=tk.X, pady=(0, 1))
                    row_frame.pack_propagate(False)
                    
                    # Contenedor para nombre
                    name_frame = ctk.CTkFrame(
                        row_frame,
                        fg_color="white",
                        corner_radius=0
                    )
                    name_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
                    
                    # Indicador de color
                    indicator = ctk.CTkFrame(
                        name_frame,
                        width=8,
                        height=8,
                        fg_color=row_color,
                        corner_radius=4
                    )
                    indicator.pack(side=tk.LEFT, padx=(0, 8), pady=15)
                    
                    # Nombre
                    ctk.CTkLabel(
                        name_frame,
                        text=name,
                        font=("Segoe UI", 12),
                        text_color="#374151"
                    ).pack(side=tk.LEFT, pady=15)
                    
                    # Frame para estadísticas
                    stats_frame = ctk.CTkFrame(
                        row_frame,
                        fg_color="white",
                        corner_radius=0,
                        width=150
                    )
                    stats_frame.pack(side=tk.RIGHT)
                    stats_frame.pack_propagate(False)
                    
                    # Cantidad
                    count_label = ctk.CTkLabel(
                        stats_frame,
                        text=f"{count}",
                        font=("Segoe UI", 12),
                        text_color="#4B5563"
                    )
                    count_label.pack(anchor="e")
                    
                    ctk.CTkLabel(
                        stats_frame,
                        text="centros",
                        font=("Segoe UI", 10),
                        text_color="#6B7280"
                    ).pack(anchor="e")
                    
                    # Frame para porcentaje y barra
                    percentage_frame = ctk.CTkFrame(
                        row_frame,
                        fg_color="white",
                        corner_radius=0,
                        width=100
                    )
                    percentage_frame.pack(side=tk.RIGHT, padx=10)
                    percentage_frame.pack_propagate(False)
                    
                    # Porcentaje
                    ctk.CTkLabel(
                        percentage_frame,
                        text=f"{percentage:.1f}%",
                        font=("Segoe UI", 12, "bold"),
                        text_color=row_color
                    ).pack(anchor="e")
                    
                    # Barra de progreso
                    progress_bg = ctk.CTkFrame(
                        percentage_frame,
                        height=4,
                        fg_color="#E5E7EB",
                        corner_radius=2,
                        width=100
                    )
                    progress_bg.pack(anchor="e", pady=(4, 0))
                    
                    progress = ctk.CTkFrame(
                        progress_bg,
                        height=4,
                        fg_color=row_color,
                        corner_radius=2,
                        width=int(percentage)
                    )
                    progress.place(relx=0, rely=0, relheight=1)
                
                # Frame para el gráfico
                chart_frame = ctk.CTkFrame(
                    content_frame,
                    fg_color="white",
                    corner_radius=0,
                    width=400
                )
                chart_frame.pack(fill=tk.BOTH, expand=True, pady=(20, 0))
                
                # Crear gráfico
                fig = Figure(figsize=(4, 3), dpi=100)
                ax = fig.add_subplot(111)
                
                # Variable para almacenar el tooltip actual
                current_tooltip = [None]
                
                if district:
                    # Obtener colores para las regiones del distrito actual
                    colors = []
                    for region in data.index:
                        try:
                            colors.append(region_colors[normalize_text(region)])
                        except KeyError:
                            colors.append('#bfdbfe' if district == 'COSTA' else '#a7f3d0')
                    
                    # Crear gráfico circular sin etiquetas
                    wedges = ax.pie(
                        percentages.values,
                        colors=colors,
                        startangle=90,
                        wedgeprops=dict(width=0.7)
                    )[0]
                    
                    def on_motion(event):
                        if current_tooltip[0]:
                            hide_tooltip(current_tooltip[0])
                            current_tooltip[0] = None
                        
                        if event.inaxes == ax:
                            for i, wedge in enumerate(wedges):
                                if not event.xdata or not event.ydata:
                                    continue
                                if wedge.contains_point([event.xdata, event.ydata]):
                                    region_name = list(percentages.index)[i]
                                    region_value = percentages[region_name]
                                    tooltip_text = f"{region_name}\n{region_value:.1f}%"
                                    current_tooltip[0] = show_tooltip(event, tooltip_text)
                                    break
                    
                else:
                    # Gráfico de barras para distritos con colores específicos
                    colors = [district_colors.get(normalize_text(district), '#bfdbfe') 
                            for district in data.index]
                    bars = ax.bar(
                        range(len(percentages)),
                        percentages.values,
                        color=colors,
                        width=0.5
                    )
                    
                    # Quitar etiquetas del eje x
                    ax.set_xticks([])
                    ax.set_yticks([])
                    
                    def on_motion(event):
                        if current_tooltip[0]:
                            hide_tooltip(current_tooltip[0])
                            current_tooltip[0] = None
                        
                        if event.inaxes == ax:
                            for i, bar in enumerate(bars):
                                if not event.xdata or not event.ydata:
                                    continue
                                if bar.contains_point([event.xdata, event.ydata]):
                                    district_name = list(percentages.index)[i]
                                    district_value = percentages[district_name]
                                    tooltip_text = f"{district_name}\n{district_value:.1f}%"
                                    current_tooltip[0] = show_tooltip(event, tooltip_text)
                                    break
                    
                    # Quitar bordes
                    for spine in ax.spines.values():
                        spine.set_visible(False)
                
                def on_leave(event):
                    if current_tooltip[0]:
                        hide_tooltip(current_tooltip[0])
                        current_tooltip[0] = None
                
                fig.tight_layout(pad=0.5)
                
                # Crear canvas y configurar eventos
                canvas = FigureCanvasTkAgg(fig, master=chart_frame)
                canvas.draw()
                widget = canvas.get_tk_widget()
                widget.pack(fill=tk.BOTH, expand=True)
                
                # Vincular eventos
                canvas.mpl_connect('motion_notify_event', on_motion)
                widget.bind('<Leave>', on_leave)

                fig.tight_layout(pad=1.5)
                
                return container

            # Crear frame principal con 3 columnas
            main_grid = ctk.CTkFrame(
                main_container,
                fg_color="transparent",
                corner_radius=0
            )
            main_grid.pack(fill=tk.BOTH, expand=True)
            
            # Configurar las columnas con pesos diferentes
            main_grid.columnconfigure(0, weight=30)  # Distrito
            main_grid.columnconfigure(1, weight=60)  # COSTA
            main_grid.columnconfigure(2, weight=30)  # INTERIOR

            # Definir colores para los distritos
            colors = {'COSTA': '#4299e1', 'INTERIOR': '#10b981'}

            # 1. Detalle por Distrito
            district_counts = centers_data['Distrito'].value_counts()
            district_section = create_data_table_with_chart(
                main_grid,
                "Detalle por Distrito",
                district_counts,
                width_ratio=0.6
            )
            district_section.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)

            # 2. Detalle de Regiones - COSTA
            costa_section = create_data_table_with_chart(
                main_grid,
                "Detalle de Regiones - COSTA",
                None,
                district='COSTA',
                width_ratio=0.6
            )
            costa_section.grid(row=0, column=1, sticky='nsew', padx=5, pady=5)

            # 3. Detalle de Regiones -
            # 3. Detalle de Regiones - INTERIOR
            interior_section = create_data_table_with_chart(
                main_grid,
                "Detalle de Regiones - INTERIOR",
                None,
                district='INTERIOR',
                width_ratio=0.6
            )
            interior_section.grid(row=0, column=2, sticky='nsew', padx=5, pady=5)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar el informe general: {str(e)}")

    def generate_group_reports(self):
        """Generar los informes de grupos finales con análisis de modulación."""
        try:                
            # Cargar datos geográficos
            geo_data = self.load_geographic_data()
            if geo_data is None:
                return
                
            # Obtener grupos finales
            final_groups = self.calculate_final_groups(self.current_plu_limit)
            
            # Crear el contenedor principal
            main_container = ctk.CTkFrame(
                self.reports_frame,
                fg_color="#F9FAFB",
                corner_radius=0,
                border_width=0
            )
            main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=0)

            def normalize_text(text):
                """Elimina tildes, espacios extra y normaliza separadores"""
                text = str(text).upper().strip()
                # Reemplazar variaciones de separadores
                text = text.replace(' / ', ' ').replace('/', ' ')
                # Eliminar tildes
                text = ''.join(c for c in unicodedata.normalize('NFD', text)
                            if unicodedata.category(c) != 'Mn')
                # Normalizar espacios múltiples
                text = ' '.join(text.split())
                return text
                
            # Definir colores base
            district_colors = {
                'COSTA': '#4299e1',
                'INTERIOR': '#10b981'
            }
            
            # Colores para regiones con nombres normalizados
            base_region_colors = {
                'BARRANQUILLA NORTE / CARTAGENA': '#2563EB',
                'BARRANQUILLA CENTRO/SUR': '#7C3AED',
                'SANTA MARTA / VALLEDUPAR': '#EC4899',
                'SINCELEJO / MONTERÍA': '#F59E0B',
                'CENTRO II': '#10B981',
                'OCCIDENTE': '#0EA5E9',
                'CAFETERO': '#94A3B8',
                'CENTRO I': '#8B5CF6'
            }
            
            # Crear diccionario de colores normalizados
            region_colors = {normalize_text(k): v for k, v in base_region_colors.items()}
            
            def get_region_color(region_name):
                """Obtener color para una región, con fallback a un color por defecto."""
                normalized_name = normalize_text(region_name)
                return region_colors.get(normalized_name, '#bfdbfe')  # Color azul claro por defecto

            def create_legend_item(parent, name, percentage, count, color, type="district"):
                """Crear un elemento de leyenda con diseño horizontal."""
                # Frame contenedor principal
                legend_frame = ctk.CTkFrame(
                    parent,
                    fg_color="white",
                    corner_radius=0,
                    height=40
                )
                legend_frame.pack(fill=tk.X, pady=2)
                legend_frame.pack_propagate(False)
                
                # Contenedor para nombre e indicador de color
                name_frame = ctk.CTkFrame(
                    legend_frame,
                    fg_color="white",
                    corner_radius=0
                )
                name_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
                
                # Indicador de color
                indicator = ctk.CTkFrame(
                    name_frame,
                    width=6,
                    height=6,
                    fg_color=color,
                    corner_radius=3
                )
                indicator.pack(side=tk.LEFT, padx=(0, 8), pady=15)
                
                # Nombre
                ctk.CTkLabel(
                    name_frame,
                    text=name,
                    font=("Segoe UI", 11),
                    text_color="#374151"
                ).pack(side=tk.LEFT, pady=15)
                
                # Frame para estadísticas
                stats_frame = ctk.CTkFrame(
                    legend_frame,
                    fg_color="white",
                    corner_radius=0,
                    width=250  # Ancho fijo para estadísticas
                )
                stats_frame.pack(side=tk.RIGHT)
                stats_frame.pack_propagate(False)
                
                # Barra de progreso - Fondo
                progress_bg = ctk.CTkFrame(
                    stats_frame,
                    height=8,
                    fg_color="#E5E7EB",
                    corner_radius=4,
                    width=100
                )
                progress_bg.pack(side=tk.LEFT, pady=(16, 0))
                
                # Barra de progreso - Relleno
                progress_fill = ctk.CTkFrame(
                    progress_bg,
                    height=8,
                    fg_color=color,
                    corner_radius=4
                )
                progress_fill.place(relx=0, rely=0, relheight=1, relwidth=percentage/100)
                
                # Porcentaje
                percentage_label = ctk.CTkLabel(
                    stats_frame,
                    text=f"{percentage:.1f}%",
                    font=("Segoe UI", 11),
                    text_color=color,
                    width=60
                )
                percentage_label.pack(side=tk.LEFT, padx=(10, 0), pady=(16, 0))
                
                # Número de centros
                count_label = ctk.CTkLabel(
                    stats_frame,
                    text=f"{count} {'centro' if count == 1 else 'centros'}",
                    font=("Segoe UI", 11),
                    text_color="#6B7280",
                    width=80
                )
                count_label.pack(side=tk.LEFT, padx=(5, 0), pady=(16, 0))
                
                return legend_frame

            def create_chart_section(parent, data, chart_type="district"):
                """
                Crear sección de gráfico.
                chart_type puede ser "district" o "region"
                """
                chart_frame = ctk.CTkFrame(
                    parent,
                    fg_color="white",
                    corner_radius=0
                )
                chart_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
                
                fig = Figure(figsize=(6, 4), dpi=100)
                ax = fig.add_subplot(111)
                
                values = [item['percentage'] for item in data]
                colors = [item['color'] for item in data]
                
                if chart_type == "district":
                    # Gráfico de barras para distritos
                    bars = ax.bar(
                        range(len(values)),
                        values,
                        color=colors,
                        width=0.6
                    )
                    
                    # Configurar ejes
                    ax.set_xticks([])
                    ax.set_yticks([])
                    
                    # Quitar bordes
                    for spine in ax.spines.values():
                        spine.set_visible(False)
                else:
                    # Gráfico circular para regiones
                    wedges = ax.pie(
                        values,
                        colors=colors,
                        startangle=90,
                        labels=None,
                        autopct=None,
                        wedgeprops=dict(width=0.5)
                    )[0]
                    
                    ax.axis('equal')
                
                # Crear canvas y agregar a frame
                canvas = FigureCanvasTkAgg(fig, master=chart_frame)
                canvas.draw()
                canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
                
                return chart_frame

            def create_data_table_with_chart(parent, title, group_data):
                """Crear tabla de datos con gráfico usando el nuevo diseño."""
                # Crear contenedor principal
                container = ctk.CTkFrame(
                    parent,
                    fg_color="white",
                    corner_radius=10,
                    border_width=1,
                    border_color="#E5E7EB"
                )
                container.pack(fill=tk.X, padx=5, pady=5)
                
                # Header frame
                header_frame = ctk.CTkFrame(
                    container,
                    fg_color="#F9FAFB",
                    corner_radius=0,
                    height=50
                )
                header_frame.pack(fill=tk.X)
                header_frame.pack_propagate(False)

                # Frame izquierdo para título
                title_frame = ctk.CTkFrame(
                    header_frame,
                    fg_color="transparent"
                )
                title_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

                # Título
                ctk.CTkLabel(
                    title_frame,
                    text=title,
                    font=("Segoe UI", 13, "bold"),
                    text_color="#111827"
                ).pack(side=tk.LEFT, padx=15, pady=15)

                # Botón de análisis
                group_num = int(title.split()[1])  # Extraer número de grupo del título
                analyze_button = ctk.CTkButton(
                    header_frame,
                    text="Ver Análisis de Modulación",
                    command=lambda g=group_num, c=group['centers']: self.show_modulation_analysis(g, c),
                    width=180,
                    height=32,
                    fg_color="#2563eb",
                    hover_color="#1d4ed8"
                )
                analyze_button.pack(side=tk.RIGHT, padx=15)
                                
                # Frame para el contenido principal
                content_frame = ctk.CTkFrame(
                    container,
                    fg_color="white",
                    corner_radius=0
                )
                content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
                
                # Configurar grid para las tres secciones
                content_frame.grid_columnconfigure(0, weight=40)  # Detalle por Distrito
                content_frame.grid_columnconfigure(1, weight=20)  # COSTA
                content_frame.grid_columnconfigure(2, weight=40)  # INTERIOR
                
                # 1. Sección de Detalle por Distrito
                district_data = group_data['Distrito'].value_counts()
                district_percentages = (district_data / len(group_data)) * 100
                
                district_col = ctk.CTkFrame(content_frame, fg_color="white")
                district_col.grid(row=0, column=0, sticky="nsew", padx=(0, 20))
                
                # Título de la sección
                ctk.CTkLabel(
                    district_col,
                    text="Detalle por Distrito",
                    font=("Segoe UI", 12, "bold"),
                    text_color="#111827"
                ).pack(anchor="w", pady=(0, 15))
                
                # Crear datos para el distrito
                district_items = []
                for district in ['COSTA', 'INTERIOR']:
                    if district in district_data.index:
                        count = district_data[district]
                        percentage = district_percentages[district]
                        district_items.append({
                            'name': district,
                            'percentage': percentage,
                            'count': count,
                            'color': district_colors[district]
                        })
                
                # Agregar leyendas y gráfico
                legend_frame = ctk.CTkFrame(district_col, fg_color="white")
                legend_frame.pack(fill=tk.X, pady=(0, 15))
                
                for item in district_items:
                    create_legend_item(
                        legend_frame,
                        item['name'],
                        item['percentage'],
                        item['count'],
                        item['color']
                    )
                
                create_chart_section(district_col, district_items, "district")
                
                # 2 y 3. Secciones de Regiones
                for idx, district in enumerate(['COSTA', 'INTERIOR']):
                    if district in district_data.index:
                        # Calcular datos para las regiones
                        district_regions = group_data[group_data['Distrito'] == district]
                        region_data = district_regions['Region'].value_counts()
                        region_percentages = (region_data / len(district_regions)) * 100
                        
                        region_col = ctk.CTkFrame(content_frame, fg_color="white")
                        region_col.grid(
                            row=0,
                            column=idx+1,
                            sticky="nsew",
                            padx=(0, 20 if idx == 0 else 0)
                        )
                        
                        # Título de la sección
                        ctk.CTkLabel(
                            region_col,
                            text=f"Detalle de Regiones - {district}",
                            font=("Segoe UI", 12, "bold"),
                            text_color="#111827"
                        ).pack(anchor="w", pady=(0, 15))
                        
                        # Frame para leyendas
                        region_legend_frame = ctk.CTkFrame(region_col, fg_color="white")
                        region_legend_frame.pack(fill=tk.X, pady=(0, 15))
                        
                        # Crear datos para las regiones
                        for region in region_data.index:
                            count = region_data[region]
                            percentage = region_percentages[region]
                            
                            create_legend_item(
                                region_legend_frame,
                                region,
                                percentage,
                                count,
                                get_region_color(region),
                                type="region"
                            )
                        
                        # Convertir datos para el gráfico
                        region_items = [{
                            'name': region,
                            'percentage': region_percentages[region],
                            'count': region_data[region],
                            'color': get_region_color(region)
                        } for region in region_data.index]
                        
                        # Crear gráfico de la región
                        create_chart_section(region_col, region_items, chart_type="region")

                return container

            # Crear el notebook (sistema de pestañas)
            notebook = ttk.Notebook(main_container, style="Custom.TNotebook")
            notebook.pack(fill=tk.BOTH, expand=True, pady=10)

            # Configurar estilo de las pestañas
            self.style.configure(
                "Custom.TNotebook.Tab",
                padding=[20, 10],
                font=('Segoe UI', 11),
                background="white"
            )

            self.style.map(
                "Custom.TNotebook.Tab",
                background=[("selected", "white")],
                foreground=[("selected", "#2563eb")],
                expand=[("selected", [1, 1, 1, 0])]
            )

            # Procesar cada grupo final
            for i, group in enumerate(final_groups, 1):
                # Crear un frame para cada pestaña
                tab_frame = ttk.Frame(notebook, style="Card.TFrame")
                notebook.add(tab_frame, text=f"Grupo final {i}")
                
                # Crear contenedor para el grupo actual dentro de la pestaña
                group_container = ctk.CTkFrame(
                    tab_frame,
                    fg_color="white",
                    corner_radius=8,
                    border_width=1,
                    border_color="#E5E7EB"
                )
                group_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

                # Datos geográficos del grupo
                group_data = geo_data[geo_data['Centro'].isin(group['centers'])]
                
                # Crear la sección de distribución existente
                create_data_table_with_chart(
                    group_container,
                    f"Grupo {i} - {len(group['centers'])} centros",
                    group_data
                )

        except Exception as e:
            messagebox.showerror("Error", f"Error al generar el informe de grupos: {str(e)}")
            print(f"Error detallado: {str(e)}")
 
    def show_modulation_analysis(self, group_num, centers):
        try:
            # Primero pedir la categoría
            category = self.get_category_input()
            if not category:
                return
                
            # Obtener los datos de modulación
            modulation_data = self.get_modulation_data(centers, category)
            if not modulation_data:
                messagebox.showerror(
                    "Error",
                    f"No se encontraron datos de modulación para la categoría {category}"
                )
                return
            
            # Crear la ventana de análisis
            analysis_window = tk.Toplevel(self.root)
            analysis_window.title(f"Análisis de Modulación - Grupo {group_num}")
            analysis_window.state('zoomed')
            
            # Frame principal con padding y fondo blanco
            main_frame = ctk.CTkFrame(
                analysis_window,
                fg_color="white",
                corner_radius=0
            )
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Frame para el título
            title_section = ttk.Frame(main_frame, style='White.TFrame')
            title_section.pack(fill=tk.X, padx=20, pady=(20, 0))
            
            # Título principal
            ttk.Label(
                title_section,
                text="Análisis de Modulación",
                font=("Segoe UI", 28, "bold"),
                foreground="#1E293B",
                background="white"
            ).pack(anchor='w')
            
            # Frame para subtítulo y total centros
            info_section = ttk.Frame(main_frame, style='White.TFrame')
            info_section.pack(fill=tk.X, padx=20, pady=(0, 10))
            
            # Subtítulo (izquierda)
            ttk.Label(
                info_section,
                text=f"Grupo {group_num} • Categoría: {category}",
                font=("Segoe UI", 11),
                foreground="#6B7280",
                background="white"
            ).pack(side=tk.LEFT)
            
            # Total centros (derecha)
            ttk.Label(
                info_section,
                text=f"Total centros: {len(centers)}",
                font=("Segoe UI", 11),
                foreground="#6B7280",
                background="white"
            ).pack(side=tk.RIGHT)
            
            # Línea separadora inferior
            # separator_bottom = ttk.Frame(main_frame, height=4, style='Line.TFrame')
            # separator_bottom.pack(fill=tk.X)
            
            # Configurar estilos
            style = ttk.Style()
            style.configure('White.TFrame', background='white')
            style.configure('Line.TFrame', background='black')
                
            # Contenedor para tabla y gráfico
            content_frame = ttk.Frame(main_frame, style='White.TFrame')  # Agregamos el estilo White.TFrame
            content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
            content_frame.grid_columnconfigure(0, weight=35)
            content_frame.grid_columnconfigure(1, weight=65)
            
            # Frame para la tabla con borde suave
            table_frame = ctk.CTkFrame(
                content_frame,
                fg_color="white",
                corner_radius=10,
                border_width=3,
                border_color="black"
            )
            table_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
            
            # Estilo moderno para la tabla
            style = ttk.Style()
            style.configure(
                "Treeview",
                background="white",
                fieldbackground="white",
                rowheight=40,
                font=("Segoe UI", 11)
            )
            style.configure(
                "Treeview.Heading",
                background="white",  # Cambiar de gris a blanco
                foreground="#111827",
                font=("Segoe UI", 11, "bold"),
                padding=10
            )
            style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])
            style.map(
                "Treeview",
                background=[('selected', '#EEF2FF')],
                foreground=[('selected', '#4F46E5')]
            )
            
            # Crear tabla
            columns = ("Modulación", "# Centros", "% del Total", "Centros")
            tree = ttk.Treeview(
                table_frame,
                columns=columns,
                show='headings',
                style="Treeview"
            )
            
            # Configurar columnas
            for col in columns:
                tree.heading(col, text=col)
                if col == "Centros":
                    tree.column(col, width=400, anchor='w')
                else:
                    tree.column(col, width=100, anchor='center')
            
            # Frame para el gráfico
            chart_frame = ctk.CTkFrame(
                content_frame,
                fg_color="white",
                corner_radius=10,
                border_width=3,
                border_color="black"
            )
            chart_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
            
            # Crear gráfico de barras
            figure = Figure(figsize=(8, 6), dpi=100)
            figure.patch.set_facecolor('#FFFFFF')
            ax = figure.add_subplot(111)
            
            # Preparar datos para el gráfico
            labels = []
            sizes = []
            total_centers = len(centers)
            colors = ['#4F46E5', '#6366F1', '#818CF8', '#A5B4FC', '#C7D2FE', '#E0E7FF']
            
            # Ordenar datos por número de módulos
            sorted_data = sorted(
                [(k, v) for k, v in modulation_data.items()],
                key=lambda x: float(x[0]) if x[0].isdigit() else float('inf')
            )
            
            for num_modulos, data in sorted_data:
                percentage = (data['count'] / total_centers) * 100
                labels.append(str(num_modulos))
                sizes.append(percentage)
            
            # Crear gráfico de barras
            bars = ax.bar(labels, sizes, color=colors[:len(labels)])

            # Modificación en la configuración del gráfico
            ax.set_xticklabels(labels, rotation=90)  # Rotar etiquetas 90 grados
            ax.tick_params(axis='x', pad=10)  # Añadir padding para las etiquetas rotadas

            # Ajustar los márgenes del gráfico para acomodar las etiquetas rotadas
            figure.subplots_adjust(bottom=0.2)
                        
            # Configurar estilo del gráfico
            ax.set_title(
                "Distribución de Módulos",
                pad=20,
                fontsize=14,
                fontweight="bold",
                color="#111827"
            )
            ax.set_ylabel("Porcentaje (%)")
            
            # Agregar valores sobre las barras
            for bar in bars:
                height = bar.get_height()
                ax.text(
                    bar.get_x() + bar.get_width()/2.,
                    height,
                    f'{height:.1f}%',
                    ha='center',
                    va='bottom',
                    color="#111827"
                )
            
            # Mejorar apariencia del gráfico
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.set_axisbelow(True)
            ax.grid(axis='y', linestyle='--', alpha=0.3)
            
            canvas = FigureCanvasTkAgg(figure, chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
            
            # Modificar la inserción de datos en la tabla para el ajuste automático
            for num_modulos, data in sorted_data:
                percentage = (data['count'] / total_centers) * 100
                
                # Formatear centros con saltos de línea cada cierto número de elementos
                centers_list = sorted(data['centers'], key=lambda x: int(x))
                centers_text = ""
                line = []
                chars_per_line = 80  # Aumentamos el número de caracteres por línea
                
                for center in centers_list:
                    # Si agregar el siguiente centro excede el límite de caracteres
                    if line and len(', '.join(line + [center])) > chars_per_line:
                        # Agregar la línea actual al texto final
                        centers_text += ', '.join(line) + '\n'
                        # Comenzar nueva línea con el centro actual
                        line = [center]
                    else:
                        # Agregar el centro a la línea actual
                        line.append(center)
                
                # Agregar la última línea si hay centros pendientes
                if line:
                    centers_text += ', '.join(line)
                
                # Insertar en la tabla
                item = tree.insert("", tk.END, values=(
                    num_modulos,
                    data['count'],
                    f"{percentage:.1f}%",
                    centers_text.strip()  # Eliminar espacios en blanco extra
                ))
                
                # Ajustar altura de la fila según el contenido
                tree.item(item, tags=('row',))

            # Función para ajustar la altura de las filas
            def adjust_row_heights():
                for item in tree.get_children():
                    centers_text = tree.item(item)['values'][3]
                    if centers_text:
                        # Contar líneas en el texto
                        num_lines = centers_text.count('\n') + 1
                        # Ajustar altura (25 píxeles por línea para optimizar espacio)
                        row_height = max(35, num_lines * 35)
                        tree.tag_configure('row', rowheight=row_height)

            # Asegurarnos que la columna de centros sea lo suficientemente ancha
            tree.column("Centros", width=700, anchor='w', stretch=True)

            # Llamar a la función después de que la ventana se haya dibujado
            analysis_window.after(100, adjust_row_heights)

            # Configurar la columna de centros para que se ajuste al contenido
            tree.column("Centros", width=400, anchor='w', stretch=True)
            
            tree.pack(fill=tk.BOTH, expand=True, padx=15, pady=(15, 5))

            # Mensaje informativo dentro del frame de la tabla
            info_frame = ctk.CTkFrame(
                table_frame,
                fg_color="#FEF2F2",
                corner_radius=10,
                height=40
            )
            info_frame.pack(fill=tk.X, padx=15, pady=(0, 15))

            ttk.Label(
                info_frame,
                text="⚠️ Los centros con datos vacíos o no encontrados requieren revisión",
                font=("Segoe UI", 11),
                foreground="#991B1B",  # Color de texto rojo más oscuro
                background="#FEF2F2",  # Mismo color de fondo que el frame
                padding=(15, 8)
            ).pack(side=tk.LEFT)

            
        except Exception as e:
            messagebox.showerror("Error", f"Error al mostrar el análisis: {str(e)}")

    def load_geographic_data(self):
        """Cargar datos geográficos desde el archivo Excel."""
        try:
            # Usar el método de la clase base para obtener la ruta del archivo
            db_file = self.get_data_path("db_maestrospdv.xlsx")
            
            # Leer archivo
            df = pd.read_excel(db_file)
        
            # Verificar existencia de columnas exactas
            required_columns = {
                'Centro': self.column_validator.find_column(df, 'CENTRO'),
                'Distrito': 'Distrito',  # Nombre exacto
                'Region': 'Región'       # Nombre exacto con tilde
            }
            
            # Verificar si las columnas existen
            missing_columns = []
            for col_name in ['Distrito', 'Región']:
                if col_name not in df.columns:
                    missing_columns.append(col_name)
            
            if missing_columns:
                raise ValueError(
                    f"No se encontraron las siguientes columnas exactas: "
                    f"{', '.join(missing_columns)}"
                )
            
            # Crear DataFrame con las columnas encontradas
            result_df = pd.DataFrame({
                'Centro': df[required_columns['Centro']],
                'Distrito': df['Distrito'],
                'Region': df['Región']
            })
            
            # Limpiar datos manteniendo mayúsculas/minúsculas originales
            for col in result_df.columns:
                result_df[col] = result_df[col].astype(str).str.strip()
                
            result_df = result_df.drop_duplicates()
            result_df = result_df.dropna()
            
            return result_df
            
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return None
    
    def generate_district_distribution(self, group_centers, geo_data):
        """Generar visualización de distribución por distrito y región."""
        # Filtrar datos para los centros del grupo
        group_data = geo_data[geo_data['Centro'].isin(group_centers)]
        
        # Crear figura con dos subplots
        fig = Figure(figsize=(12, 5))
        
        # Configurar el estilo general
        text_color = '#1e293b'  # Color oscuro para texto

        # Definir colores específicos para distritos
        district_colors = {
            'COSTA': '#4299e1',    # Azul claro
            'INTERIOR': '#10b981'  # Verde esmeralda
        }
        
        # Definir colores específicos para regiones
        region_colors = {
            'BARRANQUILLA NORTE / CARTAGENA': '#2563EB',  # Azul royal
            'BARRANQUILLA CENTRO/SUR': '#7C3AED',        # Violeta
            'SANTA MARTA / VALLEDUPAR': '#EC4899',        # Rosa intenso
            'SINCELEJO / MONTERÍA': '#F59E0B',            # Ámbar
            'CENTRO II': '#10B981',                     # Esmeralda
            'OCCIDENTE': '#0EA5E9',                     # Celeste
            'CAFETERO': '#94A3B8',                      # Gris
            'CENTRO I': '#8B5CF6'                       # Púrpura
        }
        
        # Distribución por distrito
        ax1 = fig.add_subplot(121)
        district_counts = group_data['Distrito'].value_counts()
        district_percentages = (district_counts / len(group_data)) * 100
        
        bars = ax1.bar(district_percentages.index, district_percentages.values)
        ax1.set_title('Distribución por Distrito', 
                    fontfamily='Segoe UI', 
                    fontsize=12, 
                    color=text_color)
        ax1.set_ylabel('Porcentaje', color=text_color)
        ax1.tick_params(axis='x', rotation=45, colors=text_color)
        ax1.tick_params(axis='y', colors=text_color)
        
        # Agregar etiquetas de datos mejoradas
        for bar in bars:
            height = bar.get_height()
            count = int(height * len(group_data) / 100)  # Calcular cantidad de centros
            ax1.text(
                bar.get_x() + bar.get_width()/2.,
                height,
                f'{height:.1f}%\n({count})',
                ha='center',
                va='bottom',
                fontfamily='Segoe UI',
                fontsize=9,
                color=text_color,
                fontweight='bold'
            )
        
        # Distribución por región
        ax2 = fig.add_subplot(122)
        
        # Datos para el gráfico de anillos
        region_district_data = defaultdict(lambda: defaultdict(int))
        for _, row in group_data.iterrows():
            region_district_data[row['Distrito']][row['Region']] += 1
        
        # Calcular porcentajes y totales
        total_centers = len(group_data)
        district_colors = plt.cm.Set3(np.linspace(0, 1, len(region_district_data)))
        
        # Crear gráfico de anillos con etiquetas mejoradas
        start = 0
        legend_elements = []
        
        for district, regions in region_district_data.items():
            values = [count/total_centers * 100 for count in regions.values()]
            wedges, texts, autotexts = ax2.pie(
                values,
                labels=regions.keys(),
                startangle=start,
                radius=0.8,
                labeldistance=1.1,
                autopct=lambda pct: f'{pct:.1f}%\n({int(pct*total_centers/100)})',
                pctdistance=0.75,
                wedgeprops=dict(width=0.5)
            )
            
            # Mejorar el estilo de las etiquetas
            for autotext in autotexts:
                autotext.set_fontsize(8)
                autotext.set_fontfamily('Segoe UI')
                autotext.set_color(text_color)
            for text in texts:
                text.set_fontsize(8)
                text.set_fontfamily('Segoe UI')
                text.set_color(text_color)
            
            legend_elements.extend(wedges)
            start += sum(values)
        
        ax2.set_title('Distribución Regional por Distrito', 
                    fontfamily='Segoe UI', 
                    fontsize=12,
                    color=text_color)
        
        fig.tight_layout(rect=[0, 0.05, 1, 0.95])
        return fig

    def create_report_section(self, group_num, centers, fig):
        """Crear una sección de reporte para un grupo específico."""
        # Frame para la sección
        section_frame = ttk.Frame(self.reports_frame, style="Card.TFrame")
        section_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Header frame
        header_frame = ttk.Frame(section_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(10, 5))
        
        # Título y botón en la misma línea
        title_frame = ttk.Frame(header_frame, style="Card.TFrame")
        title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Título del grupo
        title_label = ttk.Label(
            title_frame,
            text=f"Grupo {group_num}",
            font=('Segoe UI', 16, 'bold'),
            foreground='#2563eb',
            background='white'
        )
        title_label.pack(side=tk.LEFT, pady=(0, 5))
        
        # Botón de análisis
        analyze_button = ctk.CTkButton(
            header_frame,
            text="Analizar Variación",
            width=150,
            height=35,
            command=lambda: self.show_portfolio_variation(group_num, centers),
            fg_color="#2563eb",
            hover_color="#1d4ed8"
        )
        analyze_button.pack(side=tk.RIGHT, padx=10)
        
        # Subtítulo con cantidad de centros
        subtitle_label = ttk.Label(
            section_frame,
            text=f"{len(centers)} centros",
            font=('Segoe UI', 12),
            foreground='#6b7280',
            background='white'
        )
        subtitle_label.pack(pady=(0, 10))
        
        # Agregar el gráfico
        canvas = FigureCanvasTkAgg(fig, master=section_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

    def get_group_portfolio_data(self, centers, selected_fields):
        """Obtener datos del portafolio para el grupo seleccionado."""
        try:
            # Leer archivo original
            df = pd.read_excel(self.file_path_var.get())
            
            # Buscar la columna de centro ignorando mayúsculas/minúsculas y espacios
            centro_column = None
            plu_column = None
            
            for col in df.columns:
                col_clean = str(col).upper().strip()
                if 'CENTRO' in col_clean:
                    centro_column = col
                if 'PLU' in col_clean:
                    plu_column = col
            
            if not centro_column or not plu_column:
                raise ValueError(
                    f"No se encontraron las columnas requeridas.\n"
                    f"Columnas disponibles: {', '.join(df.columns)}"
                )
            
            # Filtrar por centros del grupo
            df_group = df[df[centro_column].astype(str).isin([str(c) for c in centers])]
            
            # Crear diccionario para almacenar resultados
            results = []
            
            # Obtener PLUs únicos
            unique_plus = df_group[plu_column].unique()
            
            for plu in unique_plus:
                item = {}
                # Agregar campos seleccionados
                for field in selected_fields:
                    if field in df.columns:
                        value = df[df[plu_column] == plu][field].iloc[0]
                        item[field] = value
                    else:
                        item[field] = ''  # Valor vacío si no se encuentra el campo
                
                # Agregar marca de verificación para cada centro
                for center in centers:
                    has_plu = df[
                        (df[plu_column] == plu) & 
                        (df[centro_column].astype(str) == str(center))
                    ].shape[0] > 0
                    item[str(center)] = "✓" if has_plu else ""
                
                results.append(item)
            
            return results
            
        except Exception as e:
            error_msg = str(e)
            if "No se encontraron las columnas requeridas" in error_msg:
                messagebox.showerror(
                    "Error", 
                    "No se encontraron las columnas necesarias en el archivo.\n"
                    "Asegúrese de que el archivo contenga las columnas:\n"
                    "- Centro (o similar)\n"
                    "- PLU_SAP (o similar)"
                )
            else:
                messagebox.showerror(
                    "Error",
                    f"Error al obtener datos: {error_msg}\n"
                    "Por favor, verifique el formato del archivo."
                )
            return []

    def get_available_columns(self):
        """Obtener las columnas disponibles del archivo original."""
        try:
            df = pd.read_excel(self.file_path_var.get())
            # Filtrar columnas, excluyendo la de centro que se agrega automáticamente
            columns = [col for col in df.columns if 'CENTRO' not in str(col).upper()]
            return columns
        except Exception as e:
            messagebox.showerror(
                "Error", 
                "Error al leer las columnas del archivo.\n"
                "Por favor, verifique que el archivo sea accesible."
            )
            return ['PLU_SAP', 'Articulo']  # Columnas por defecto

    def find_column_match(self, df, variations):
        """
        Busca una columna en el DataFrame que coincida con alguna de las variaciones proporcionadas.
        
        Args:
            df: DataFrame de pandas
            variations: Lista de posibles nombres de columna
            
        Returns:
            str: Nombre de la columna encontrada o None si no se encuentra
        """
        # Limpiar nombres de columnas del DataFrame
        df_columns = {col: str(col).upper().strip() for col in df.columns}
        
        # 1. Buscar coincidencia exacta
        for col, clean_col in df_columns.items():
            if clean_col in [var.upper().strip() for var in variations]:
                return col
        
        # 2. Buscar coincidencia parcial
        for col, clean_col in df_columns.items():
            for variation in variations:
                var_clean = variation.upper().strip()
                if var_clean in clean_col or clean_col in var_clean:
                    return col
        
        # 3. Buscar por palabras clave
        for col, clean_col in df_columns.items():
            for variation in variations:
                var_parts = variation.upper().strip().split()
                if all(part in clean_col for part in var_parts):
                    return col
        
        return None

    def normalize_dataframe_columns(self, df):
        """
        Normaliza las columnas del DataFrame buscando las variaciones comunes de nombres.
        
        Args:
            df: DataFrame original
            
        Returns:
            DataFrame: DataFrame con columnas normalizadas
        """
        column_variations = {
            'Centro': [
                'Centro', 'Centros', 'Cod_Centro', 'Codigo_Centro', 'Centro_Id',
                'Id_Centro', 'Num_Centro', 'Centro_Codigo'
            ],
            'Categoria': [
                'Categoria', 'Categorias', 'Cat', 'Categoria_Prod',
                'Categoria_Producto', 'Clase_Producto'
            ],
            'Subcategoria': [
                'Subcategoria', 'Sub_Categoria', 'Subcategorias', 'Sub_Categorias',
                'SubCat', 'Sub_Cat', 'Subclase'
            ],
            'Segmento': [
                'Segmento', 'Segmentos', 'Seg', 'Segmento_Prod',
                'Segmento_Producto', 'Linea_Producto'
            ],
            'PLU_SAP': [
                'PLU_SAP', 'PLU', 'Codigo_PLU', 'Cod_PLU', 'SKU',
                'Codigo_Producto', 'Cod_Producto'
            ],
            'Articulo': [
                'Articulo', 'Articulos', 'Descripcion', 'Desc',
                'Nombre_Producto', 'Producto'
            ]
        }
        
        column_mapping = {}
        missing_columns = []
        
        # Buscar coincidencias para cada columna requerida
        for std_name, variations in column_variations.items():
            found_col = self.find_column_match(df, variations)
            if found_col:
                column_mapping[found_col] = std_name
            else:
                missing_columns.append(std_name)
        
        if missing_columns:
            error_msg = (
                f"No se encontraron las siguientes columnas requeridas: "
                f"{', '.join(missing_columns)}\n"
                f"Columnas disponibles: {', '.join(df.columns)}"
            )
            raise ValueError(error_msg)
        
        # Renombrar columnas y retornar nuevo DataFrame
        return df.rename(columns=column_mapping)

    def show_portfolio_variation(self, group_num, centers):
        """Mostrar ventana de análisis de variación de portafolio."""
        try:
            # Leer el archivo Excel
            df = pd.read_excel(self.file_path_var.get(), sheet_name='Sheet1')
            
            # Validar columnas requeridas (CENTRO y PLU_SAP)
            required_columns = self.column_validator.validate_required_columns(
                df, ['CENTRO', 'PLU_SAP']
            )
            
            # Intentar encontrar columnas opcionales
            optional_columns = ['CATEGORIA', 'SUBCATEGORIA', 'SEGMENTO', 'ARTICULO']
            for col_type in optional_columns:
                try:
                    found_col = self.column_validator.find_column(df, col_type, raise_error=False)
                    if found_col:
                        required_columns[col_type] = found_col
                    else:
                        # Si no se encuentra la columna, crear una columna con '-'
                        df[col_type] = '-'
                        required_columns[col_type] = col_type
                except:
                    # Si hay cualquier error, crear una columna con '-'
                    df[col_type] = '-'
                    required_columns[col_type] = col_type

            # Primero filtrar usando el nombre original de la columna
            centro_col = required_columns['CENTRO']
            df_filtered = df[df[centro_col].astype(str).isin([str(c) for c in centers])]

            # Crear una copia para evitar SettingWithCopyWarning
            df_filtered = df_filtered.copy()

            # Renombrar las columnas al formato esperado por SimplePivotTable
            column_mapping = {
                required_columns['CENTRO']: 'Centro',
                required_columns['PLU_SAP']: 'PLU_SAP',
                required_columns['CATEGORIA']: 'Categoria',
                required_columns['SUBCATEGORIA']: 'Subcategoria',
                required_columns['SEGMENTO']: 'Segmento',
                required_columns['ARTICULO']: 'Articulo'
            }
            
            # Aplicar el renombramiento
            df_filtered = df_filtered.rename(columns=column_mapping)
            
            # Verificar que tenemos todas las columnas necesarias
            expected_columns = ['Centro', 'PLU_SAP', 'Categoria', 'Subcategoria', 'Segmento', 'Articulo']
            for col in expected_columns:
                if col not in df_filtered.columns:
                    df_filtered[col] = '-'
            
            # Crear la ventana de pivot
            SimplePivotTable(
                parent=self.root,
                data=df_filtered,
                title=f"Análisis de Variación - Grupo {group_num}"
            )
            
        except ValueError as e:
            if "columnas requeridas" in str(e):
                messagebox.showerror(
                    "Error",
                    "Error en el formato del archivo:\n\n"
                    "No se encontraron las columnas básicas necesarias (CENTRO, PLU_SAP).\n"
                    "Por favor, verifique que el archivo contenga estas columnas."
                )
            else:
                messagebox.showerror("Error", str(e))
        except Exception as e:
            messagebox.showerror(
                "Error", 
                f"Error al analizar la variación: {str(e)}\n"
                f"Columnas encontradas: {', '.join(df_filtered.columns if 'df_filtered' in locals() else [])}"
            )

    def generate_reports(self):
        """Generar reportes para todos los grupos."""
        try:
            # Cargar datos geográficos
            geo_data = self.load_geographic_data()
            if geo_data is None:
                return
            
            # Obtener grupos finales
            final_groups = self.calculate_final_groups(self.current_plu_limit)
            
            # Crear reporte para cada grupo
            for i, group in enumerate(final_groups, 1):
                # Generar gráficos
                fig = self.generate_district_distribution(group['centers'], geo_data)
                
                # Crear sección de reporte
                self.create_report_section(i, group['centers'], fig)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar reportes: {str(e)}")

    def create_title_row(self, ws, title, row=1, subtitle=None):
        """Crear fila de título con formato y opcional subtítulo"""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = self.title_font
        cell.alignment = self.left_alignment
        
        if subtitle:
            sub_cell = ws.cell(row=row+1, column=1, value=subtitle)
            sub_cell.font = self.subtitle_font
            sub_cell.alignment = self.left_alignment
            return row + 3
        return row + 2

    def add_table_headers(self, ws, headers, row):
        """Agregar encabezados de tabla con formato"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
        return row + 1

    def adjust_column_widths(self, worksheet, min_width=8, max_width=60):
        """
        Ajusta el ancho de las columnas según el contenido.
        
        Args:
            worksheet: Worksheet de openpyxl
            min_width: Ancho mínimo de columna
            max_width: Ancho máximo de columna
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            # Calcular el ancho máximo necesario
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        # Considerar saltos de línea
                        if '\n' in str(cell.value):
                            cell_length = max(len(line) for line in str(cell.value).split('\n'))
                        max_length = max(max_length, cell_length)
                except:
                    pass
            
            # Ajustar el ancho con padding
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = max(min(adjusted_width, max_width), min_width)

    def configure_group_final_sheet(self, worksheet):
        """
        Configura el formato específico para las hojas de grupos finales.
        
        Args:
            worksheet: Worksheet de openpyxl
        """
        # Fijar ancho de columnas específicas
        fixed_width_columns = {'A': 40, 'B': 40, 'D': 40, 'E': 40}
        for col, width in fixed_width_columns.items():
            worksheet.column_dimensions[col].width = width
        
        # Activar wrap text para todas las celdas
        for row in worksheet.rows:
            for cell in row:
                if cell.value:  # Solo configurar celdas con contenido
                    cell.alignment = Alignment(
                        horizontal='left',
                        vertical='center',
                        wrap_text=True
                    )

    def create_centers_sheet(self, workbook, title, centers_data, base_font, header_font, header_fill):
        """
        Crea una hoja para listar centros con formato profesional.
        
        Args:
            workbook: Workbook de openpyxl
            title: Título de la hoja
            centers_data: Lista de tuplas (centro, plus)
            base_font: Fuente base
            header_font: Fuente para encabezados
            header_fill: Relleno para encabezados
        """
        ws = workbook.create_sheet(title)
        current_row = 1
        
        # Título
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = Font(name='Segoe UI', size=14, bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 2
        
        # Encabezados
        headers = ["Centro", "Cantidad de PLUs", "PLUs"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Datos
        for center, plus_set in centers_data:
            row_data = [
                center,
                len(plus_set),
                ", ".join(map(str, sorted(plus_set)))
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = base_font
                cell.alignment = Alignment(
                    horizontal='left' if col == 3 else 'center',
                    vertical='center',
                    wrap_text=True
                )
            current_row += 1
        
        # Ajustar ancho de columnas
        self.adjust_column_widths(ws)
        
        return ws

    def format_pivot_tables(self, workbook):
        """
        Aplica formato de tabla dinámica y ajuste de texto a todas las hojas.
        
        Args:
            workbook: Workbook de openpyxl
        """
        for ws in workbook.worksheets:
            # Aplicar wrap_text y alineación a todas las celdas con contenido
            for row in ws.rows:
                for cell in row:
                    if cell.value:  # Solo configurar celdas con contenido
                        cell.alignment = Alignment(
                            horizontal='left',
                            vertical='center',
                            wrap_text=True
                        )
            
            # Ajustar ancho de columnas para todas las hojas
            self.adjust_column_widths(ws)
            
            # Aplicar formato de tabla solo a las hojas que no son de grupos finales
            if not ws.title.startswith("Grupo Final"):
                try:
                    # Encontrar el rango de datos
                    data_range = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
                    
                    # Crear tabla con estilo personalizado
                    table = Table(
                        displayName=f"Table_{ws.title.replace(' ', '_')}",
                        ref=data_range
                    )
                    
                    # Aplicar estilo de tabla
                    style = TableStyleInfo(
                        name="TableStyleMedium20",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                    table.tableStyleInfo = style
                                
                    # Agregar tabla a la hoja
                    ws.add_table(table)
                    
                except Exception:
                    # Ignorar errores si la hoja no puede tener formato de tabla
                    continue

    def normalize_text(self, text):
        """Elimina tildes, normaliza espacios y símbolos"""
        text = str(text).strip()
        # Reemplazar variaciones de separadores
        text = text.replace(' / ', ' ').replace('/', ' ')
        # Eliminar tildes
        text = ''.join(c for c in unicodedata.normalize('NFD', text)
                    if unicodedata.category(c) != 'Mn')
        # Normalizar espacios
        text = ' '.join(text.split())
        return text

    def get_group_name(self, group):
        """
        Obtiene el nombre personalizado del grupo basado en sus criterios.
        Sigue la misma lógica que create_group_name en la clase CustomGroupResults.
        
        Args:
            group (dict): Diccionario con los datos del grupo incluyendo criteria_values
            
        Returns:
            str: Nombre del grupo basado en sus criterios
        """
        if 'criteria_values' not in group:
            return "Grupo sin criterios"

        # Procesar cada criterio y su valor
        name_parts = []
        for criterion, value in group['criteria_values'].items():
            # Normalizar el valor
            value = self.normalize_text(value)
            
            if criterion == "Clúster":
                # Para Clúster, si el valor es "-", dejar un espacio
                if value != "-":
                    name_parts.append(str(value))
            elif criterion == "Modulación":
                # Para Modulación, agregar "M" al final
                if value not in ["Sin datos", "No encontrado"]:
                    name_parts.append(f"{value}M")
            else:
                name_parts.append(str(value))

        # Unir con guiones evitando guiones duplicados o innecesarios
        group_name = ""
        for i, part in enumerate(name_parts):
            if i > 0:
                # Si la parte anterior o actual es solo un guión, no agregar guión adicional
                if part != "-" and name_parts[i-1] != "-":
                    group_name += "-"
            group_name += part

        return group_name

    def add_variation_analysis_sheets(self, writer):
        """
        Agrega hojas de análisis de variación para cada grupo final al archivo Excel.
        """
        try:
            # Obtener grupos finales
            final_groups = self.calculate_final_groups(self.current_plu_limit)
            
            # Leer específicamente Sheet1
            df = pd.read_excel(self.file_path_var.get(), sheet_name='Sheet1')
            
            # Validar columnas requeridas (CENTRO y PLU_SAP)
            required_columns = self.column_validator.validate_required_columns(
                df, ['CENTRO', 'PLU_SAP']
            )
            
            # Intentar encontrar columnas opcionales
            optional_columns = ['CATEGORIA', 'SUBCATEGORIA', 'SEGMENTO', 'ARTICULO']
            for col_type in optional_columns:
                try:
                    found_col = self.column_validator.find_column(df, col_type, raise_error=False)
                    if found_col:
                        required_columns[col_type] = found_col
                    else:
                        # Si no se encuentra la columna, crear una columna con '-'
                        df[col_type] = '-'
                        required_columns[col_type] = col_type
                except:
                    # Si hay cualquier error, crear una columna con '-'
                    df[col_type] = '-'
                    required_columns[col_type] = col_type

            # Limpiar y preparar los datos
            for col_name in required_columns.values():
                df[col_name] = df[col_name].fillna('-').astype(str).str.strip()
            
            # Procesar cada grupo final
            for group_num, group in enumerate(final_groups, 1):
                try:
                    sheet_name = f'Análisis GF. {group_num}'
                    
                    # Filtrar datos para el grupo actual
                    centers = [str(c) for c in group['centers']]
                    df_group = df[df[required_columns['CENTRO']].isin(centers)].copy()
                    
                    # Crear una columna auxiliar para el conteo
                    df_group['count'] = 1
                    
                    # Configurar las columnas para el pivot
                    index_columns = [
                        required_columns['CATEGORIA'],
                        required_columns['SUBCATEGORIA'],
                        required_columns['SEGMENTO'],
                        required_columns['PLU_SAP'],
                        required_columns['ARTICULO']
                    ]
                    
                    # Crear el pivot table
                    pivot = pd.pivot_table(
                        df_group,
                        index=index_columns,
                        columns=[required_columns['CENTRO']],
                        values='count',
                        aggfunc='sum',
                        fill_value=0
                    )
                    
                    # Convertir a 1s y 0s
                    pivot = (pivot > 0).astype(int)
                    
                    # Resetear índice
                    pivot = pivot.reset_index()
                    
                    # Renombrar columnas al formato esperado
                    column_mapping = {
                        required_columns['CATEGORIA']: 'Categoria',
                        required_columns['SUBCATEGORIA']: 'Subcategoria',
                        required_columns['SEGMENTO']: 'Segmento',
                        required_columns['PLU_SAP']: 'PLU_SAP',
                        required_columns['ARTICULO']: 'Articulo'
                    }
                    
                    # Renombrar columnas de índice
                    pivot = pivot.rename(columns=column_mapping)
                    
                    # Ordenar columnas
                    fixed_cols = ['Categoria', 'Subcategoria', 'Segmento', 'PLU_SAP', 'Articulo']
                    center_cols = sorted([col for col in pivot.columns if col not in fixed_cols])
                    pivot = pivot[fixed_cols + center_cols]
                    
                    # Escribir a Excel
                    pivot.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Obtener la hoja y aplicar formato
                    worksheet = writer.sheets[sheet_name]
                    
                    # Quitar las líneas de cuadrícula
                    worksheet.sheet_view.showGridLines = False

                    # Aplicar formato base a toda la hoja
                    worksheet.font = Font(name='Segoe UI', size=10)

                    # Formato para encabezados
                    header_font = Font(name='Segoe UI', size=11, bold=True)
                    header_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # Aplicar formato a los encabezados
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment

                    # Ajustar anchos de columna y aplicar formatos
                    for col in range(1, worksheet.max_column + 1):
                        column_letter = get_column_letter(col)
                        
                        # Obtener el máximo ancho necesario para la columna
                        max_length = 0
                        for row in range(1, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=col)
                            if cell.value:
                                try:
                                    max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass

                        # Ajustar ancho basado en el contenido
                        if worksheet.cell(row=1, column=col).value in ['Categoria', 'Subcategoria', 'Segmento', 'Articulo']:
                            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)
                        elif 'PLU' in str(worksheet.cell(row=1, column=col).value):
                            worksheet.column_dimensions[column_letter].width = 15
                        else:
                            worksheet.column_dimensions[column_letter].width = 8

                    # Aplicar formato a las celdas de datos
                    normal_font = Font(name='Segoe UI', size=10)
                    normal_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    center_alignment = Alignment(horizontal='center', vertical='center')

                    for row in worksheet.iter_rows(min_row=2):
                        for idx, cell in enumerate(row, 1):
                            cell.font = normal_font
                            
                            # Alineación basada en el tipo de columna
                            if idx <= 5:  # Primeras 5 columnas (categoría, subcategoría, etc.)
                                cell.alignment = normal_alignment
                            else:  # Columnas de centros
                                cell.alignment = center_alignment

                    # Congelar panel superior
                    worksheet.freeze_panes = 'A2'
                    
                    print(f"Se exportó el análisis del Grupo Final {group_num}")
                    
                except Exception as e:
                    print(f"Error al procesar Grupo Final {group_num}: {str(e)}")
                    continue
                        
        except Exception as e:
            print(f"Error al exportar análisis de variación: {str(e)}")
            raise

    def add_initial_sheet(self, writer):
        try:
            # Leer la hoja Sheet1 del archivo original
            df = pd.read_excel(self.file_path_var.get(), sheet_name='Sheet1')
            
            # Función para normalizar texto (eliminar tildes y mayúsculas)
            def normalize_text(text):
                """Elimina tildes y convierte a mayúsculas"""
                # Convertir a string por si acaso
                text = str(text)
                # Convertir a mayúsculas
                text = text.upper()
                # Normalizar caracteres (NFD descompone los caracteres con tilde)
                text = unicodedata.normalize('NFD', text)
                # Eliminar los caracteres diacríticos
                text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
                return text.strip()

            # Columnas exactas que queremos (sin variaciones)
            exact_columns = {
                'DIRECCION NAL.': ['Dirección Nal.', 'DIRECCION NAL', 'DIRECCION NACIONAL'],
                'CLUSTER': ['Cluster', 'CLUSTER'],
                'LLAVECP': ['LLAVECP', 'LLAVE CP', 'LLAVE_CP'],
                'MARCA': ['Marca', 'MARCA'],
                'CLASEMARCA': ['ClaseMarca', 'CLASE MARCA', 'CLASE_MARCA'],
                'VENTA 6 MESES': ['Venta 6 Meses', 'VENTA 6 MESES', 'VENTA_6_MESES']
            }
            
            # Columnas que tienen variaciones en ColumnValidator
            column_types = [
                'CENTRO', 'PLU_SAP', 'CATEGORIA', 'SUBCATEGORIA', 
                'SEGMENTO', 'ARTICULO'
            ]
            
            # Normalizar nombres de columnas del DataFrame
            df.columns = [str(col) for col in df.columns]  # Convertir a string
            normalized_columns = {normalize_text(col): col for col in df.columns}
            
            # Usar el validador para encontrar las columnas con variaciones
            found_columns = {}
            for col_type in column_types:
                try:
                    found_col = self.column_validator.find_column(df, col_type, raise_error=False)
                    if found_col:
                        found_columns[col_type] = found_col
                except Exception:
                    continue
            
            # Buscar columnas exactas normalizando nombres
            exact_found_columns = {}
            for col_key, variations in exact_columns.items():
                for variation in variations:
                    normalized_variation = normalize_text(variation)
                    if normalized_variation in normalized_columns.keys():
                        exact_found_columns[col_key] = normalized_columns[normalized_variation]
                        break
            
            # Crear lista de todas las columnas a mantener
            columns_to_keep = []
            
            # Agregar columnas encontradas por el validador
            columns_to_keep.extend(found_columns.values())
            
            # Agregar columnas exactas encontradas
            columns_to_keep.extend(exact_found_columns.values())
            
            # Verificar si tenemos al menos la columna Centro para poder agregar el grupo
            if 'CENTRO' not in found_columns:
                raise ValueError("La columna Centro es requerida para generar la hoja Sheet1")
            
            # Crear nuevo DataFrame con todas las columnas encontradas
            df_new = df[columns_to_keep].copy()
            
            # Obtener grupos finales
            final_groups = self.calculate_final_groups(self.current_plu_limit)
            
            # Crear diccionario de mapeo centro -> grupo
            centro_grupo_map = {}
            for group_num, group in enumerate(final_groups, 1):
                for center in group['centers']:
                    centro_grupo_map[str(center)] = f'Grupo Final {group_num}'
            
            # Insertar columna de Grupo al inicio
            centro_col = found_columns['CENTRO']
            df_new.insert(0, 'Grupo', df_new[centro_col].astype(str).map(centro_grupo_map).fillna('-'))
            
            # Escribir a Excel
            df_new.to_excel(writer, sheet_name='Sheet1', index=False)
            
            # Aplicar formato
            worksheet = writer.sheets['Sheet1']
            
            # Formato para encabezados
            header_format = Font(name='Segoe UI', size=11, bold=True)
            header_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            
            # Aplicar formato a encabezados
            for col in range(1, len(df_new.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_format
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Formato para datos y formato especial para Venta 6 Meses
            for row in range(2, len(df_new) + 2):
                for col in range(1, len(df_new.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = Font(name='Segoe UI', size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Aplicar formato de contabilidad a Venta 6 Meses
                    if normalize_text(df_new.columns[col-1]) == 'VENTA 6 MESES':
                        cell.number_format = '"$"#,##0_);("$"#,##0)'
            
            # Ajustar anchos de columna
            for col in worksheet.columns:
                max_length = 0
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)
            
            # Agregar bordes
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df_new) + 1):
                for cell in row:
                    cell.border = border
            
            # Congelar panel superior
            worksheet.freeze_panes = 'A2'
            
        except Exception as e:
            print(f"Error al crear hoja Sheet1: {str(e)}")
            raise

    def add_consolidated_sheet(self, writer):
        """
        Agrega una hoja consolidada con información detallada de cada centro.
        """
        try:
            # Cargar datos maestros
            maestro_file = self.get_data_path("db_maestrospdv.xlsx")
            
            if not os.path.exists(maestro_file):
                raise FileNotFoundError("No se encontró el archivo db_maestrospdv.xlsx")
                
            # Leer datos maestros
            df_maestro = pd.read_excel(maestro_file)
            
            # Obtener grupos finales
            final_groups = self.calculate_final_groups(self.current_plu_limit)
            
            # Crear DataFrame para la hoja consolidada
            consolidated_data = []
            
            # Procesar cada grupo final
            for group_num, group in enumerate(final_groups, 1):
                # Procesar cada centro del grupo
                for center in sorted(group['centers']):
                    # Buscar información del centro en datos maestros
                    centro_info = df_maestro[df_maestro['Centro'].astype(str).str.strip() == str(center)]
                    
                    # Obtener cantidad de PLUs para este centro
                    if center in self.unique_portfolios:
                        plus_count = len(self.unique_portfolios[center])
                    else:
                        for centers, plus in self.identical_portfolios.items():
                            if str(center) in [str(c) for c in centers]:
                                plus_count = len(plus)
                                break
                        else:
                            plus_count = 0
                    
                    # Agregar fila al consolidado
                    row = {
                        'Grupo': f'Grupo Final {group_num}',
                        'Centro': center,
                        'Formato': centro_info['Formato'].iloc[0] if not centro_info.empty else '-',
                        'Región': centro_info['Región'].iloc[0] if not centro_info.empty else '-',
                        'Ciudad': centro_info['Ciudad'].iloc[0] if not centro_info.empty else '-',
                        'Estrato': centro_info['Estrato'].iloc[0] if not centro_info.empty else '-',
                        'Cant. PLUs': plus_count
                    }
                    consolidated_data.append(row)
            
            # Agregar centros sin recomendación
            for center in sorted(self.non_compatible):
                # Buscar información del centro en datos maestros
                centro_info = df_maestro[df_maestro['Centro'].astype(str).str.strip() == str(center)]
                
                # Obtener cantidad de PLUs del centro
                plus_count = len(self.unique_portfolios[center]) if center in self.unique_portfolios else 0
                
                # Agregar fila al consolidado
                row = {
                    'Grupo': '-',
                    'Centro': center,
                    'Formato': centro_info['Formato'].iloc[0] if not centro_info.empty else '-',
                    'Región': centro_info['Región'].iloc[0] if not centro_info.empty else '-',
                    'Ciudad': centro_info['Ciudad'].iloc[0] if not centro_info.empty else '-',
                    'Estrato': centro_info['Estrato'].iloc[0] if not centro_info.empty else '-',
                    'Cant. PLUs': plus_count
                }
                consolidated_data.append(row)

            # Crear DataFrame
            df_consolidated = pd.DataFrame(consolidated_data)
            
            # Escribir a Excel con formato
            df_consolidated.to_excel(writer, sheet_name='Consolidado', index=False)
            
            # Obtener la hoja
            worksheet = writer.sheets['Consolidado']
            
            # Configurar formato
            header_format = {
                'font': Font(name='Segoe UI', size=11, bold=True),
                'fill': PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            }
            
            # Aplicar formato a encabezados
            for col in range(1, len(df_consolidated.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_format['font']
                cell.fill = header_format['fill']
                cell.alignment = header_format['alignment']
            
            # Formato para celdas de datos
            data_format = Font(name='Segoe UI', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Aplicar formato a datos
            for row in range(2, len(df_consolidated) + 2):
                for col in range(1, len(df_consolidated.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_format
                    
                    # Alineación específica por columna
                    if col in [1, 2, 7]:  # Grupo, Centro, Cant. PLUs
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment
            
            # Ajustar anchos de columna
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = min(adjusted_width, 40)
            
            # Agregar bordes a la tabla
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df_consolidated) + 1):
                for cell in row:
                    cell.border = border
            
            # Congelar panel superior
            worksheet.freeze_panes = 'A2'
            
        except Exception as e:
            print(f"Error al crear hoja consolidada: {str(e)}")
            raise

    def extract_group_info(self, group_frame):
        """
        Extrae la información de un grupo desde su frame.
        
        Args:
            group_frame: Frame que contiene la información del grupo
            
        Returns:
            dict: Diccionario con la información del grupo o None si no se puede extraer
        """
        try:
            group_info = {}
            
            # Buscar los labels que contienen la información
            for widget in group_frame.winfo_children():
                if isinstance(widget, ttk.Label):
                    text = widget.cget("text")
                    if "Centros:" in text:
                        # Extraer y limpiar la lista de centros
                        centers_text = text.split("Centros:")[1].strip()
                        centers = [c.strip() for c in centers_text.split(",")]
                        group_info['centers'] = centers
                        
                    # Añadir más extracciones según sea necesario
                    # Por ejemplo, criterios de agrupación si los hay
            
            return group_info if 'centers' in group_info else None
            
        except Exception as e:
            print(f"Error al extraer información del grupo: {str(e)}")
            return None

    def show_export_options(self):
        """
        Muestra una ventana de diálogo para seleccionar las hojas a exportar.
        """
        dialog = tk.Toplevel(self.root)
        dialog.title("Opciones de Exportación")
        dialog.grab_set()  # Hacer el diálogo modal
        
        # Centrar el diálogo
        window_width = 500
        window_height = 700
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Frame principal con padding y fondo blanco
        main_frame = ctk.CTkFrame(
            dialog,
            fg_color="white",
            corner_radius=10
        )
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Título
        ctk.CTkLabel(
            main_frame,
            text="Seleccione las hojas a exportar",
            font=("Segoe UI", 16, "bold"),
            text_color="#1e293b"
        ).pack(pady=(0, 20))
        
        # Variables para las opciones
        options = {
            'sheet1': tk.BooleanVar(value=True),
            'resumen': tk.BooleanVar(value=True),
            'grupos_identicos': tk.BooleanVar(value=True),
            'grupos_finales': tk.BooleanVar(value=True),
            'analisis_grupos': tk.BooleanVar(value=False),
            'distribucion_region': tk.BooleanVar(value=False),
            'analisis_modulacion': tk.BooleanVar(value=False),
            'centros_unicos': tk.BooleanVar(value=True),
            'centros_no_compatibles': tk.BooleanVar(value=True),
            'analisis_variacion': tk.BooleanVar(value=True),
            'consolidado': tk.BooleanVar(value=True),
            'agrupacion_personalizada': tk.BooleanVar(value=True),
            'agrupacion_personalizada_modulacion': tk.BooleanVar(value=True),
            'agrupacion_personalizada_variacion': tk.BooleanVar(value=True),
            'consolidado_personalizado': tk.BooleanVar(value=True)
        }

        # Frame para las opciones principales
        options_frame = ctk.CTkFrame(
            main_frame,
            fg_color="transparent"
        )
        options_frame.pack(fill=tk.BOTH, expand=True, padx=10)

        def update_suboptions_state():
            state = "normal" if options['analisis_grupos'].get() else "disabled"
            distribucion_check.configure(state=state)
            modulacion_check.configure(state=state)

        # Crear opciones
        ctk.CTkCheckBox(
            options_frame,
            text="Sheet1 (Datos originales)",
            variable=options['sheet1'],
            font=("Segoe UI", 12),
            text_color="#374151"
        ).pack(anchor="w", pady=5)

        ctk.CTkCheckBox(
            options_frame,
            text="Resumen General",
            variable=options['resumen'],
            font=("Segoe UI", 12),
            text_color="#374151"
        ).pack(anchor="w", pady=5)

        ctk.CTkCheckBox(
            options_frame,
            text="Grupos Idénticos",
            variable=options['grupos_identicos'],
            font=("Segoe UI", 12),
            text_color="#374151"
        ).pack(anchor="w", pady=5)

        ctk.CTkCheckBox(
            options_frame,
            text="Grupos Finales",
            variable=options['grupos_finales'],
            font=("Segoe UI", 12),
            text_color="#374151"
        ).pack(anchor="w", pady=5)

        # Grupo de análisis con subopciones
        analisis_check = ctk.CTkCheckBox(
            options_frame,
            text="Análisis de Grupos Finales",
            variable=options['analisis_grupos'],
            command=update_suboptions_state,
            font=("Segoe UI", 12),
            text_color="#374151"
        )
        analisis_check.pack(anchor="w", pady=5)

        # Frame para subopciones con padding
        sub_frame = ctk.CTkFrame(
            options_frame,
            fg_color="transparent"
        )
        sub_frame.pack(fill=tk.X, padx=(30, 0))

        distribucion_check = ctk.CTkCheckBox(
            sub_frame,
            text="Distribución por Región",
            variable=options['distribucion_region'],
            font=("Segoe UI", 12),
            text_color="#374151"
        )
        distribucion_check.pack(anchor="w", pady=2)

        modulacion_check = ctk.CTkCheckBox(
            sub_frame,
            text="Análisis de Modulación",
            variable=options['analisis_modulacion'],
            font=("Segoe UI", 12),
            text_color="#374151"
        )
        modulacion_check.pack(anchor="w", pady=2)

        # Resto de opciones
        ctk.CTkCheckBox(
            options_frame,
            text="Centros Únicos",
            variable=options['centros_unicos'],
            font=("Segoe UI", 12),
            text_color="#374151"
        ).pack(anchor="w", pady=5)

        ctk.CTkCheckBox(
            options_frame,
            text="Centros No Compatibles",
            variable=options['centros_no_compatibles'],
            font=("Segoe UI", 12),
            text_color="#374151"
        ).pack(anchor="w", pady=5)

        ctk.CTkCheckBox(
            options_frame,
            text="Análisis de Variación",
            variable=options['analisis_variacion'],
            font=("Segoe UI", 12),
            text_color="#374151"
        ).pack(anchor="w", pady=5)

        ctk.CTkCheckBox(
            options_frame,
            text="Consolidado",
            variable=options['consolidado'],
            font=("Segoe UI", 12),
            text_color="#374151"
        ).pack(anchor="w", pady=5)

        # Agrupación personalizada con subopciones
        agrupacion_personalizada_check = ctk.CTkCheckBox(
            options_frame,
            text="Agrupación Personalizada",
            variable=options['agrupacion_personalizada'],
            command=update_suboptions_state,
            font=("Segoe UI", 12),
            text_color="#374151"
        )
        agrupacion_personalizada_check.pack(anchor="w", pady=5)

        # Frame para subopciones de agrupación personalizada
        sub_frame_personalizada = ctk.CTkFrame(
            options_frame,
            fg_color="transparent"
        )
        sub_frame_personalizada.pack(fill=tk.X, padx=(30, 0))

        modulacion_personalizada_check = ctk.CTkCheckBox(
            sub_frame_personalizada,
            text="Análisis de Modulación",
            variable=options['agrupacion_personalizada_modulacion'],
            font=("Segoe UI", 12),
            text_color="#374151",
            state="disabled"
        )
        modulacion_personalizada_check.pack(anchor="w", pady=2)

        variacion_personalizada_check = ctk.CTkCheckBox(
            sub_frame_personalizada,
            text="Análisis de Variación",
            variable=options['agrupacion_personalizada_variacion'],
            font=("Segoe UI", 12),
            text_color="#374151",
            state="disabled"
        )
        variacion_personalizada_check.pack(anchor="w", pady=2)

        ctk.CTkCheckBox(
                sub_frame_personalizada,
                text="Consolidado II",
                variable=options['consolidado_personalizado'],
                font=("Segoe UI", 12),
                text_color="#374151",
                state="disabled"
            ).pack(anchor="w", pady=2)

        # Frame para botones
        button_frame = ctk.CTkFrame(
            main_frame,
            fg_color="transparent"
        )
        button_frame.pack(fill=tk.X, pady=(20, 0))

        # Variable para almacenar el resultado
        result = [None]

        def on_accept():
            result[0] = {k: v.get() for k, v in options.items()}
            dialog.destroy()

        def on_cancel():
            dialog.destroy()

        # Botones
        ctk.CTkButton(
            button_frame,
            text="Cancelar",
            command=on_cancel,
            fg_color="#f1f5f9",
            hover_color="#e2e8f0",
            text_color="#64748b",
            width=100
        ).pack(side=tk.RIGHT, padx=5)

        ctk.CTkButton(
            button_frame,
            text="Exportar",
            command=on_accept,
            width=100
        ).pack(side=tk.RIGHT, padx=5)

        # Actualizar estado inicial de subopciones
        update_suboptions_state()

        # Esperar hasta que se cierre el diálogo
        dialog.wait_window()

        return result[0]

    def add_custom_grouping_sheet(self, writer, groups, category=None):
        try:
            # Crear hoja para agrupación personalizada
            ws = writer.book.create_sheet("Agrupación Personalizada")
            current_row = 1

            # Título principal con formato
            cell = ws.cell(row=current_row, column=1, value="Análisis de Agrupación Personalizada")
            cell.font = Font(name='Segoe UI', size=14, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 2

            # Para cada grupo
            for group in groups:
                # Obtener nombre del grupo
                group_name = self.get_group_name(group)
                
                # Nombre del grupo
                cell = ws.cell(row=current_row, column=1, value=group_name)
                cell.font = Font(name='Segoe UI', size=12, bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                current_row += 1

                # Centros
                ws.cell(row=current_row, column=1, value="Centros:").font = Font(name='Segoe UI', size=11)
                current_row += 1
                # Lista de centros
                centers_str = ", ".join(sorted([str(center) for center in group['centers']]))
                ws.cell(row=current_row, column=1, value=centers_str).font = Font(name='Segoe UI', size=11)
                current_row += 1

                # Criterios de agrupación
                if 'criteria_values' in group:
                    ws.cell(row=current_row, column=1, value="Criterios de Agrupación:").font = Font(name='Segoe UI', size=11)
                    current_row += 1
                    
                    for criterion, value in group['criteria_values'].items():
                        cell_criterion = ws.cell(row=current_row, column=1, value=f"{criterion}:")
                        cell_criterion.font = Font(name='Segoe UI', size=11)
                        
                        cell_value = ws.cell(row=current_row, column=2, value=value)
                        cell_value.font = Font(name='Segui UI', size=11)
                        current_row += 1
                    current_row += 1

                # Análisis de Modulación
                if category:
                    modulation_data = self.get_modulation_data(group['centers'], category)
                    if modulation_data:
                        # Título de modulación
                        cell = ws.cell(row=current_row, column=1, value="Análisis de Modulación")
                        cell.font = Font(name='Segoe UI', size=11, bold=True)
                        current_row += 1

                        # Headers
                        headers = ["No. Módulos", "Cantidad", "Centros"]
                        for idx, header in enumerate(headers, 1):
                            cell = ws.cell(row=current_row, column=idx, value=header)
                            cell.font = Font(name='Segoe UI', size=11, bold=True)
                            cell.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        current_row += 1

                        # Datos de modulación
                        for modulos, data in modulation_data.items():
                            ws.cell(row=current_row, column=1, value=modulos).font = Font(name='Segoe UI', size=11)
                            ws.cell(row=current_row, column=2, value=data['count']).font = Font(name='Segoe UI', size=11)
                            ws.cell(row=current_row, column=3, value=", ".join(sorted([str(c) for c in data['centers']])))
                            current_row += 1
                        current_row += 1

                # Espacio entre grupos
                current_row += 1

            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    if cell.value:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = min(adjusted_width, 100)

            # Aplicar bordes a todas las celdas con contenido
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

        except Exception as e:
            print(f"Error al exportar agrupación personalizada: {str(e)}")
            raise

    def add_variation_analysis_for_group(self, writer, centers, sheet_name):
        """
        Añade una hoja de análisis de variación para un grupo personalizado.
        
        Args:
            writer: ExcelWriter objeto
            centers: Lista de centros del grupo
            sheet_name: Nombre de la hoja
        """
        try:
            # Leer el archivo Excel
            df = pd.read_excel(self.file_path_var.get(), sheet_name='Sheet1')
            
            # Validar columnas requeridas (CENTRO y PLU_SAP)
            required_columns = self.column_validator.validate_required_columns(
                df, ['CENTRO', 'PLU_SAP']
            )
            centro_col = required_columns['CENTRO']
            plu_col = required_columns['PLU_SAP']
            
            # Intentar encontrar columnas opcionales
            optional_columns = ['CATEGORIA', 'SUBCATEGORIA', 'SEGMENTO', 'ARTICULO']
            for col_type in optional_columns:
                try:
                    found_col = self.column_validator.find_column(df, col_type, raise_error=False)
                    if found_col:
                        required_columns[col_type] = found_col
                    else:
                        # Si no se encuentra la columna, crear una columna con '-'
                        df[col_type] = '-'
                        required_columns[col_type] = col_type
                except:
                    # Si hay cualquier error, crear una columna con '-'
                    df[col_type] = '-'
                    required_columns[col_type] = col_type
            
            # Filtrar datos para los centros del grupo
            centers = [str(c) for c in centers]  # Asegurar que los centros sean strings
            df_group = df[df[centro_col].astype(str).isin(centers)].copy()
            
            # Crear una columna auxiliar para el conteo
            df_group['count'] = 1
            
            # Configurar las columnas para el pivot
            index_columns = [
                required_columns['CATEGORIA'],
                required_columns['SUBCATEGORIA'],
                required_columns['SEGMENTO'],
                required_columns['PLU_SAP'],
                required_columns['ARTICULO']
            ]
            
            # Crear el pivot table
            pivot = pd.pivot_table(
                df_group,
                index=index_columns,
                columns=[centro_col],
                values='count',
                aggfunc='sum',
                fill_value=0
            )
            
            # Convertir a 1s y 0s
            pivot = (pivot > 0).astype(int)
            
            # Resetear índice
            pivot = pivot.reset_index()
            
            # Renombrar columnas al formato esperado
            column_mapping = {
                required_columns['CATEGORIA']: 'Categoria',
                required_columns['SUBCATEGORIA']: 'Subcategoria',
                required_columns['SEGMENTO']: 'Segmento',
                required_columns['PLU_SAP']: 'PLU_SAP',
                required_columns['ARTICULO']: 'Articulo'
            }
            
            # Renombrar columnas de índice
            pivot = pivot.rename(columns=column_mapping)
            
            # Ordenar columnas
            fixed_cols = ['Categoria', 'Subcategoria', 'Segmento', 'PLU_SAP', 'Articulo']
            center_cols = sorted([col for col in pivot.columns if col not in fixed_cols])
            pivot = pivot[fixed_cols + center_cols]
            
            # Escribir a Excel con formato
            pivot.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]            
            
            # Aplicar formato
            worksheet = writer.sheets[sheet_name]
            
            # Quitar líneas de cuadrícula
            worksheet.sheet_view.showGridLines = False
            
            # Formato base para toda la hoja
            worksheet.font = Font(name='Segoe UI', size=10)
            
            # Formato para encabezados
            header_font = Font(name='Segoe UI', size=11, bold=True)
            header_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            
            # Aplicar formato a los encabezados
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Ajustar ancho de columna según el contenido
                column_letter = get_column_letter(col)
                if worksheet.cell(row=1, column=col).value in ['Categoria', 'Subcategoria', 'Segmento', 'Articulo']:
                    worksheet.column_dimensions[column_letter].width = 40
                elif 'PLU' in str(worksheet.cell(row=1, column=col).value):
                    worksheet.column_dimensions[column_letter].width = 15
                else:
                    worksheet.column_dimensions[column_letter].width = 8
            
            # Formato para las celdas de datos
            for row in worksheet.iter_rows(min_row=2):
                for idx, cell in enumerate(row, 1):
                    cell.font = Font(name='Segoe UI', size=10)
                    if idx <= 5:  # Columnas de texto
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:  # Columnas de centros
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Congelar panel superior
            worksheet.freeze_panes = 'A2'

        except Exception as e:
            print(f"Error al procesar análisis de variación para grupo personalizado: {str(e)}")
            raise

    def add_custom_grouping_consolidated(self, writer, groups, category=None):
        """
        Agrega una hoja consolidada con información detallada de cada centro en grupos personalizados.
        
        Args:
            writer: ExcelWriter objeto
            groups: Lista de grupos personalizados
            category: Categoría opcional para el análisis de modulación
        """
        try:
            # Cargar datos maestros
            maestro_file = self.get_data_path("db_maestrospdv.xlsx")
            
            if not os.path.exists(maestro_file):
                raise FileNotFoundError("No se encontró el archivo db_maestrospdv.xlsx")
                
            # Leer datos maestros
            df_maestro = pd.read_excel(maestro_file)
            
            # Crear DataFrame para la hoja consolidada
            consolidated_data = []
            
            # Procesar cada grupo
            for group in groups:
                # Obtener nombre del grupo
                group_name = self.get_group_name(group)
                
                # Si hay análisis de modulación, obtener los datos
                modulation_data = None
                if category:
                    modulation_data = self.get_modulation_data(group['centers'], category)
                
                # Procesar cada centro del grupo
                for center in sorted(group['centers']):
                    # Buscar información del centro en datos maestros
                    centro_info = df_maestro[df_maestro['Centro'].astype(str).str.strip() == str(center)]
                    
                    # Obtener cantidad de PLUs para este centro
                    if center in self.unique_portfolios:
                        plus_count = len(self.unique_portfolios[center])
                    else:
                        for centers, plus in self.identical_portfolios.items():
                            if str(center) in [str(c) for c in centers]:
                                plus_count = len(plus)
                                break
                        else:
                            plus_count = 0
                    
                    # Obtener número de módulos si existe el análisis
                    num_modulos = '-'
                    if modulation_data:
                        for key, data in modulation_data.items():
                            if str(center) in [str(c) for c in data['centers']]:
                                num_modulos = key
                                break

                    # Crear fila de datos
                    row = {
                        'Grupo': group_name,
                        'Centro': center,
                        'Formato': centro_info['Formato'].iloc[0] if not centro_info.empty else '-',
                        'Región': centro_info['Región'].iloc[0] if not centro_info.empty else '-',
                        'Ciudad': centro_info['Ciudad'].iloc[0] if not centro_info.empty else '-',
                        'Estrato': centro_info['Estrato'].iloc[0] if not centro_info.empty else '-',
                        'Cant. PLUs': plus_count
                    }
                    
                    # Agregar número de módulos si hay análisis de modulación
                    if category:
                        row['No. Módulos'] = num_modulos
                    
                    consolidated_data.append(row)
            
            # Crear DataFrame
            df_consolidated = pd.DataFrame(consolidated_data)
            
            # Escribir a Excel con formato
            df_consolidated.to_excel(writer, sheet_name='Consolidado II', index=False)
            
            # Obtener la hoja
            worksheet = writer.sheets['Consolidado II']
            
            # Configurar formato
            header_format = {
                'font': Font(name='Segoe UI', size=11, bold=True),
                'fill': PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            }
            
            # Aplicar formato a encabezados
            for col in range(1, len(df_consolidated.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_format['font']
                cell.fill = header_format['fill']
                cell.alignment = header_format['alignment']
            
            # Formato para celdas de datos
            data_format = Font(name='Segoe UI', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Aplicar formato a datos
            for row in range(2, len(df_consolidated) + 2):
                for col in range(1, len(df_consolidated.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.font = data_format
                    
                    # Alineación específica por columna
                    if col in [2, 7, 8]:  # Centro, Cant. PLUs, No. Módulos
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment
            
            # Ajustar anchos de columna
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = min(adjusted_width, 40)
            
            # Agregar bordes a la tabla
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df_consolidated) + 1):
                for cell in row:
                    cell.border = border
            
            # Congelar panel superior
            worksheet.freeze_panes = 'A2'
            
        except Exception as e:
            print(f"Error al crear hoja consolidada II: {str(e)}")
            raise

    def export_to_excel(self):
        # Mostrar diálogo de opciones
        export_options = self.show_export_options()
        if not export_options:  # Usuario canceló
            return
        
        try:
            # Abrir diálogo para guardar archivo
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar archivo Excel"
            )
            
            if not file_path:
                return

            # Validar si el archivo existe
            if os.path.exists(file_path):
                if not messagebox.askyesno(
                    "Archivo existente",
                    "El archivo ya existe. ¿Desea sobrescribirlo?"
                ):
                    return
                    
            # Mostrar diálogo de carga
            loading_window = self.show_loading_spinner("Preparando exportación...")
            
            # Crear un Excel writer
            writer = pd.ExcelWriter(file_path, engine='openpyxl')

            # Cargar datos según las opciones seleccionadas
            self.update_loading_message(loading_window, "Cargando datos necesarios...")
            
            # Cargar datos básicos que siempre se necesitan
            df = pd.read_excel(self.file_path_var.get())
            if df.empty:
                raise ValueError("El archivo Excel está vacío")
                
            final_groups = self.calculate_final_groups(self.current_plu_limit)
            
            # Cargar datos geográficos solo si son necesarios
            geo_data = None
            if export_options['analisis_grupos'] and export_options['distribucion_region']:
                geo_data = self.load_geographic_data()
                
            # Preguntar categoría solo si se necesita el análisis de modulación
            category = None
            if export_options['analisis_grupos'] and export_options['analisis_modulacion']:
                category = self.get_category_input()

            # Configurar estilos base
            base_font = Font(name='Segoe UI', size=11)
            header_font = Font(name='Segoe UI', size=11, bold=True)
            title_font = Font(name='Segoe UI', size=14, bold=True)
            subtitle_font = Font(name='Segoe UI', size=12, bold=True)
            
            # Alineaciones
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Colores
            header_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
            warning_fill = PatternFill(start_color='FEF2F2', end_color='FEF2F2', fill_type='solid')
            success_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
            alternate_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            
            # 1. Resumen General
            if export_options['resumen']:
                try:
                    self.update_loading_message(loading_window, "Generando resumen general...")
                    ws_summary = writer.book.create_sheet("Resumen General", 0)
                    current_row = 1

                    # Estadísticas generales
                    current_row = self.create_title_row(ws_summary, "Estadísticas Generales")
                    
                    stats_headers = ["Métrica", "Valor", "Porcentaje"]
                    current_row = self.add_table_headers(ws_summary, stats_headers, current_row)

                    # Calcular estadísticas
                    total_centers = int(self.total_centers.get())
                    unique_centers = int(self.unique_centers.get())
                    identical_centers = total_centers - unique_centers
                    initial_masters = len(self.identical_portfolios) + len(self.unique_portfolios)
                    final_masters = len(final_groups) + len(self.non_compatible)
                    optimization_percentage = ((initial_masters - final_masters) / initial_masters * 100 
                                            if initial_masters > 0 else 0)
                    
                    stats_data = [
                        ("Total de Centros", total_centers, "100%"),
                        ("Centros en Grupos", identical_centers, f"{(identical_centers/total_centers)*100:.1f}%"),
                        ("Centros Únicos", unique_centers, f"{(unique_centers/total_centers)*100:.1f}%"),
                        ("Grupos Idénticos Iniciales", len(self.identical_portfolios), "-"),
                        ("Planogramas Másteres Iniciales", initial_masters, "-"),
                        ("Planogramas Másteres Finales", final_masters, "-"),
                        ("Optimización de Planogramas", final_masters - initial_masters, f"{optimization_percentage:.1f}%"),
                        ("Centros sin Recomendación", len(self.non_compatible), 
                        f"{(len(self.non_compatible)/total_centers)*100:.1f}%")
                    ]

                    for stat in stats_data:
                        for col, value in enumerate(stat, 1):
                            cell = ws_summary.cell(row=current_row, column=col, value=value)
                            cell.font = base_font
                            cell.alignment = center_alignment if col > 1 else left_alignment
                        current_row += 1

                except Exception as e:
                    print(f"Error al crear resumen general: {str(e)}")


            # 2. Grupos Idénticos
            if export_options['grupos_identicos']:
                try:
                    self.update_loading_message(loading_window, "Procesando grupos idénticos...")
                    ws_identical = writer.book.create_sheet("Grupos Idénticos")
                    current_row = 1

                    current_row = self.create_title_row(ws_identical, "Grupos con Portafolios Idénticos")
                    
                    # Información detallada de grupos
                    headers = ["Grupo", "Centros", "Cantidad de Centros", "PLUs", "Cantidad de PLUs"]
                    current_row = self.add_table_headers(ws_identical, headers, current_row)

                    for i, (centers, plus) in enumerate(self.identical_portfolios.items(), 1):
                        row_data = [
                            f"Grupo {i}",
                            ", ".join(sorted(centers)),
                            len(centers),
                            ", ".join(map(str, sorted(plus))),
                            len(plus)
                        ]
                        
                        for col, value in enumerate(row_data, 1):
                            cell = ws_identical.cell(row=current_row, column=col, value=value)
                            cell.font = base_font
                            cell.alignment = left_alignment if col in [2, 4] else center_alignment
                        current_row += 1

                except Exception as e:
                    print(f"Error al procesar grupos idénticos: {str(e)}")

            # 3. Grupos Finales
            if export_options['grupos_finales']:
                try:
                    self.update_loading_message(loading_window, "Generando información de grupos finales...")
                    ws_final = writer.book.create_sheet("Grupos Finales")
                    current_row = 1

                    current_row = self.create_title_row(ws_final, "Grupos Finales Optimizados")
                    
                    headers = ["Grupo", "Centros", "Cantidad de Centros", "PLUs Diferentes", 
                            "Cantidad PLUs Dif.", "Total PLUs"]
                    current_row = self.add_table_headers(ws_final, headers, current_row)

                    if export_options['analisis_grupos']:
                        for i, group in enumerate(final_groups, 1):
                            try:
                                diff_plus, all_plus = self.calculate_total_different_plus(group['centers'])
                                
                                row_data = [
                                    f"Grupo Final {i}",
                                    ", ".join(sorted(group['centers'])),
                                    len(group['centers']),
                                    ", ".join(map(str, sorted(diff_plus))),
                                    len(diff_plus),
                                    len(all_plus)
                                ]
                                
                                for col, value in enumerate(row_data, 1):
                                    cell = ws_final.cell(row=current_row, column=col, value=value)
                                    cell.font = base_font
                                    cell.alignment = left_alignment if col in [2, 4] else center_alignment
                                current_row += 1

                                # Crear hoja individual para cada grupo final
                                self.update_loading_message(loading_window, f"Procesando Grupo Final {i}...")
                                ws_group = writer.book.create_sheet(f"Grupo Final {i}")

                                # En cada hoja de grupo final, después de crear la hoja:
                                self.configure_group_final_sheet(ws_group)

                                group_row = 1  # Inicializar group_row

                                # 1. Información básica del grupo
                                group_row = self.create_title_row(
                                    ws_group, 
                                    f"Grupo Final {i}", 
                                    row=1,
                                    subtitle=f"Total centros: {len(group['centers'])}"
                                )

                                # Lista de centros
                                headers = ["Centros", "PLUs Diferentes", "Total PLUs"]
                                group_row = self.add_table_headers(ws_group, headers, group_row)

                                # Datos básicos del grupo
                                group_data = [
                                    ", ".join(sorted(group['centers'])),
                                    ", ".join(map(str, sorted(diff_plus))),
                                    len(all_plus)
                                ]

                                for col, value in enumerate(group_data, 1):
                                    cell = ws_group.cell(row=group_row, column=col, value=value)
                                    cell.font = base_font
                                    cell.alignment = left_alignment
                                group_row += 2

                                # 2. Resumen por región
                                if export_options['distribucion_region'] and geo_data is not None:
                                    group_row = self.create_title_row(ws_group, "Distribución por Región", group_row)
                                    
                                    # Obtener datos del grupo
                                    group_geo_data = geo_data[geo_data['Centro'].isin(group['centers'])]
                                    
                                    # Análisis por distrito
                                    district_headers = ["Distrito", "Región", "Cantidad", "Porcentaje", "Centros"]
                                    group_row = self.add_table_headers(ws_group, district_headers, group_row)
                                    
                                    district_colors = {
                                        'COSTA': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
                                        'INTERIOR': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                                    }

                                    for district in group_geo_data['Distrito'].unique():
                                        district_data = group_geo_data[group_geo_data['Distrito'] == district]
                                        
                                        for region in district_data['Region'].unique():
                                            region_data = district_data[district_data['Region'] == region]
                                            centers = region_data['Centro'].tolist()
                                            percentage = (len(centers) / len(group['centers'])) * 100
                                            
                                            row_data = [
                                                district,
                                                region,
                                                len(centers),
                                                f"{percentage:.1f}%",
                                                ", ".join(sorted(centers))
                                            ]
                                            
                                            for col, value in enumerate(row_data, 1):
                                                cell = ws_group.cell(row=group_row, column=col, value=value)
                                                cell.font = base_font
                                                cell.alignment = left_alignment if col == 5 else center_alignment
                                                cell.fill = district_colors.get(district, alternate_fill)
                                            group_row += 1
                                    group_row += 2

                                    # 3. Análisis de modulación
                                    if export_options['analisis_modulacion'] and category:  # Si se seleccionó una categoría para el análisis
                                        group_row = self.create_title_row(
                                            ws_group,
                                            "Análisis de Modulación",
                                            group_row,
                                            subtitle=f"Categoría: {category}"
                                        )
                                        
                                        modulation_data = self.get_modulation_data(group['centers'], category)
                                        if modulation_data:
                                            headers = ["No. Módulos", "Cantidad de Centros", "Porcentaje", "Centros"]
                                            group_row = self.add_table_headers(ws_group, headers, group_row)
                                            
                                            total_centers = sum(data['count'] for data in modulation_data.values())
                                            
                                            for num_modulos, data in modulation_data.items():
                                                percentage = (data['count'] / total_centers) * 100
                                                row_data = [
                                                    num_modulos,
                                                    data['count'],
                                                    f"{percentage:.1f}%",
                                                    ", ".join(sorted(data['centers']))
                                                ]
                                                
                                                for col, value in enumerate(row_data, 1):
                                                    cell = ws_group.cell(row=group_row, column=col, value=value)
                                                    cell.font = base_font
                                                    cell.alignment = left_alignment if col == 4 else center_alignment
                                                    
                                                    # Colorear según el tipo de dato
                                                    if num_modulos == "Datos vacíos":
                                                        cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                                                    elif num_modulos == "No encontrados":
                                                        cell.fill = warning_fill
                                                group_row += 1
                                        else:
                                            cell = ws_group.cell(
                                                row=group_row,
                                                column=1,
                                                value="No se encontraron datos de modulación para este grupo"
                                            )
                                            cell.font = base_font
                                            cell.fill = warning_fill
                                            group_row += 1

                                    # Ajustar columnas al final
                                    for column in ws_group.columns:
                                        max_length = 0
                                        column_letter = get_column_letter(column[0].column)
                                        
                                        for cell in column:
                                            try:
                                                if len(str(cell.value)) > max_length:
                                                    max_length = len(str(cell.value))
                                            except:
                                                pass
                                        
                                        adjusted_width = (max_length + 2)
                                        ws_group.column_dimensions[column_letter].width = min(adjusted_width, 100)

                                    # Ajustar altura de filas
                                    for row in ws_group.rows:
                                        max_height = 0
                                        for cell in row:
                                            if cell.value:
                                                text_lines = str(cell.value).count('\n') + 1
                                                line_height = 15  # altura aproximada por línea
                                                needed_height = text_lines * line_height
                                                max_height = max(max_height, needed_height)
                                        
                                        if max_height > 15:  # altura mínima
                                            ws_group.row_dimensions[row[0].row].height = max_height

                            except Exception as e:
                                 print(f"Error al procesar Grupo Final {i}: {str(e)}")

                    self.update_loading_message(loading_window, "Creando hojas adicionales...")

                except Exception as e:
                    print(f"Error al generar grupos finales: {str(e)}")


            # Centros Únicos
            if export_options['centros_unicos']:
                try:
                    self.update_loading_message(loading_window, "Procesando centros únicos...")
                    unique_centers_data = [
                        (center, plus_set) 
                        for center, plus_set in self.unique_portfolios.items()
                        if center not in self.non_compatible
                    ]
                    self.create_centers_sheet(
                        writer.book,
                        "Centros únicos",
                        unique_centers_data,
                        base_font,
                        header_font,
                        header_fill
                    )
                except Exception as e:
                    print(f"Error al procesar centros únicos: {str(e)}")

            # Centros No Compatibles
            if export_options['centros_no_compatibles']:
                try:
                    self.update_loading_message(loading_window, "Procesando centros no compatibles...")
                    non_compatible_data = [
                        (center, self.unique_portfolios[center])
                        for center in sorted(self.non_compatible)
                    ]
                    self.create_centers_sheet(
                        writer.book,
                        "Centros no compatibles",
                        non_compatible_data,
                        base_font,
                        header_font,
                        header_fill
                    )
                except Exception as e:
                    print(f"Error al procesar centros no compatibles: {str(e)}")

            # Ajustar formato de todas las hojas
            self.update_loading_message(loading_window, "Aplicando formato final...")
            self.format_pivot_tables(writer.book)

            # Análisis de Variación
            if export_options['analisis_variacion']:
                try:
                    self.update_loading_message(loading_window, "Exportando análisis de variación...")
                    self.add_variation_analysis_sheets(writer)
                except Exception as e:
                    print(f"Error al agregar hojas de análisis de variación: {str(e)}")

            # Consolidado
            if export_options['consolidado']:
                try:
                    self.update_loading_message(loading_window, "Generando hoja consolidada...")
                    self.add_consolidated_sheet(writer)
                except Exception as e:
                    print(f"Error al agregar hoja consolidada: {str(e)}")

            # Sheet1 (Datos originales)
            if export_options['sheet1']:
                try:
                    self.update_loading_message(loading_window, "Copiando datos originales...")
                    self.add_initial_sheet(writer)
                except Exception as e:
                    print(f"Error al copiar datos originales: {str(e)}")

            # agrupación personalizada
            if export_options['agrupacion_personalizada']:
                try:
                    # Verificar si existen grupos personalizados
                    if hasattr(self, 'custom_groups') and self.custom_groups:
                        self.update_loading_message(loading_window, "Exportando agrupación personalizada...")
                        
                        # Análisis de modulación solo si está seleccionado
                        category = None
                        if export_options['agrupacion_personalizada_modulacion']:
                            category = self.get_category_input()
                        
                        # Exportar grupos personalizados
                        self.add_custom_grouping_sheet(writer, self.custom_groups, category)
                        
                        # Consolidado II solo si está seleccionado
                        if export_options['consolidado_personalizado']:
                            self.update_loading_message(loading_window, "Generando consolidado de grupos personalizados...")
                            self.add_custom_grouping_consolidated(writer, self.custom_groups, category)
                        
                        # Análisis de variación solo si está seleccionado
                        if export_options['agrupacion_personalizada_variacion']:
                            self.update_loading_message(loading_window, 
                                "Exportando análisis de variación para grupos personalizados...")
                            for i, group in enumerate(self.custom_groups, 1):
                                group_name = self.get_group_name(group)
                                sheet_name = f'Var. {group_name}'
                                self.add_variation_analysis_for_group(writer, group['centers'], sheet_name)
                                
                    else:
                        messagebox.showwarning("Aviso", "No se ha realizado una agrupación personalizada.")
                        
                except Exception as e:
                    print(f"Error al exportar agrupación personalizada: {str(e)}")

            # Finalizar exportación
            writer.close()
            
            # Verificar archivo y mostrar mensajes
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                if file_size > 0:
                    try:
                        pd.ExcelFile(file_path)
                        loading_window.destroy()
                        messagebox.showinfo(
                            "Exportación Exitosa",
                            f"Archivo Excel exportado correctamente\n"
                            f"Ubicación: {file_path}\n"
                            f"Tamaño: {file_size/1024:.1f} KB"
                        )
                        if messagebox.askyesno("Abrir Archivo", "¿Desea abrir el archivo exportado?"):
                            os.startfile(file_path)
                    except Exception as e:
                        loading_window.destroy()
                        messagebox.showerror("Error", f"El archivo se creó pero podría estar corrupto: {str(e)}")
                else:
                    loading_window.destroy()
                    messagebox.showerror("Error", "El archivo se creó pero está vacío")
            else:
                loading_window.destroy()
                messagebox.showerror("Error", "No se pudo crear el archivo")
                    
        except Exception as e:
            if 'loading_window' in locals():
                loading_window.destroy()
            messagebox.showerror("Error", f"Error al exportar el archivo: {str(e)}")

    def create_results_tabs(self, parent):
            self.notebook = ttk.Notebook(parent, style="Custom.TNotebook")
            self.notebook.pack(fill=tk.BOTH, expand=True)
            
            # Pestaña de portafolios idénticos
            identical_frame = ttk.Frame(self.notebook, style="Card.TFrame")
            self.notebook.add(identical_frame, text="Portafolios Idénticos")
            identical_frame.pack_propagate(False)
            
            # Canvas con scrollbar para los grupos
            canvas = tk.Canvas(identical_frame, bg='white', highlightthickness=0)
            scrollbar = ttk.Scrollbar(identical_frame, orient="vertical", command=canvas.yview)
            
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            # Frame para el grid de grupos
            self.groups_grid = ttk.Frame(canvas, style="Card.TFrame")
            canvas_window = canvas.create_window((0, 0), window=self.groups_grid, anchor='nw')
            
            # Configurar scroll para esta pestaña
            self.bind_scrolling(canvas, scrollbar)
            
            # Configurar el grid con 3 columnas
            for i in range(3):
                self.groups_grid.columnconfigure(i, weight=1)
            
            canvas_window = canvas.create_window((0, 0), window=self.groups_grid, anchor='nw')
            
            # Pestaña de portafolios únicos
            unique_frame = ttk.Frame(self.notebook, style="Card.TFrame")
            self.notebook.add(unique_frame, text="Portafolios Únicos")
            unique_frame.pack_propagate(False)
            
            unique_canvas = tk.Canvas(unique_frame, bg='white', highlightthickness=0)
            unique_scrollbar = ttk.Scrollbar(unique_frame, orient="vertical", command=unique_canvas.yview)
            
            unique_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            unique_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            # Crear unique_grid
            self.unique_grid = ttk.Frame(unique_canvas, style="Card.TFrame")
            unique_canvas_window = unique_canvas.create_window((0, 0), window=self.unique_grid, anchor='nw')

            # Configurar scroll para esta pestaña
            self.bind_scrolling(unique_canvas, unique_scrollbar)
            
            # Configurar columnas para unique_grid
            for i in range(3):
                self.unique_grid.columnconfigure(i, weight=1)
            
            unique_canvas_window = unique_canvas.create_window((0, 0), window=self.unique_grid, anchor='nw')
            
            # Pestaña de recomendaciones
            recommendations_frame = ttk.Frame(self.notebook, style="Card.TFrame")
            self.notebook.add(recommendations_frame, text="Recomendaciones")
            recommendations_frame.pack_propagate(False)
            
            recommendations_canvas = tk.Canvas(recommendations_frame, bg='white', highlightthickness=0)
            recommendations_scrollbar = ttk.Scrollbar(recommendations_frame, orient="vertical", command=recommendations_canvas.yview)
            
            recommendations_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            recommendations_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            # Crear recommendations_grid
            self.recommendations_grid = ttk.Frame(recommendations_canvas, style="Card.TFrame")
            recommendations_window = recommendations_canvas.create_window((0, 0), window=self.recommendations_grid, anchor='nw')
            recommendations_canvas.configure(yscrollcommand=recommendations_scrollbar.set)
            self.bind_scrolling(recommendations_canvas, recommendations_scrollbar)
            
            # Configurar columnas para recommendations_grid
            for i in range(3):
                self.recommendations_grid.columnconfigure(i, weight=1)
            
            recommendations_canvas_window = recommendations_canvas.create_window((0, 0), window=self.recommendations_grid, anchor='nw')
            
            # Pestaña de Recomendaciones II
            group_recommendations_frame = ttk.Frame(self.notebook, style="Card.TFrame")
            self.notebook.add(group_recommendations_frame, text="Recomendaciones II")
            group_recommendations_frame.pack_propagate(False)
            
            group_recommendations_canvas = tk.Canvas(group_recommendations_frame, bg='white', highlightthickness=0)
            group_recommendations_scrollbar = ttk.Scrollbar(group_recommendations_frame, orient="vertical", command=group_recommendations_canvas.yview)
            
            group_recommendations_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            group_recommendations_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            # Crear group_recommendations_grid
            self.group_recommendations_grid = ttk.Frame(group_recommendations_canvas, style="Card.TFrame")
            group_recommendations_window = group_recommendations_canvas.create_window((0, 0), window=self.group_recommendations_grid, anchor='nw')
            group_recommendations_canvas.configure(yscrollcommand=group_recommendations_scrollbar.set)
            self.bind_scrolling(group_recommendations_canvas, group_recommendations_scrollbar)
            
            # Configurar columnas para group_recommendations_grid
            for i in range(3):
                self.group_recommendations_grid.columnconfigure(i, weight=1)
            
            group_recommendations_canvas_window = group_recommendations_canvas.create_window((0, 0), window=self.group_recommendations_grid, anchor='nw')

            # Pestaña de Resumen
            summary_frame = ttk.Frame(self.notebook, style="Card.TFrame")
            self.notebook.add(summary_frame, text="Resumen")
            summary_frame.pack_propagate(False)
            
            summary_canvas = tk.Canvas(summary_frame, bg='white', highlightthickness=0)
            summary_scrollbar = ttk.Scrollbar(summary_frame, orient="vertical", command=summary_canvas.yview)
            
            summary_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            summary_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            # Crear summary_grid
            self.summary_grid = ttk.Frame(summary_canvas, style="Card.TFrame")
            summary_window = summary_canvas.create_window((0, 0), window=self.summary_grid, anchor='nw')
            summary_canvas.configure(yscrollcommand=summary_scrollbar.set)
            self.bind_scrolling(summary_canvas, summary_scrollbar)
            
            # Configurar columnas para summary_grid
            for i in range(3):
                self.summary_grid.columnconfigure(i, weight=1)
            
            summary_canvas_window = summary_canvas.create_window((0, 0), window=self.summary_grid, anchor='nw')
            
            # Configurar los eventos de scroll y redimensionamiento para todos los canvas
            def configure_canvas_events(canvas, canvas_window, frame):
                def _on_frame_configure(event=None):
                    canvas.configure(scrollregion=canvas.bbox("all"))
                    canvas.itemconfig(canvas_window, width=canvas.winfo_width())
                    
                def _on_canvas_configure(event=None):
                    canvas.itemconfig(canvas_window, width=event.width)
                    _on_frame_configure()
                    
                frame.bind("<Configure>", _on_frame_configure)
                canvas.bind("<Configure>", _on_canvas_configure)
            
            # Aplicar configuración a todos los canvas
            configure_canvas_events(canvas, canvas_window, self.groups_grid)
            configure_canvas_events(unique_canvas, unique_canvas_window, self.unique_grid)
            configure_canvas_events(recommendations_canvas, recommendations_canvas_window, self.recommendations_grid)
            configure_canvas_events(group_recommendations_canvas, group_recommendations_canvas_window, self.group_recommendations_grid)
            configure_canvas_events(summary_canvas, summary_canvas_window, self.summary_grid)
            
            # Configurar el comportamiento del scrollbar para todos los canvas
            def _on_mousewheel(event):
                current_tab = self.notebook.select()
                tab_name = self.notebook.tab(current_tab, "text")
                
                if tab_name == "Portafolios Idénticos":
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                elif tab_name == "Portafolios Únicos":
                    unique_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                elif tab_name == "Recomendaciones":
                    recommendations_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                elif tab_name == "Recomendaciones II":
                    group_recommendations_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                elif tab_name == "Resumen":
                    summary_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            
            # Bind mousewheel event to all frames
            identical_frame.bind("<MouseWheel>", _on_mousewheel)
            unique_frame.bind("<MouseWheel>", _on_mousewheel)
            recommendations_frame.bind("<MouseWheel>", _on_mousewheel)
            group_recommendations_frame.bind("<MouseWheel>", _on_mousewheel)
            summary_frame.bind("<MouseWheel>", _on_mousewheel)
            
            # Configurar yscrollcommand para todos los canvas
            canvas.configure(yscrollcommand=scrollbar.set)
            unique_canvas.configure(yscrollcommand=unique_scrollbar.set)
            recommendations_canvas.configure(yscrollcommand=recommendations_scrollbar.set)
            group_recommendations_canvas.configure(yscrollcommand=group_recommendations_scrollbar.set)
            summary_canvas.configure(yscrollcommand=summary_scrollbar.set)
            
            return self.notebook

    def bind_scrolling(self, canvas, scrollbar):
        """
        Vincula los eventos de scroll del mouse a un canvas.
        
        Args:
            canvas: El canvas que recibirá los eventos de scroll
            scrollbar: La barra de desplazamiento asociada
        """
        def on_mousewheel(event):
            # Obtener el widget actual bajo el cursor
            widget = event.widget
            
            # Verificar si el canvas es visible y activo
            if not canvas.winfo_viewable():
                return
                
            # Determinar la dirección del scroll
            if event.delta:
                # Para Windows/MacOS
                delta = -1 * (event.delta // 120)
            else:
                # Para Linux
                if event.num == 4:
                    delta = -1
                else:
                    delta = 1
            
            # Realizar el scroll
            canvas.yview_scroll(delta, "units")
        
        # Vincular eventos de scroll
        canvas.bind_all("<MouseWheel>", on_mousewheel)  # Windows/MacOS
        canvas.bind_all("<Button-4>", on_mousewheel)    # Linux scroll up
        canvas.bind_all("<Button-5>", on_mousewheel)    # Linux scroll down
        
        # Desvinculación automática cuando el canvas no está visible
        def on_tab_change(event):
            if not canvas.winfo_viewable():
                canvas.unbind_all("<MouseWheel>")
                canvas.unbind_all("<Button-4>")
                canvas.unbind_all("<Button-5>")
            else:
                canvas.bind_all("<MouseWheel>", on_mousewheel)
                canvas.bind_all("<Button-4>", on_mousewheel)
                canvas.bind_all("<Button-5>", on_mousewheel)
        
        # Vincular al evento de cambio de pestaña
        canvas.bind("<Visibility>", on_tab_change)

    def update_ui(self):
        # Actualizar solo cada 100ms
        if not hasattr(self, '_last_update'):
            self._last_update = 0
        
        current_time = time.time() * 1000
        if current_time - self._last_update > 100:
            self.root.update_idletasks()
            self._last_update = current_time

    def load_items_lazily(self, parent_frame, items, create_item_func):
        BATCH_SIZE = 10
        current_batch = 0
        
        def load_next_batch():
            nonlocal current_batch
            start_idx = current_batch * BATCH_SIZE
            end_idx = start_idx + BATCH_SIZE
            batch_items = items[start_idx:end_idx]
            
            for i, item in enumerate(batch_items, start=start_idx):
                row = i // 3
                col = i % 3
                frame = create_item_func(parent_frame, *item)
                frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                
            current_batch += 1
            
            # Si quedan más items, programar la siguiente carga
            if end_idx < len(items):
                parent_frame.after(10, load_next_batch)
        
        load_next_batch()

    def create_group_merger_frame(self, parent, groups_list, different_plus):
        # Crear frame principal para la tarjeta
        group_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Crear el frame CustomTkinter para el contenido
        content_frame = ctk.CTkFrame(
            group_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Frame para el header
        header_frame = ttk.Frame(content_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))
        
        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)
        
        # Crear un Canvas para el ícono SVG
        icon_size = 24
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg='white',
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(10, 5), pady=10)
        
        def draw_icon(is_expanded):
            icon_canvas.delete("all")
            # Crear un botón circular como fondo
            icon_canvas.create_oval(
                0, 0, icon_size, icon_size,
                fill="#4F46E5",  # Color púrpura para grupos fusionados
                outline=""
            )
            # Dibujar el ícono
            if is_expanded:
                icon_canvas.create_line(
                    8, 10, 12, 14, 16, 10,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
            else:
                icon_canvas.create_line(
                    10, 8, 14, 12, 10, 16,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
        
        # Contenedor para el texto
        text_container = ttk.Frame(header_frame, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))
        
        # Lista de variables matemáticas para usar en los títulos
        math_vars = ['X', 'Y', 'Z', 'α', 'β', 'γ', 'δ', 'ε', 'ζ', 'η', 'θ', 'ι', 'κ', 'λ', 'μ', 'ν', 'ξ', 'π', 'ρ', 'σ', 'τ', 'υ', 'φ', 'χ', 'ψ', 'ω']
        
        # Obtener los números de grupo
        groups = list(self.identical_portfolios.keys())
        group_numbers = []
        for group in groups_list:
            try:
                group_numbers.append(str(groups.index(group) + 1))
            except ValueError:
                continue
        
        # Asignar variable matemática basada en el índice del grupo
        if not hasattr(self, 'math_var_counter'):
            self.math_var_counter = 0
        
        math_var = math_vars[self.math_var_counter % len(math_vars)]
        self.math_var_counter += 1
        
        # Título
        title_label = ttk.Label(
            text_container,
            text=f"Grupos {math_var}",
            font=('Segoe UI', 14, 'bold'),
            foreground='#4F46E5',  # Color púrpura para grupos fusionados
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Calcular el máximo de PLUs diferentes
        centers_plus_dict = {}
        for centers in groups_list:
            centers_str = str(centers)  # Convertir a string para usar como clave
            plus_set = self.identical_portfolios[centers]
            centers_plus_dict[centers_str] = plus_set
        
        max_missing, _, _ = self.calculate_plu_differences_multi(centers_plus_dict)
        
        # Estadísticas actualizadas
        stats_text = f"Max. PLUs dif: {max_missing}"
        stats_label = ttk.Label(
            text_container,
            text=stats_text,
            font=('Segoe UI', 12),
            foreground='#6b7280',
            background='white'
        )
        stats_label.pack(anchor='w')
        
        # Frame para el contenido expandible
        content_detail_frame = ttk.Frame(content_frame, style="ContentCard.TFrame")
        
        # Mostrar los grupos a unir
        groups_text = f"Grupos a unir: {', '.join(group_numbers)}"
        groups_label = ttk.Label(
            content_detail_frame,
            text=groups_text,
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        groups_label.pack(anchor='w', padx=15, pady=(10, 5))
        
        # Mostrar los PLUs diferentes
        plues_label = ttk.Label(
            content_detail_frame,
            text=f"PLUs diferentes: {', '.join(map(str, sorted(different_plus)))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        plues_label.pack(anchor='w', padx=15, pady=(0, 10))
        
        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                content_detail_frame.pack(fill=tk.X)
            else:
                content_detail_frame.pack_forget()
            draw_icon(is_expanded.get())
        
        # Hacer que todo el header sea clickeable
        header_frame.bind("<Button-1>", lambda e: toggle_content())
        icon_canvas.bind("<Button-1>", lambda e: toggle_content())
        
        # Configurar el hover effect
        def on_enter(e):
            icon_canvas.configure(cursor="hand2")
            header_frame.configure(cursor="hand2")
            
        def on_leave(e):
            icon_canvas.configure(cursor="")
            header_frame.configure(cursor="")
        
        header_frame.bind("<Enter>", on_enter)
        header_frame.bind("<Leave>", on_leave)
        icon_canvas.bind("<Enter>", on_enter)
        icon_canvas.bind("<Leave>", on_leave)
        
        # Dibujar el ícono inicial
        draw_icon(False)
        
        return group_frame

    def create_grouped_recommendation_frame(self, parent, unique_centers, suggested_group, group_number, different_plus):
        # Crear frame principal para la tarjeta
        rec_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Crear el frame CustomTkinter para el contenido
        content_frame = ctk.CTkFrame(
            rec_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Frame para el header
        header_frame = ttk.Frame(content_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))
        
        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)
        
        # Crear un Canvas para el ícono SVG
        icon_size = 24
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg='white',
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(10, 5), pady=10)
        
        def draw_icon(is_expanded):
            icon_canvas.delete("all")
            # Crear un botón circular como fondo
            icon_canvas.create_oval(
                0, 0, icon_size, icon_size,
                fill="#2563eb",  # Color azul para recomendaciones
                outline=""
            )
            if is_expanded:
                icon_canvas.create_line(
                    8, 10, 12, 14, 16, 10,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
            else:
                icon_canvas.create_line(
                    10, 8, 14, 12, 10, 16,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
        
        # Contenedor para el texto
        text_container = ttk.Frame(header_frame, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))
        
        # Determinar el título basado en la cantidad de centros únicos
        if len(unique_centers) == 1:
            title_text = f"Centro {next(iter(unique_centers))}"
        else:
            if not hasattr(self, 'letter_counter'):
                self.letter_counter = 0
            letter = chr(65 + self.letter_counter)
            title_text = f"Grupo {letter}"  # Cambiado de "Centros" a "Grupo"
            self.letter_counter += 1
        
        # Título
        title_label = ttk.Label(
            text_container,
            text=title_text,
            font=('Segoe UI', 14, 'bold'),
            foreground='#2563eb',
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Calcular el máximo de PLUs diferentes
        centers_plus_dict = {
            'grupo_existente': set(self.identical_portfolios[suggested_group])
        }
        for centro in unique_centers:
            centers_plus_dict[centro] = set(self.unique_portfolios[centro])
        
        max_missing, _, _ = self.calculate_plu_differences_multi(centers_plus_dict)
        
        # Estadísticas actualizadas
        stats_text = f"Grupo {group_number} • Max. PLUs dif: {max_missing}"
        stats_label = ttk.Label(
            text_container,
            text=stats_text,
            font=('Segoe UI', 12),
            foreground='#6b7280',
            background='white'
        )
        stats_label.pack(anchor='w')
        
        # Frame para el contenido expandible
        content_detail_frame = ttk.Frame(content_frame, style="ContentCard.TFrame")
        
        # Centros únicos
        unique_centers_label = ttk.Label(
            content_detail_frame,
            text=f"Centros con PU: {', '.join(sorted(unique_centers))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        unique_centers_label.pack(anchor='w', padx=15, pady=(10, 5))
        
        # Centros del grupo
        group_centers_label = ttk.Label(
            content_detail_frame,
            text=f"Centros del grupo: {', '.join(sorted(suggested_group))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        group_centers_label.pack(anchor='w', padx=15, pady=5)
        
        # PLUs diferentes
        plues_label = ttk.Label(
            content_detail_frame,
            text=f"PLUs diferentes: {', '.join(map(str, sorted(different_plus)))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        plues_label.pack(anchor='w', padx=15, pady=(0, 10))
        
        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                content_detail_frame.pack(fill=tk.X)
            else:
                content_detail_frame.pack_forget()
            draw_icon(is_expanded.get())
        
        # Hacer que todo el header sea clickeable
        header_frame.bind("<Button-1>", lambda e: toggle_content())
        icon_canvas.bind("<Button-1>", lambda e: toggle_content())
        
        # Configurar el hover effect
        def on_enter(e):
            icon_canvas.configure(cursor="hand2")
            header_frame.configure(cursor="hand2")
        
        def on_leave(e):
            icon_canvas.configure(cursor="")
            header_frame.configure(cursor="")
        
        header_frame.bind("<Enter>", on_enter)
        header_frame.bind("<Leave>", on_leave)
        icon_canvas.bind("<Enter>", on_enter)
        icon_canvas.bind("<Leave>", on_leave)
        
        # Dibujar el ícono inicial
        draw_icon(False)
        
        return rec_frame

    def create_recommendation_frame(self, parent, unique_center, suggested_group, group_number, different_plus):
        # Crear frame principal para la tarjeta
        rec_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Frame para el header de la tarjeta
        header_frame = ttk.Frame(rec_frame, style="ModernCard.TFrame")
        header_frame.pack(fill=tk.X, padx=10, pady=5)
        
        # Frame para el contenido expandible
        content_frame = ttk.Frame(rec_frame, style="ModernCard.TFrame")
        
        # Configurar el estado de expansión
        if not hasattr(self, 'recommendation_states'):
            self.recommendation_states = {}
        self.recommendation_states[unique_center] = tk.BooleanVar(value=False)
        
        def toggle_recommendation():
            current_state = self.recommendation_states[unique_center].get()
            self.recommendation_states[unique_center].set(not current_state)
            if self.recommendation_states[unique_center].get():
                content_frame.pack(fill=tk.X)
                expand_btn.configure(text="▼")
            else:
                content_frame.pack_forget()
                expand_btn.configure(text="▶")
        
        # Contenedor izquierdo para el botón y la información
        left_container = ttk.Frame(header_frame, style="ModernCard.TFrame")
        left_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Botón de expandir/colapsar con estilo moderno
        expand_btn = ctk.CTkButton(
            left_container,
            text="▶",
            width=30,
            height=30,
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            command=toggle_recommendation
        )
        expand_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        # Contenedor para el título y las estadísticas
        info_container = ttk.Frame(left_container, style="ModernCard.TFrame")
        info_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Ajustar el título para que se ajuste al ancho disponible
        def truncate_text(text, max_width, font):
            if not text:
                return text
                
            test_label = ttk.Label(info_container, text=text, font=font)
            text_width = test_label.winfo_reqwidth()
            test_label.destroy()
            
            if text_width <= max_width:
                return text
                
            # Agregar puntos suspensivos y ajustar el texto
            while text and text_width > max_width:
                text = text[:-1]
                test_label = ttk.Label(info_container, text=text + "...", font=font)
                text_width = test_label.winfo_reqwidth()
                test_label.destroy()
                
            return text + "..."
        
        # Título con texto ajustado
        title_text = f"Centro {unique_center}"
        title_font = ('Segoe UI', 14, 'bold')
        max_width = parent.winfo_width() // 3 - 100  # Ajustar según el ancho de la columna
        
        title_label = ttk.Label(
            info_container,
            text=title_text,
            style="ModernCardTitle.TLabel",
            font=title_font,
            wraplength=max_width
        )
        title_label.pack(anchor='w')
        
        # Frame para las estadísticas en línea
        stats_frame = ttk.Frame(info_container, style="ModernCard.TFrame")
        stats_frame.pack(anchor='w', fill=tk.X)
        
        # Información del grupo y cantidad de PLUs diferentes
        stats_text = f"Grupo {group_number} • {len(different_plus)} PLUs diferentes"
        stats_label = ttk.Label(
            stats_frame,
            text=stats_text,
            style="ModernCardStats.TLabel",
            font=('Segoe UI', 12),
            wraplength=max_width
        )
        stats_label.pack(anchor='w')
        
        # Contenido expandible con texto ajustado
        group_text = f"Centros del grupo: {', '.join(suggested_group)}"
        group_info_label = ttk.Label(
            content_frame,
            text=group_text,
            style="ModernCardContent.TLabel",
            wraplength=max_width
        )
        group_info_label.pack(pady=5, padx=10, anchor='w', fill=tk.X)
        
        plues_text = f"PLUs dif: {', '.join(map(str, different_plus))}"
        plues_label = ttk.Label(
            content_frame,
            text=plues_text,
            style="ModernCardContent.TLabel",
            wraplength=max_width
        )
        plues_label.pack(pady=(0, 10), padx=10, anchor='w', fill=tk.X)
        
        return rec_frame

    def create_unique_center_frame(self, parent, center_num, plu_list, non_compatible=False):
        # Crear frame principal para la tarjeta
        group_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Crear el frame CustomTkinter para el contenido
        content_frame = ctk.CTkFrame(
            group_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Frame para el header
        header_frame = ttk.Frame(content_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))
        
        # Contenedor para el botón y la información
        button_container = ttk.Frame(header_frame, style="Card.TFrame")
        button_container.pack(side=tk.LEFT, padx=10, pady=10)
        
        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)
        
        # Configuración de colores para el botón
        if non_compatible:
            fg_color = "#FF4747"      # Rojo para advertencias
            hover_color = "#FF6B6B"   # Rojo más claro para hover
        else:
            fg_color = "#2563eb"      # Azul estándar
            hover_color = "#1d4ed8"   # Azul más oscuro para hover

        # Botón de expandir con estilo moderno
        expand_btn = ctk.CTkButton(
            button_container,
            text="▶",
            width=32,
            height=32,
            fg_color=fg_color,
            hover_color=hover_color,
            corner_radius=8,
            command=lambda: toggle_content()
        )
        expand_btn.pack(side=tk.LEFT)
        
        # Contenedor para el texto
        text_container = ttk.Frame(header_frame, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))
        
        # Título con el centro
        title_label = ttk.Label(
            text_container,
            text=f"Centro {center_num}",
            font=('Segoe UI', 14, 'bold'),
            foreground='#2563eb' if not non_compatible else '#FF4747',
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Estadísticas
        stats_text = f"{len(plu_list):,} PLUs"
        stats_label = ttk.Label(
            text_container,
            text=stats_text,
            font=('Segoe UI', 12),
            foreground='#6b7280',
            background='white'
        )
        stats_label.pack(anchor='w')
        
        # Frame para el contenido expandible
        content_detail_frame = ttk.Frame(content_frame, style="ContentCard.TFrame")
        
        # Lista de PLUs
        plues_label = ttk.Label(
            content_detail_frame,
            text=f"PLUs: {', '.join(map(str, sorted(plu_list)))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        plues_label.pack(anchor='w', padx=15, pady=(10, 10))
        
        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                content_detail_frame.pack(fill=tk.X)
                expand_btn.configure(text="▼")
            else:
                content_detail_frame.pack_forget()
                expand_btn.configure(text="▶")
        
        return group_frame

    def create_unique_recommendation_frame(self, parent, centers, different_plus, next_group_number):
        # Crear frame principal para la tarjeta
        group_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Crear el frame CustomTkinter para el contenido
        content_frame = ctk.CTkFrame(
            group_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Frame para el header
        header_frame = ttk.Frame(content_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))
        
        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)
        
        # Crear un Canvas para el ícono SVG
        icon_size = 24
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg='white',
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(10, 5), pady=10)
        
        def draw_icon(is_expanded):
            icon_canvas.delete("all")
            icon_canvas.create_oval(
                0, 0, icon_size, icon_size,
                fill="#2563eb",
                outline=""
            )
            if is_expanded:
                icon_canvas.create_line(
                    8, 10, 12, 14, 16, 10,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
            else:
                icon_canvas.create_line(
                    10, 8, 14, 12, 10, 16,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
        
        # Contenedor para el texto
        text_container = ttk.Frame(header_frame, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))
        
        # Título usando PG X
        title_label = ttk.Label(
            text_container,
            text=f"PG {next_group_number}",
            font=('Segoe UI', 14, 'bold'),
            foreground='#2563eb',
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Calcular el máximo de PLUs diferentes
        centers_plus_dict = {}
        for center in centers:
            centers_plus_dict[center] = set(self.unique_portfolios[center])
        
        max_missing, _, _ = self.calculate_plu_differences_multi(centers_plus_dict)
        
        # Estadísticas actualizadas
        stats_text = f"Max. PLUs dif: {max_missing}"
        stats_label = ttk.Label(
            text_container,
            text=stats_text,
            font=('Segoe UI', 12),
            foreground='#6b7280',
            background='white'
        )
        stats_label.pack(anchor='w')
        
        # Frame para el contenido expandible
        content_detail_frame = ttk.Frame(content_frame, style="ContentCard.TFrame")
        
        # Mostrar los centros
        centers_label = ttk.Label(
            content_detail_frame,
            text=f"Centros: {', '.join(sorted(centers))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        centers_label.pack(anchor='w', padx=15, pady=(10, 5))
        
        # Mostrar PLUs diferentes
        plues_label = ttk.Label(
            content_detail_frame,
            text=f"PLUs diferentes: {', '.join(map(str, sorted(different_plus)))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        plues_label.pack(anchor='w', padx=15, pady=(0, 10))
        
        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                content_detail_frame.pack(fill=tk.X)
            else:
                content_detail_frame.pack_forget()
            draw_icon(is_expanded.get())
        
        # Hacer que todo el header sea clickeable
        header_frame.bind("<Button-1>", lambda e: toggle_content())
        icon_canvas.bind("<Button-1>", lambda e: toggle_content())
        
        # Configurar el hover effect
        def on_enter(e):
            icon_canvas.configure(cursor="hand2")
            header_frame.configure(cursor="hand2")
            
        def on_leave(e):
            icon_canvas.configure(cursor="")
            header_frame.configure(cursor="")
        
        header_frame.bind("<Enter>", on_enter)
        header_frame.bind("<Leave>", on_leave)
        icon_canvas.bind("<Enter>", on_enter)
        icon_canvas.bind("<Leave>", on_leave)
        
        # Dibujar el ícono inicial
        draw_icon(False)
        
        return group_frame

    def create_distribution_chart(self, parent):
        canvas = tk.Canvas(parent, bg='white', highlightthickness=0)
        canvas.pack(fill=tk.BOTH, expand=True, padx=20)
        
        def update_chart(unique=0, identical=0):
            canvas.delete("all")
            total = unique + identical
            if total == 0:
                return
            
            # Calcular dimensiones del gráfico principal
            width = canvas.winfo_width()
            height = canvas.winfo_height()
            diameter = min(width, height) - 50  # Ajustado al quitar espacio de leyenda
            
            # Centro del gráfico principal
            center_x = width // 2 - 80
            center_y = height // 2  # Centrado vertical al quitar leyenda
            
            # Colores modernos y profesionales
            colors = {
                'in_groups': '#4338CA',
                'unique': '#F59E0B',
                'with_rec': '#10B981',
                'no_rec': '#EF4444'
            }
            
            # Variables para el tooltip
            tooltip_window = None
            
            def show_tooltip(event, text):
                nonlocal tooltip_window
                hide_tooltip()
                
                tooltip_window = tk.Toplevel()
                tooltip_window.wm_overrideredirect(True)
                tooltip_window.wm_geometry(f"+{event.x_root + 10}+{event.y_root + 10}")
                
                tooltip_frame = ttk.Frame(tooltip_window, style='Tooltip.TFrame')
                tooltip_frame.pack(fill='both', expand=True)
                
                ttk.Label(
                    tooltip_frame,
                    text=text,
                    style='Tooltip.TLabel',
                    justify='left',
                    padding=(10, 5)
                ).pack()
                
                tooltip_window.lift()
                
            def hide_tooltip():
                nonlocal tooltip_window
                if tooltip_window:
                    tooltip_window.destroy()
                    tooltip_window = None
            
            # Calcular porcentajes (necesarios para tooltips)
            with_rec = unique - len(self.non_compatible)
            no_rec = len(self.non_compatible)
            
            # Calcular ángulos para el gráfico principal
            unique_angle = (unique / total) * 360
            in_groups_angle = (identical / total) * 360
            
            # Ajustar el ángulo inicial para tener el segmento de únicos a la izquierda
            start_angle = 0 - (unique_angle / 2)
            
            # Variables de animación
            animation_duration = 1000
            frames_per_second = 100
            total_frames = int(animation_duration / (1000 / frames_per_second))
            current_frame = [0]
            
            def ease_out_quad(t):
                return t * (2 - t)
            
            def animate_frame():
                canvas.delete("arc")
                
                progress = current_frame[0] / total_frames
                eased_progress = ease_out_quad(progress)
                
                # Dibujar sección "En grupos" con animación
                current_in_groups = in_groups_angle * eased_progress
                if current_in_groups > 0:
                    in_groups_arc = canvas.create_arc(
                        center_x - diameter//2, center_y - diameter//2,
                        center_x + diameter//2, center_y + diameter//2,
                        start=start_angle + (unique_angle * eased_progress),
                        extent=current_in_groups,
                        fill=colors['in_groups'],
                        outline='white',
                        width=2,
                        tags=('arc', 'in_groups')
                    )
                
                # Dibujar sección "Únicos" con animación
                current_unique = unique_angle * eased_progress
                if current_unique > 0:
                    unique_arc = canvas.create_arc(
                        center_x - diameter//2, center_y - diameter//2,
                        center_x + diameter//2, center_y + diameter//2,
                        start=start_angle,
                        extent=current_unique,
                        fill=colors['unique'],
                        outline='white',
                        width=2,
                        tags=('arc', 'unique')
                    )
                
                # Subgráfico para centros únicos con animación
                if unique > 0:
                    sub_diameter = diameter * 0.45
                    sub_x = center_x + diameter//2 + 50
                    sub_y = center_y
                    
                    # Calcular ángulos para el subgráfico
                    with_rec_angle = (with_rec / unique) * 360
                    no_rec_angle = (no_rec / unique) * 360
                    
                    # Dibujar subgráfico con animación
                    current_with_rec = with_rec_angle * eased_progress
                    if current_with_rec > 0:
                        with_rec_arc = canvas.create_arc(
                            sub_x - sub_diameter//2, sub_y - sub_diameter//2,
                            sub_x + sub_diameter//2, sub_y + sub_diameter//2,
                            start=0,
                            extent=current_with_rec,
                            fill=colors['with_rec'],
                            outline='white',
                            width=2,
                            tags=('arc', 'with_rec')
                        )
                    
                    current_no_rec = no_rec_angle * eased_progress
                    if current_no_rec > 0:
                        no_rec_arc = canvas.create_arc(
                            sub_x - sub_diameter//2, sub_y - sub_diameter//2,
                            sub_x + sub_diameter//2, sub_y + sub_diameter//2,
                            start=with_rec_angle * eased_progress,
                            extent=current_no_rec,
                            fill=colors['no_rec'],
                            outline='white',
                            width=2,
                            tags=('arc', 'no_rec')
                        )
                
                # Continuar la animación si no ha terminado
                if current_frame[0] < total_frames:
                    current_frame[0] += 1
                    canvas.after(int(1000/frames_per_second), animate_frame)
                else:
                    # Bind eventos para tooltips una vez que la animación termine
                    canvas.tag_bind('with_rec', '<Enter>', 
                        lambda e: show_tooltip(e, f'Con recomendación: {with_rec} centros\n{(with_rec/unique*100):.1f}% de centros únicos'))
                    canvas.tag_bind('with_rec', '<Leave>', lambda e: hide_tooltip())
                    
                    canvas.tag_bind('no_rec', '<Enter>', 
                        lambda e: show_tooltip(e, f'Sin recomendación: {no_rec} centros\n{(no_rec/unique*100):.1f}% de centros únicos'))
                    canvas.tag_bind('no_rec', '<Leave>', lambda e: hide_tooltip())
                    
                    canvas.tag_bind('in_groups', '<Enter>', 
                        lambda e: show_tooltip(e, f'En grupos: {identical} centros\n{(identical/total*100):.1f}% del total'))
                    canvas.tag_bind('in_groups', '<Leave>', lambda e: hide_tooltip())
                    
                    canvas.tag_bind('unique', '<Enter>', 
                        lambda e: show_tooltip(e, f'Únicos: {unique} centros\n{(unique/total*100):.1f}% del total'))
                    canvas.tag_bind('unique', '<Leave>', lambda e: hide_tooltip())
            
            # Iniciar la animación
            current_frame[0] = 0
            animate_frame()

        canvas.bind('<Configure>', lambda e: update_chart(
            int(self.unique_centers.get()), 
            int(self.total_centers.get()) - int(self.unique_centers.get())
        ))
        
        return update_chart

    def configure_styles(self):
        # Configurar el tema base
        FONT_FAMILY = "Segoe UI"

        # Estilos base
        self.style.configure(
            "Card.TFrame",
            background='white',
            relief='flat'
        )
        
        self.style.configure(
            "ContentCard.TFrame",
            background='#f9fafb',
            borderwidth=0
        )

        # Actualizar el estilo ModernCard para incluir sombra y bordes
        self.style.configure(
            "ModernCard.TFrame",
            background='white',
            relief='solid',
            borderwidth=1
        )

        # Estilos para el título principal
        self.style.configure(
            "Title.TLabel",
            font=(FONT_FAMILY, 32, 'bold'),
            background='#f8fafc',
            foreground='#1e293b'
        )

        # Estilo para separadores
        self.style.configure(
            "BlackSeparator.TFrame",
            background='black'
        )

        # Estilos para estadísticas
        self.style.configure(
            "Stats.TLabel",
            background='white',
            foreground='#2563eb',
            font=(FONT_FAMILY, 14, 'bold')
        )

        self.style.configure(
            "StatsDesc.TLabel",
            background='white',
            foreground='#64748b',
            font=(FONT_FAMILY, 12)
        )

        # Estilos para tarjetas con advertencia
        self.style.configure(
            "WarningStats.TLabel",
            background='white',
            foreground='#E53E3E',
            font=(FONT_FAMILY, 14, 'bold')
        )

        # Estilo para estadísticas en azul
        self.style.configure(
            "BlueStats.TLabel",
            background='white',
            foreground='#4F46E5',
            font=(FONT_FAMILY, 14, 'bold')
        )

        # Estilo para el notebook
        self.style.configure(
            "Custom.TNotebook",
            background="white",
            borderwidth=0,
            tabmargin=0
        )

        self.style.configure(
            "Custom.TNotebook.Tab",
            padding=[20, 10],
            background="white",
            borderwidth=0
        )

        self.style.map("Custom.TNotebook.Tab",
            background=[("selected", "white")],
            expand=[("selected", [1, 1, 1, 0])],
            borderwidth=[("selected", 0)]
        )

        # Estilos para tooltips
        self.style.configure(
            "Tooltip.TFrame",
            background='#1e293b'
        )

        self.style.configure(
            "Tooltip.TLabel",
            background='#1e293b',
            foreground='white',
            font=(FONT_FAMILY, 11)
        )

    def create_stat_card(self, parent, title, variable, color, column):
        # Crear el frame para la tarjeta
        card = ttk.Frame(parent, style="Card.TFrame")
        card.grid(row=0, column=column, padx=10, pady=5, sticky="nsew")
        
        # Frame interno con padding
        card_inner = ttk.Frame(
            card,
            style="Card.TFrame",
            padding=(10, 10)  # Reducir padding vertical
        )
        card_inner.pack(fill=tk.BOTH, expand=True)
        
        # Frame superior para el número
        number_frame = ttk.Frame(card_inner, style="Card.TFrame")
        number_frame.pack(fill=tk.X)
        
        # Número centrado
        number_label = ttk.Label(
            number_frame,
            textvariable=variable,
            font=('Segoe UI', 28, 'bold'),
            foreground=color,
            background='white',
            anchor='center',
            justify='center'
        )
        number_label.pack()
        
        # Frame para el título
        title_frame = ttk.Frame(card_inner, style="Card.TFrame")
        title_frame.pack(fill=tk.X)
        
        if "Másteres" in title:
            # Frame contenedor para ambas líneas
            text_container = ttk.Frame(title_frame, style="Card.TFrame")
            text_container.pack(fill=tk.X)
            
            # Primera línea: "Planogramas"
            ttk.Label(
                text_container,
                text="Planogramas",
                font=('Segoe UI', 11),
                foreground='#64748b',
                background='white',
                anchor='center',
                justify='center'
            ).pack(fill=tk.X)
            
            # Segunda línea: "Másteres Iniciales/Finales"
            ttk.Label(
                text_container,
                text="Másteres " + ("Iniciales" if "Iniciales" in title else "Finales"),
                font=('Segoe UI', 11),
                foreground='#64748b',
                background='white',
                anchor='center',
                justify='center'
            ).pack(fill=tk.X)
        else:
            # Título normal de una línea
            ttk.Label(
                title_frame,
                text=title,
                font=('Segoe UI', 11),
                foreground='#64748b',
                background='white',
                anchor='center',
                justify='center'
            ).pack(fill=tk.X)
        
        # Asegurar altura fija pero más compacta
        card_inner.pack_propagate(False)
        card_inner.configure(height=130)  # Reducir altura
        
        return card

    def create_group_frame(self, parent, group_num, centers, count, timestamp, plu_count):
        # Crear frame principal para la tarjeta
        group_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Crear el frame CustomTkinter para el contenido
        content_frame = ctk.CTkFrame(
            group_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Frame para el header
        header_frame = ttk.Frame(content_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))
        
        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)
        
        # Crear un Canvas para el ícono SVG
        icon_size = 24
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg='white',
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(10, 5), pady=10)
        
        # SVG para el ícono de ChevronRight
        chevron_right = """
        <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" 
            stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
            <polyline points="9 18 15 12 9 6"></polyline>
        </svg>
        """
        
        # SVG para el ícono de ChevronDown
        chevron_down = """
        <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" 
            stroke="white" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
            <polyline points="6 9 12 15 18 9"></polyline>
        </svg>
        """
        
        def draw_icon(is_expanded):
            icon_canvas.delete("all")
            # Crear un botón circular como fondo
            icon_canvas.create_oval(
                0, 0, icon_size, icon_size,
                fill="#2563eb",
                outline=""
            )
            # Convertir el SVG a PhotoImage
            if is_expanded:
                svg_data = chevron_down
            else:
                svg_data = chevron_right
                
            # Aquí dibujamos el SVG directamente usando las coordenadas
            if is_expanded:
                # Dibujar chevron down
                icon_canvas.create_line(
                    8, 10, 12, 14, 16, 10,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
            else:
                # Dibujar chevron right
                icon_canvas.create_line(
                    10, 8, 14, 12, 10, 16,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
        
        # Contenedor para el texto
        text_container = ttk.Frame(header_frame, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))
        
        # Título del grupo
        title_label = ttk.Label(
            text_container,
            text=f"Grupo {group_num}",
            font=('Segoe UI', 14, 'bold'),
            foreground='#2563eb',
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Estadísticas
        stats_text = f"{count} centros • {plu_count:,} PLUs"
        stats_label = ttk.Label(
            text_container,
            text=stats_text,
            font=('Segoe UI', 12),
            foreground='#6b7280',
            background='white'
        )
        stats_label.pack(anchor='w')
        
        # Frame para el contenido expandible
        content_detail_frame = ttk.Frame(content_frame, style="ContentCard.TFrame")
        
        # Centros
        centers_label = ttk.Label(
            content_detail_frame,
            text=f"Centros: {centers}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        centers_label.pack(anchor='w', padx=15, pady=(10, 5))
        
        # Timestamp
        timestamp_label = ttk.Label(
            content_detail_frame,
            text=f"Última actualización: {timestamp}",
            font=('Segoe UI', 10),
            foreground='#6b7280',
            background='#f9fafb'
        )
        timestamp_label.pack(anchor='w', padx=15, pady=(0, 10))
        
        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                content_detail_frame.pack(fill=tk.X)
            else:
                content_detail_frame.pack_forget()
            draw_icon(is_expanded.get())
        
        # Hacer que todo el header sea clickeable
        header_frame.bind("<Button-1>", lambda e: toggle_content())
        icon_canvas.bind("<Button-1>", lambda e: toggle_content())
        
        # Configurar el hover effect
        def on_enter(e):
            icon_canvas.configure(cursor="hand2")
            header_frame.configure(cursor="hand2")
            
        def on_leave(e):
            icon_canvas.configure(cursor="")
            header_frame.configure(cursor="")
        
        header_frame.bind("<Enter>", on_enter)
        header_frame.bind("<Leave>", on_leave)
        icon_canvas.bind("<Enter>", on_enter)
        icon_canvas.bind("<Leave>", on_leave)
        
        # Dibujar el ícono inicial
        draw_icon(False)
        
        return group_frame

    def create_unique_center_frame(self, parent, center_num, plu_list, non_compatible=False):
        # Crear frame principal para la tarjeta
        group_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Crear el frame CustomTkinter para el contenido
        content_frame = ctk.CTkFrame(
            group_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Frame para el header
        header_frame = ttk.Frame(content_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))
        
        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)
        
        # Configuración de colores
        if non_compatible:
            bg_color = "#FF4747"      # Rojo para advertencias
        else:
            bg_color = "#2563eb"      # Azul estándar
        
        # Crear un Canvas para el ícono SVG
        icon_size = 24
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg='white',
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(10, 5), pady=10)
        
        def draw_icon(is_expanded):
            icon_canvas.delete("all")
            # Crear un botón circular como fondo
            icon_canvas.create_oval(
                0, 0, icon_size, icon_size,
                fill=bg_color,
                outline=""
            )
            # Dibujar el ícono
            if is_expanded:
                # Dibujar chevron down
                icon_canvas.create_line(
                    8, 10, 12, 14, 16, 10,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
            else:
                # Dibujar chevron right
                icon_canvas.create_line(
                    10, 8, 14, 12, 10, 16,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
        
        # Contenedor para el texto
        text_container = ttk.Frame(header_frame, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))
        
        # Título con el centro
        title_label = ttk.Label(
            text_container,
            text=f"Centro {center_num}",
            font=('Segoe UI', 14, 'bold'),
            foreground=bg_color,
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Estadísticas
        stats_text = f"{len(plu_list):,} PLUs"
        stats_label = ttk.Label(
            text_container,
            text=stats_text,
            font=('Segoe UI', 12),
            foreground='#6b7280',
            background='white'
        )
        stats_label.pack(anchor='w')
        
        # Frame para el contenido expandible
        content_detail_frame = ttk.Frame(content_frame, style="ContentCard.TFrame")
        
        # Lista de PLUs
        plues_label = ttk.Label(
            content_detail_frame,
            text=f"PLUs: {', '.join(map(str, sorted(plu_list)))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        plues_label.pack(anchor='w', padx=15, pady=(10, 10))
        
        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                content_detail_frame.pack(fill=tk.X)
            else:
                content_detail_frame.pack_forget()
            draw_icon(is_expanded.get())
        
        # Hacer que todo el header sea clickeable
        header_frame.bind("<Button-1>", lambda e: toggle_content())
        icon_canvas.bind("<Button-1>", lambda e: toggle_content())
        
        # Configurar el hover effect
        def on_enter(e):
            icon_canvas.configure(cursor="hand2")
            header_frame.configure(cursor="hand2")
            
        def on_leave(e):
            icon_canvas.configure(cursor="")
            header_frame.configure(cursor="")
        
        header_frame.bind("<Enter>", on_enter)
        header_frame.bind("<Leave>", on_leave)
        icon_canvas.bind("<Enter>", on_enter)
        icon_canvas.bind("<Leave>", on_leave)
        
        # Dibujar el ícono inicial
        draw_icon(False)
        
        return group_frame

    def create_final_group_frame(self, parent, group_num, centers, total_plus):
        """
        Crea un frame para mostrar un grupo final con un botón de análisis compacto.
        
        Args:
            parent: Frame padre donde se creará este frame
            group_num: Número del grupo
            centers: Conjunto de centros en el grupo
            total_plus: Conjunto de PLUs del grupo
        """
        # Crear frame principal para la tarjeta
        group_frame = ttk.Frame(parent, style="ModernCard.TFrame")
        
        # Crear el frame CustomTkinter para el contenido
        content_frame = ctk.CTkFrame(
            group_frame,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#e2e8f0"
        )
        content_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Frame para el header
        header_frame = ttk.Frame(content_frame, style="Card.TFrame")
        header_frame.pack(fill=tk.X, pady=(0, 1))
        
        # Variable para controlar el estado expandido/colapsado
        is_expanded = tk.BooleanVar(value=False)
        
        # Crear un Canvas para el ícono de expansión
        icon_size = 24
        icon_canvas = tk.Canvas(
            header_frame,
            width=icon_size,
            height=icon_size,
            bg='white',
            highlightthickness=0
        )
        icon_canvas.pack(side=tk.LEFT, padx=(10, 5), pady=10)
        
        def draw_expand_icon(is_expanded):
            icon_canvas.delete("all")
            # Crear un botón circular como fondo
            icon_canvas.create_oval(
                0, 0, icon_size, icon_size,
                fill="#16a34a",  # Verde para grupos finales
                outline=""
            )
            # Dibujar el ícono
            if is_expanded:
                icon_canvas.create_line(
                    8, 10, 12, 14, 16, 10,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
            else:
                icon_canvas.create_line(
                    10, 8, 14, 12, 10, 16,
                    fill="white", width=2, capstyle="round", joinstyle="round"
                )
        
        # Contenedor para el texto y el botón
        content_container = ttk.Frame(header_frame, style="Card.TFrame")
        content_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 10))
        
        # Frame izquierdo para título y estadísticas
        text_container = ttk.Frame(content_container, style="Card.TFrame")
        text_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Título
        title_label = ttk.Label(
            text_container,
            text=f"Grupo Final {group_num}",
            font=('Segoe UI', 14, 'bold'),
            foreground='#16a34a',
            background='white'
        )
        title_label.pack(anchor='w')
        
        # Estadísticas
        stats_text = f"{len(centers)} centros • {len(total_plus):,} PLUs"
        stats_label = ttk.Label(
            text_container,
            text=stats_text,
            font=('Segoe UI', 12),
            foreground='#6b7280',
            background='white'
        )
        stats_label.pack(anchor='w')
        
        # Frame para el botón
        button_container = ttk.Frame(content_container, style="Card.TFrame")
        button_container.pack(side=tk.RIGHT, padx=10)
        
        # Variable para controlar el tooltip
        tooltip_window = None
        
        def show_tooltip(event):
            nonlocal tooltip_window
            if tooltip_window is None:
                tooltip_window = tk.Toplevel()
                tooltip_window.wm_overrideredirect(True)
                tooltip_window.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
                
                frame = ttk.Frame(tooltip_window, style="Tooltip.TFrame", padding=5)
                frame.pack(fill='both', expand=True)
                
                label = ttk.Label(
                    frame,
                    text="Analizar Variación",
                    style="Tooltip.TLabel",
                    background='#1e293b',
                    foreground='white',
                    font=('Segoe UI', 10)
                )
                label.pack()
        
        def hide_tooltip(event=None):
            nonlocal tooltip_window
            if tooltip_window:
                tooltip_window.destroy()
                tooltip_window = None
                
        # Botón de análisis
        analyze_button = ctk.CTkButton(
            button_container,
            text="",
            width=35,
            height=35,
            command=lambda: self.show_portfolio_variation(group_num, centers),
            fg_color="#16a34a",
            hover_color="#15803d",
            corner_radius=8
        )
        analyze_button.pack(side=tk.RIGHT)

        # Crear el ícono directamente en el botón
        def create_analysis_icon(color):
            icon_size = 20
            icon_canvas = tk.Canvas(
                analyze_button,
                width=icon_size,
                height=icon_size,
                bg=color,
                highlightthickness=0
            )
            icon_canvas.place(relx=0.5, rely=0.5, anchor="center")

            # Dibujar el ícono
            # Tabla base
            icon_canvas.create_rectangle(
                4, 4, 16, 16,
                outline="white",
                width=1.5
            )
            
            # Líneas horizontales de la tabla
            icon_canvas.create_line(
                4, 8, 16, 8,
                fill="white",
                width=1
            )
            icon_canvas.create_line(
                4, 12, 16, 12,
                fill="white",
                width=1
            )
            
            # Línea de tendencia
            icon_canvas.create_line(
                6, 14, 9, 10, 12, 12, 15, 6,
                fill="white",
                width=1.5,
                smooth=True
            )

            # Hacer que el canvas sea clickeable
            icon_canvas.bind("<Button-1>", lambda e: self.show_portfolio_variation(group_num, centers))
            return icon_canvas

        # Crear el ícono inicial
        analysis_icon = create_analysis_icon(analyze_button._fg_color)

        # Actualizar el ícono cuando el botón cambia de estado
        def update_icon_color(e):
            analysis_icon.configure(bg=analyze_button._hover_color)

        def restore_icon_color(e):
            analysis_icon.configure(bg=analyze_button._fg_color)
        
        # Frame para el contenido expandible
        content_detail_frame = ttk.Frame(content_frame, style="ContentCard.TFrame")
        
        # Centros
        centers_label = ttk.Label(
            content_detail_frame,
            text=f"Centros: {', '.join(sorted(centers))}",
            wraplength=350,
            justify='left',
            font=('Segoe UI', 11),
            foreground='#4b5563',
            background='#f9fafb'
        )
        centers_label.pack(anchor='w', padx=15, pady=(10, 10))
        
        def toggle_content():
            current_state = is_expanded.get()
            is_expanded.set(not current_state)
            if is_expanded.get():
                content_detail_frame.pack(fill=tk.X)
            else:
                content_detail_frame.pack_forget()
            draw_expand_icon(is_expanded.get())
        
        # Configurar eventos de expansión
        header_frame.bind("<Button-1>", lambda e: toggle_content())
        icon_canvas.bind("<Button-1>", lambda e: toggle_content())
        
        # Vincular eventos del botón
        analyze_button.bind('<Enter>', lambda e: (update_icon_color(e), show_tooltip(e)))
        analyze_button.bind('<Leave>', lambda e: (restore_icon_color(e), hide_tooltip(e)))
        
        # Hacer que el ícono responda al click y tooltips
        analysis_icon.bind('<Enter>', lambda e: (update_icon_color(e), show_tooltip(e)))
        analysis_icon.bind('<Leave>', lambda e: (restore_icon_color(e), hide_tooltip(e)))
        
        # Efecto hover para el header expandible
        def on_enter(e):
            icon_canvas.configure(cursor="hand2")
            header_frame.configure(cursor="hand2")
            
        def on_leave(e):
            icon_canvas.configure(cursor="")
            header_frame.configure(cursor="")
        
        header_frame.bind("<Enter>", on_enter)
        header_frame.bind("<Leave>", on_leave)
        icon_canvas.bind("<Enter>", on_enter)
        icon_canvas.bind("<Leave>", on_leave)
        
        # Dibujar el ícono inicial de expansión
        draw_expand_icon(False)
        
        return group_frame

    def calculate_total_different_plus(self, centers_list):
        """
        Calcula el total de PLUs diferentes en un grupo de centros.
        Un PLU es diferente si está presente en algunos centros pero no en todos.
        
        Args:
            centers_list: Conjunto de centros a analizar
            
        Returns:
            tuple: (different_plus, all_plus)
                - different_plus: Conjunto de PLUs que difieren entre centros
                - all_plus: Conjunto de todos los PLUs del grupo
        """
        # Obtener PLUs de cada centro
        centers_plus = {}
        for center in centers_list:
            if center in self.unique_portfolios:
                centers_plus[center] = set(self.unique_portfolios[center])
            else:
                for centers, plus in self.identical_portfolios.items():
                    if center in centers:
                        centers_plus[center] = set(plus)
                        break
        
        # Encontrar PLUs presentes en todos los centros y PLUs diferentes
        if not centers_plus:
            return set(), set()
            
        all_plus = set.union(*centers_plus.values())
        common_plus = set.intersection(*centers_plus.values())
        different_plus = all_plus - common_plus
        
        return different_plus, all_plus

    def calculate_final_groups(self, plu_limit):
        """
        Calcula los grupos finales donde cada grupo puede tener máximo 10 PLUs diferentes en total.
        Un PLU se considera diferente si está presente en algunos centros del grupo pero no en otros.
        
        Returns:
            list: Lista de diccionarios con los grupos finales
        """
        assigned_centers = set()
        final_groups = []

        # FASE 1: Fusionar grupos existentes
        groups = list(self.identical_portfolios.items())
        n = len(groups)
        used_groups = set()
        
        for i in range(n):
            if i in used_groups:
                continue
                
            current_group = {
                'centers': set(groups[i][0]),
                'plus': set(groups[i][1])
            }
            used_groups.add(i)
            
            # Intentar fusionar con otros grupos
            changed = True
            while changed:
                changed = False
                for j in range(n):
                    if j in used_groups:
                        continue
                        
                    # Probar fusión
                    test_centers = current_group['centers'] | set(groups[j][0])
                    diff_plus, _ = self.calculate_total_different_plus(test_centers)
                    
                    if len(diff_plus) <= plu_limit:
                        # Fusionar grupos
                        current_group['centers'] = test_centers
                        current_group['plus'].update(groups[j][1])
                        used_groups.add(j)
                        changed = True
            
            # Añadir grupo a grupos finales
            final_groups.append(current_group)
            assigned_centers.update(current_group['centers'])
        
        # FASE 2: Intentar agregar centros únicos a grupos existentes
        unassigned_centers = set(self.unique_portfolios.keys()) - assigned_centers
        non_compatible = []
        
        for center in unassigned_centers:
            assigned = False
            
            # Intentar agregar a cada grupo existente
            for group in final_groups:
                test_centers = group['centers'] | {center}
                diff_plus, _ = self.calculate_total_different_plus(test_centers)
                
                if len(diff_plus) <= plu_limit: 
                    group['centers'].add(center)
                    group['plus'].update(self.unique_portfolios[center])
                    assigned = True
                    assigned_centers.add(center)
                    break
            
            if not assigned:
                non_compatible.append(center)
        
        # Formar nuevos grupos con centros no compatibles
        remaining_centers = non_compatible.copy()
        non_compatible = []
        
        while remaining_centers:
            center = remaining_centers.pop(0)
            current_group = {
                'centers': {center},
                'plus': set(self.unique_portfolios[center])
            }
            
            # Intentar agregar otros centros restantes
            for other_center in remaining_centers[:]:
                test_centers = current_group['centers'] | {other_center}
                diff_plus, _ = self.calculate_total_different_plus(test_centers)
                
                if len(diff_plus) <= plu_limit:
                    current_group['centers'].add(other_center)
                    current_group['plus'].update(self.unique_portfolios[other_center])
                    remaining_centers.remove(other_center)
            
            if len(current_group['centers']) > 1:
                final_groups.append(current_group)
            else:
                non_compatible.append(center)
        
        # Actualizar lista de centros no compatibles
        self.non_compatible = sorted(non_compatible)
        
        # Ordenar grupos por tamaño
        final_groups.sort(key=lambda x: len(x['centers']), reverse=True)
        
        # Imprimir estadísticas
        print("\nDetalle de grupos finales:")
        for i, group in enumerate(final_groups, 1):
            diff_plus, all_plus = self.calculate_total_different_plus(group['centers'])
            print(f"\nGrupo Final {i}:")
            print(f"Centros ({len(group['centers'])}): {sorted(group['centers'])}")
            print(f"PLUs diferentes ({len(diff_plus)}): {sorted(diff_plus)}")
            print(f"Total PLUs en el grupo: {len(all_plus)}")
        
        print(f"\nEstadísticas finales:")
        print(f"Grupos idénticos originales: {len(self.identical_portfolios)}")
        print(f"Grupos después de fusiones: {len(final_groups)}")
        print(f"Centros sin agrupar: {len(self.non_compatible)}")
        print(f"Total centros agrupados: {sum(len(g['centers']) for g in final_groups)}")
        
        return final_groups

    def update_summary_tab(self):
        # Limpiar el grid existente
        for widget in self.summary_grid.winfo_children():
            widget.destroy()
        
        # Calcular grupos finales
        final_groups = self.calculate_final_groups(self.current_plu_limit)
        
        # Frame para estadísticas principales
        stats_frame = ttk.Frame(self.summary_grid, style="Card.TFrame")
        stats_frame.grid(row=0, column=0, columnspan=3, sticky='ew', padx=5, pady=5)
        
        # Calcular estadísticas
        groups_orig = len(self.identical_portfolios)
        groups_final = len(final_groups)
        groups_reduction = groups_orig - groups_final
        total_centers = sum(len(group['centers']) for group in final_groups)
        total_ungrouped = len(self.non_compatible)
        
        optimization_percentage = (groups_reduction / groups_orig * 100) if groups_orig > 0 else 0
        grouping_efficiency = (total_centers / (total_centers + total_ungrouped) * 100) if (total_centers + total_ungrouped) > 0 else 0
        
        # Frame para los gráficos
        charts_frame = ttk.Frame(self.summary_grid, style="Card.TFrame")
        charts_frame.grid(row=1, column=0, columnspan=3, sticky='nsew', padx=5, pady=5)
        
        # Configurar el grid para los gráficos
        charts_frame.columnconfigure(0, weight=1)
        charts_frame.columnconfigure(1, weight=1)
        
        # Canvas para el gráfico de embudo (izquierda)
        funnel_canvas = tk.Canvas(
            charts_frame,
            width=400,
            height=300,
            bg='white',
            highlightthickness=0
        )
        funnel_canvas.grid(row=0, column=0, padx=10, pady=10, sticky='nsew')
        
        # Canvas para el gráfico de barras (derecha)
        bar_canvas = tk.Canvas(
            charts_frame,
            width=400,
            height=300,
            bg='white',
            highlightthickness=0
        )
        bar_canvas.grid(row=0, column=1, padx=10, pady=10, sticky='nsew')
        
        # Dibujar gráfico de embudo
        def draw_funnel():
            width = funnel_canvas.winfo_width()
            height = funnel_canvas.winfo_height()
            
            # Calcular los valores necesarios usando las nuevas métricas
            masteres_iniciales = len(self.identical_portfolios) + len(self.unique_portfolios)
            final_groups = self.calculate_final_groups(self.current_plu_limit)
            masteres_finales = len(final_groups) + len(self.non_compatible)

            # Validar para evitar división por cero
            if masteres_iniciales == 0:
                optimization_percentage = 0
            else:
                optimization_percentage = ((masteres_iniciales - masteres_finales) / masteres_iniciales) * 100
                    
            # Título del gráfico
            funnel_canvas.create_text(
                width/2, 30,
                text="Optimización de Planogramas",
                font=('Segoe UI', 14, 'bold'),
                fill='#1F2937'
            )
            
            # Calcular dimensiones del embudo
            funnel_width_top = width * 0.7
            funnel_width_bottom = width * 0.3
            funnel_height = height * 0.5
            start_y = 80
            
            # Dibujar el embudo con degradado
            steps = 50
            step_height = funnel_height / steps
            
            for i in range(steps):
                y1 = start_y + (i * step_height)
                y2 = y1 + step_height
                
                width_top = funnel_width_top - (i * (funnel_width_top - funnel_width_bottom) / steps)
                width_bottom = funnel_width_top - ((i + 1) * (funnel_width_top - funnel_width_bottom) / steps)
                
                x1_top = width/2 - width_top/2
                x2_top = width/2 + width_top/2
                x1_bottom = width/2 - width_bottom/2
                x2_bottom = width/2 + width_bottom/2
                
                intensity = int(180 + (75 * (i / steps)))
                color = f'#{intensity:02x}{intensity:02x}FF'
                
                funnel_canvas.create_polygon(
                    x1_top, y1,
                    x2_top, y1,
                    x2_bottom, y2,
                    x1_bottom, y2,
                    fill=color,
                    outline='white',
                    width=1
                )
            
            # Texto superior actualizado
            funnel_canvas.create_text(
                width/2, start_y - 20,
                text=f"Planogramas Másteres Iniciales: {masteres_iniciales}",
                font=('Segoe UI', 12),
                fill='#1F2937'
            )
            
            # Texto inferior actualizado
            funnel_canvas.create_text(
                width/2, start_y + funnel_height + 20,
                text=f"Planogramas Másteres Finales: {masteres_finales}",
                font=('Segoe UI', 12),
                fill='#1F2937'
            )        

        # Dibujar gráfico de barras
        def draw_bars():
            width = bar_canvas.winfo_width()
            height = bar_canvas.winfo_height()
            
            # Título del gráfico
            bar_canvas.create_text(
                width/2, 30,
                text="Distribución de Centros",
                font=('Segoe UI', 14, 'bold'),
                fill='#1F2937'
            )
            
            # Configuración de barras
            bar_width = 60
            bar_spacing = 40
            max_height = height * 0.6
            start_y = height - 50
            
            # Escala para las barras
            max_value = max(total_centers, total_ungrouped)
            scale = max_height / max_value if max_value > 0 else 1
            
            # Primera barra - Centros agrupados
            height_1 = total_centers * scale
            x1 = width/3
            bar_canvas.create_rectangle(
                x1 - bar_width/2, start_y - height_1,
                x1 + bar_width/2, start_y,
                fill='#10B981',
                outline='white'
            )
            
            # Etiqueta y valor
            bar_canvas.create_text(
                x1, start_y + 20,
                text="Agrupados",
                font=('Segoe UI', 10),
                fill='#1F2937'
            )
            bar_canvas.create_text(
                x1, start_y - height_1 - 20,
                text=str(total_centers),
                font=('Segoe UI', 12, 'bold'),
                fill='#1F2937'
            )
            
            # Segunda barra - Centros sin grupo
            height_2 = total_ungrouped * scale
            x2 = 2 * width/3
            bar_canvas.create_rectangle(
                x2 - bar_width/2, start_y - height_2,
                x2 + bar_width/2, start_y,
                fill='#EF4444',
                outline='white'
            )
            
            # Etiqueta y valor
            bar_canvas.create_text(
                x2, start_y + 20,
                text="Sin grupo",
                font=('Segoe UI', 10),
                fill='#1F2937'
            )
            bar_canvas.create_text(
                x2, start_y - height_2 - 20,
                text=str(total_ungrouped),
                font=('Segoe UI', 12, 'bold'),
                fill='#1F2937'
            )
        
        # Frame para métricas adicionales
        metrics_frame = ttk.Frame(self.summary_grid, style="Card.TFrame")
        metrics_frame.grid(row=2, column=0, columnspan=3, sticky='ew', padx=5, pady=5)
        
        # Crear dos tarjetas de métricas
        metric_card_1 = ctk.CTkFrame(
            metrics_frame,
            fg_color="#EEF2FF",
            corner_radius=8
        )
        metric_card_1.pack(side=tk.LEFT, expand=True, fill='both', padx=5, pady=5)

        ttk.Label(
            metric_card_1,
            text="Reducción de planogramas",
            font=('Segoe UI', 11),
            foreground='#4338CA',
            background='#EEF2FF'
        ).pack(pady=(10,0))

        # Calcular los valores necesarios usando las nuevas métricas
        masteres_iniciales = len(self.identical_portfolios) + len(self.unique_portfolios)
        final_groups = self.calculate_final_groups(self.current_plu_limit)
        masteres_finales = len(final_groups) + len(self.non_compatible)
        planogramas_reduction = masteres_iniciales - masteres_finales

        ttk.Label(
            metric_card_1,
            text=f"{planogramas_reduction} planogramas menos",
            font=('Segoe UI', 16, 'bold'),
            foreground='#312E81',
            background='#EEF2FF'
        ).pack()

        optimization_percentage = ((masteres_iniciales - masteres_finales) / masteres_iniciales) * 100 if masteres_iniciales > 0 else 0
        ttk.Label(
            metric_card_1,
            text=f"{optimization_percentage:.1f}% de optimización",
            font=('Segoe UI', 11),
            foreground='#4338CA',
            background='#EEF2FF'
        ).pack(pady=(0,10))
        
        metric_card_2 = ctk.CTkFrame(
            metrics_frame,
            fg_color="#ECFDF5",
            corner_radius=8
        )
        metric_card_2.pack(side=tk.LEFT, expand=True, fill='both', padx=5, pady=5)
        
        ttk.Label(
            metric_card_2,
            text="Eficiencia de agrupación",
            font=('Segoe UI', 11),
            foreground='#047857',
            background='#ECFDF5'
        ).pack(pady=(10,0))
        
        ttk.Label(
            metric_card_2,
            text=f"{grouping_efficiency:.1f}%",
            font=('Segoe UI', 16, 'bold'),
            foreground='#064E3B',
            background='#ECFDF5'
        ).pack()
        
        ttk.Label(
            metric_card_2,
            text="de centros agrupados",
            font=('Segoe UI', 11),
            foreground='#047857',
            background='#ECFDF5'
        ).pack(pady=(0,10))
        
        # Dibujar los gráficos inicialmente
        draw_funnel()
        draw_bars()
        
        # Configurar actualización de gráficos al redimensionar
        def on_resize(event):
            funnel_canvas.delete('all')
            bar_canvas.delete('all')
            draw_funnel()
            draw_bars()
        
        funnel_canvas.bind('<Configure>', on_resize)
        bar_canvas.bind('<Configure>', on_resize)
        
        # Continuar con los grupos finales como antes...
        separator = ttk.Frame(self.summary_grid, height=2, style="BlackSeparator.TFrame")
        separator.grid(row=3, column=0, columnspan=3, sticky='ew', pady=15)
        
        # Mostrar grupos finales
        for i, group in enumerate(final_groups, 1):
            group_frame = self.create_final_group_frame(
                self.summary_grid,
                i,
                group['centers'],
                group['plus']
            )
            row = ((i - 1) // 3) + 4  # +4 para dejar espacio para los elementos anteriores
            col = (i - 1) % 3
            group_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
            self.summary_grid.grid_rowconfigure(row, weight=1)
            last_row = row

        # Mostrar centros sin recomendación al final
        if self.non_compatible:
            separator = ttk.Frame(self.summary_grid, height=2, style="BlackSeparator.TFrame")
            separator.grid(row=last_row + 1, column=0, columnspan=3, sticky='ew', pady=15)
            
            title_label = ttk.Label(
                self.summary_grid,
                text="Centros sin recomendación",
                font=('Segoe UI', 16, 'bold'),
                foreground='#EF4444',
                background='white'
            )
            title_label.grid(row=last_row + 2, column=0, columnspan=3, sticky='w', padx=5, pady=(0, 15))
            
            for i, centro in enumerate(sorted(self.non_compatible), 1):
                center_frame = self.create_unique_center_frame(
                    self.summary_grid,
                    centro,
                    self.unique_portfolios[centro],
                    non_compatible=True
                )
                row = last_row + 3 + ((i - 1) // 3)
                col = (i - 1) % 3
                center_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                self.summary_grid.grid_rowconfigure(row, weight=1)

    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx;*.xls")]
        )
        if filename:
            self.file_path_var.set(filename)
            self.status_var.set("Archivo seleccionado - Listo para analizar")
            
    def find_identical_and_unique_portfolios(self, file_path):
        try:
            # Leer el archivo Excel
            df = pd.read_excel(file_path, sheet_name='Sheet1')
            
            # Validar columnas requeridas
            try:
                columns = self.column_validator.validate_required_columns(
                    df, ['CENTRO', 'PLU_SAP']
                )
                centro_col = columns['CENTRO']
                plu_col = columns['PLU_SAP']
            except ValueError as e:
                raise Exception(f"Error en la validación de columnas: {str(e)}")
                
            # Verificar si hay datos
            if df.empty:
                raise Exception("El archivo Excel está vacío")
            
            # Verificar valores nulos
            null_counts = df[[centro_col, plu_col]].isnull().sum()
            if null_counts.any():
                null_info = "\n".join([
                    f"{col}: {count} valores nulos" 
                    for col, count in null_counts.items() 
                    if count > 0
                ])
                raise Exception(f"Se encontraron valores nulos en:\n{null_info}")
            
            # Limpiar y procesar datos
            df[centro_col] = df[centro_col].astype(str).str.strip()
            df[plu_col] = df[plu_col].astype(str).str.strip()
            
            # Procesar los datos y convertir explícitamente a sets
            centro_portfolios = df.groupby(centro_col)[plu_col].apply(set).reset_index()
                
            portfolio_dict = {}
            for idx, row in centro_portfolios.iterrows():
                portfolio = tuple(sorted(row[plu_col]))
                centro = str(row[centro_col])
                if portfolio in portfolio_dict:
                    portfolio_dict[portfolio].append(centro)
                else:
                    portfolio_dict[portfolio] = [centro]
            
            # Separar portafolios idénticos y únicos, manteniendo los PLUs como sets
            identical_portfolios = {
                tuple(sorted(centros)): set(portfolio)  # Convertir explícitamente a set
                for portfolio, centros in portfolio_dict.items() 
                if len(centros) > 1
            }
            
            unique_portfolios = {
                centros[0]: set(portfolio)  # Convertir explícitamente a set
                for portfolio, centros in portfolio_dict.items() 
                if len(centros) == 1
            }
            
            return identical_portfolios, unique_portfolios
            
        except Exception as e:
            raise Exception(f"Error al procesar el archivo Excel: {str(e)}")

    def calculate_plu_differences_multi(self, centers_plus_dict):
        """
        Calcula la diferencia simétrica de PLUs entre múltiples centros.
        
        Args:
            centers_plus_dict: Diccionario donde las claves son los centros y los valores son sus conjuntos de PLUs
            
        Returns:
            tuple: (symmetric_diff_size, center_ref, all_differences)
                - symmetric_diff_size: Tamaño de la diferencia simétrica
                - center_ref: Centro de referencia (primer centro)
                - all_differences: Conjunto de todos los PLUs diferentes entre los centros
        """
        if len(centers_plus_dict) < 2:
            return 0, None, set()
        
        # Tomar el primer conjunto como referencia
        center_ref, base_plus = next(iter(centers_plus_dict.items()))
        all_differences = set()
        
        # Calcular diferencia simétrica acumulativa
        for center, plus_set in centers_plus_dict.items():
            if center != center_ref:
                diff = base_plus.symmetric_difference(plus_set)
                all_differences.update(diff)
        
        return len(all_differences), center_ref, all_differences

    def analyze_group_mergers(self, plu_limit):
        """
        Encuentra fusiones de grupos usando un enfoque voraz optimizado.
        Solo usa diferencias simétricas de 10 PLUs como límite.
        """
        group_recommendations = []
        groups = list(self.identical_portfolios.items())
        n = len(groups)
        used_groups = set()

        # Precalcular y almacenar solo las mejores conexiones para cada grupo
        group_connections = {i: [] for i in range(n)}
        
        # Calcular diferencias simétricas solo una vez
        for i in range(n):
            for j in range(i + 1, n):
                diff = groups[i][1].symmetric_difference(groups[j][1])
                if len(diff) <= plu_limit: 
                    group_connections[i].append((j, diff))
                    group_connections[j].append((i, diff))
        
        def build_group_from_seed(seed_idx):
            """Construye un grupo comenzando desde un grupo semilla."""
            if seed_idx in used_groups:
                return None
                
            current_group = {seed_idx}
            group_plus = set(groups[seed_idx][1])
            current_differences = set()
            
            # Conjunto de grupos candidatos conectados al grupo actual
            candidates = set(idx for idx, _ in group_connections[seed_idx])
            
            while candidates:
                best_addition = None
                min_new_diffs = float('inf')
                best_total_diffs = None
                
                # Probar cada candidato
                for candidate in candidates:
                    if candidate in used_groups:
                        continue
                    
                    # Calcular nuevas diferencias al agregar este candidato
                    test_diffs = current_differences.copy()
                    for member in current_group:
                        for connected_idx, diff in group_connections[member]:
                            if connected_idx == candidate:
                                test_diffs.update(diff)
                    
                    if len(test_diffs) <= plu_limit and len(test_diffs) < min_new_diffs:  # Cambiado a 10 PLUs
                        min_new_diffs = len(test_diffs)
                        best_addition = candidate
                        best_total_diffs = test_diffs
                
                if best_addition is None:
                    break
                    
                # Agregar el mejor candidato al grupo
                current_group.add(best_addition)
                current_differences = best_total_diffs
                
                # Actualizar candidatos con nuevas conexiones
                new_candidates = set(idx for idx, _ in group_connections[best_addition])
                candidates = (candidates | new_candidates) - current_group - used_groups
            
            if len(current_group) > 1:
                return {
                    'groups': [groups[idx][0] for idx in current_group],
                    'differences': current_differences,
                    'indices': current_group
                }
            return None

        # Procesar grupos en orden de más conexiones a menos
        group_order = sorted(
            range(n), 
            key=lambda x: len(group_connections[x]), 
            reverse=True
        )
        
        # Construir grupos
        for seed_idx in group_order:
            result = build_group_from_seed(seed_idx)
            if result:
                group_recommendations.append({
                    'Groups': result['groups'],
                    'PLU_Differences': sorted(list(result['differences']))
                })
                used_groups.update(result['indices'])
                
                # Verificar si ya procesamos todos los grupos
                if len(used_groups) == n:
                    break

        return group_recommendations

    def analyze_unique_portfolios(self, identical_portfolios, unique_portfolios, plu_limit):
        """
        Analiza los portafolios únicos con enfoque optimizado.
        Usa límite de 10 PLUs para diferencias simétricas.
        """
        recommendations = []
        non_compatible = []
        assigned_centers = set()
        unique_groups = []
        
        # Convertir grupos existentes a formato manejable
        existing_groups = []
        for idx, (centers, plu_set) in enumerate(identical_portfolios.items()):
            existing_groups.append({
                'centers': set(centers),
                'plus': set(plu_set),
                'group_number': idx + 1
            })
        
        # Pre-calcular diferencias para centros únicos
        center_differences = {}
        for center in unique_portfolios:
            center_plus = set(unique_portfolios[center])
            best_diff = float('inf')
            best_group = None
            best_differences = None
            
            for group in existing_groups:
                diff = center_plus.symmetric_difference(group['plus'])
                if len(diff) <= plu_limit and len(diff) < best_diff:  # Cambiado a 10 PLUs
                    best_diff = len(diff)
                    best_group = group
                    best_differences = diff
            
            if best_group:
                center_differences[center] = {
                    'group': best_group,
                    'differences': best_differences,
                    'diff_size': len(best_differences)
                }
        
        # Asignar centros a grupos existentes basado en las mejores diferencias
        for center, data in sorted(
            center_differences.items(), 
            key=lambda x: x[1]['diff_size']
        ):
            if center not in assigned_centers:
                recommendations.append({
                    'Centro Único': center,
                    'Grupo Sugerido': tuple(sorted(data['group']['centers'])),
                    'Grupo Número': data['group']['group_number'],
                    'PLU Diferentes': sorted(list(data['differences']))
                })
                assigned_centers.add(center)
        
        # Procesar centros restantes para nuevos grupos
        remaining_centers = set(unique_portfolios.keys()) - assigned_centers
        
        # Pre-calcular diferencias entre centros restantes
        remaining_connections = {}
        for center1 in remaining_centers:
            plus1 = set(unique_portfolios[center1])
            remaining_connections[center1] = []
            
            for center2 in remaining_centers:
                if center1 != center2:
                    diff = plus1.symmetric_difference(unique_portfolios[center2])
                    if len(diff) <= plu_limit:  # Cambiado a 10 PLUs
                        remaining_connections[center1].append((center2, diff))
        
        # Construir nuevos grupos usando enfoque voraz
        while remaining_centers:
            # Seleccionar centro con más conexiones como semilla
            seed_center = max(
                remaining_centers,
                key=lambda x: len(remaining_connections.get(x, []))
            )
            
            current_group = {seed_center}
            current_differences = set()
            candidates = set(c for c, _ in remaining_connections.get(seed_center, []))
            
            while candidates:
                best_addition = None
                min_new_diffs = float('inf')
                best_total_diffs = None
                
                for candidate in candidates:
                    if candidate not in remaining_centers:
                        continue
                    
                    # Calcular nuevas diferencias
                    test_diffs = current_differences.copy()
                    seed_plus = set(unique_portfolios[seed_center])
                    candidate_plus = set(unique_portfolios[candidate])
                    new_diff = seed_plus.symmetric_difference(candidate_plus)
                    test_diffs.update(new_diff)
                    
                    if len(test_diffs) <= plu_limit and len(test_diffs) < min_new_diffs:  # Cambiado a 10 PLUs
                        min_new_diffs = len(test_diffs)
                        best_addition = candidate
                        best_total_diffs = test_diffs
                
                if best_addition is None:
                    break
                    
                current_group.add(best_addition)
                current_differences = best_total_diffs
                candidates = set(c for c, _ in remaining_connections.get(best_addition, [])) - current_group
            
            # Procesar el grupo formado
            if len(current_group) > 1:
                unique_groups.append({
                    'Centros': sorted(list(current_group)),
                    'PLU Diferentes': sorted(list(current_differences))
                })
            else:
                non_compatible.extend(current_group)
                
            # Remover centros procesados
            remaining_centers -= current_group
        
        return recommendations, sorted(non_compatible), unique_groups

    def analyze_portfolios(self):
        file_path = self.file_path_var.get()
        if not file_path:
            messagebox.showerror("Error", "Por favor seleccione un archivo Excel")
            return
        
        # Obtener el límite de PLUs diferentes
        plu_limit = self.get_plu_limit()
        if plu_limit is None:  # Usuario canceló
            return
            
        self.current_plu_limit = plu_limit

        try:
            # Crear y mostrar el spinner
            loading_window = tk.Toplevel(self.root)
            loading_window.overrideredirect(True)
            loading_window.transient(self.root)
            loading_window.configure(bg='white')
            
            # Configurar el tamaño y posición
            window_width = 300
            window_height = 100
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            # Calcular posición para centrar
            x = (screen_width - window_width) // 2
            y = (screen_height - window_height) // 2
            
            # Establecer geometría y atributos
            loading_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
            loading_window.attributes('-topmost', True)
            
            # Agregar un borde
            loading_window.configure(borderwidth=1, relief='solid')
            
            # Crear un marco con borde redondeado
            loading_frame = ttk.Frame(loading_window, style="Card.TFrame")
            loading_frame.pack(expand=True, fill='both', padx=2, pady=2)
            
            # Canvas para el spinner
            canvas_size = 40
            canvas = tk.Canvas(
                loading_frame,
                width=canvas_size,
                height=canvas_size,
                bg='white',
                highlightthickness=0
            )
            canvas.pack(pady=(10, 5))
            
            # Crear el spinner circular
            def create_arc(start, extent):
                padding = 5
                canvas.create_arc(
                    padding, padding,
                    canvas_size - padding, canvas_size - padding,
                    start=start,
                    extent=extent,
                    fill='#2563eb',
                    width=0
                )
            
            # Texto de estado
            status_label = ttk.Label(
                loading_frame,
                text="Analizando archivo...",
                style="StatsDesc.TLabel"
            )
            status_label.pack(pady=(0, 10))
            
            rotation = [0]
            
            def update_spinner():
                if loading_window.winfo_exists():
                    canvas.delete("all")
                    create_arc(rotation[0], 300)
                    rotation[0] = (rotation[0] + 10) % 360
                    loading_window.after(10, update_spinner)
            
            # Iniciar la animación del spinner
            update_spinner()
            
            # Forzar la actualización de la ventana
            loading_window.update()
            loading_window.deiconify()
            
            # Hacer que la ventana principal procese eventos
            self.root.update_idletasks()
            self.root.update()
            
            # Limpiar grupos existentes
            status_label.config(text="Preparando análisis...")
            loading_window.update()
            
            for widget in self.groups_grid.winfo_children():
                widget.destroy()
            for widget in self.unique_grid.winfo_children():
                widget.destroy()
            for widget in self.recommendations_grid.winfo_children():
                widget.destroy()
            for widget in self.group_recommendations_grid.winfo_children():
                widget.destroy()
            
            # Resetear el contador de letras
            self.letter_counter = 0
            
            # Realizar el análisis
            status_label.config(text="AAAAnalizando portafolios...")
            loading_window.update()
            
            self.identical_portfolios, self.unique_portfolios = self.find_identical_and_unique_portfolios(file_path)
            
            # Obtener recomendaciones
            status_label.config(text="Generando recomendaciones...")
            loading_window.update()
            
            # Pasar el límite a los métodos que lo necesitan
            self.recommendations, self.non_compatible, unique_groups = self.analyze_unique_portfolios(
                self.identical_portfolios,
                self.unique_portfolios,
                plu_limit
            )
            
            # Obtener recomendaciones de fusión
            status_label.config(text="Analizando fusiones posibles...")
            loading_window.update()
            
            group_recommendations = self.analyze_group_mergers(plu_limit)
            
            # Convertir no compatibles a set
            self.non_compatible_set = set(self.non_compatible)
            
            # Calcular estadísticas
            status_label.config(text="Calculando estadísticas...")
            loading_window.update()
            
            unique_count = len(self.unique_portfolios)
            identical_count = sum(len(centers) for centers in self.identical_portfolios.keys())
            initial_masters = len(self.identical_portfolios) + unique_count
            final_groups = self.calculate_final_groups(plu_limit)
            if not final_groups and not self.non_compatible:
                # Si no hay grupos finales ni centros no compatibles, crear un grupo vacío
                final_groups = [{'centers': set(), 'plus': set()}] 
            final_masters = len(final_groups) + len(self.non_compatible)

            # Usa animate_stat para cada estadística
            self.animate_stat(self.total_centers, identical_count + unique_count)
            self.animate_stat(self.unique_centers, unique_count)
            self.animate_stat(self.identical_groups, len(self.identical_portfolios))
            self.animate_stat(self.initial_masters, initial_masters)
            self.animate_stat(self.final_masters, final_masters)
            
            # Actualizar interfaz
            status_label.config(text="Actualizando interfaz...")
            loading_window.update()
            
            # Actualizar interfaz con todas las estadísticas
            self.distribution_chart(unique_count, identical_count)
            self.total_centers.set(str(identical_count + unique_count))
            self.unique_centers.set(str(unique_count))
            self.identical_groups.set(str(len(self.identical_portfolios)))
            self.initial_masters.set(str(initial_masters))
            self.final_masters.set(str(final_masters))
            
            # Actualizar grids en lotes
            status_label.config(text="Actualizando resultados...")
            loading_window.update()
            
            # Mostrar portafolios idénticos
            batch_size = 10
            for i, (centros, productos) in enumerate(self.identical_portfolios.items(), 1):
                if i % batch_size == 0:
                    status_label.config(text=f"Procesando grupo {i} de {len(self.identical_portfolios)}...")
                    loading_window.update()
                    self.root.update_idletasks()
                
                group_frame = self.create_group_frame(
                    parent=self.groups_grid,
                    group_num=i,
                    centers=", ".join(centros),
                    count=len(centros),
                    timestamp=datetime.now().strftime("%Y-%m-%d %H:%M"),
                    plu_count=len(productos)
                )
                row = (i - 1) // 3
                col = (i - 1) % 3
                group_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                self.groups_grid.grid_rowconfigure(row, weight=1)
            
            # Mostrar portafolios únicos
            for i, (centro, plu_list) in enumerate(self.unique_portfolios.items(), 1):
                if i % batch_size == 0:
                    status_label.config(text=f"Procesando centro único {i} de {len(self.unique_portfolios)}...")
                    loading_window.update()
                    self.root.update_idletasks()
                
                is_non_compatible = centro in self.non_compatible
                center_frame = self.create_unique_center_frame(
                    self.unique_grid,
                    centro,
                    plu_list,
                    non_compatible=is_non_compatible
                )
                row = (i - 1) // 3
                col = (i - 1) % 3
                center_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                self.unique_grid.grid_rowconfigure(row, weight=1)
            
            # Agrupar recomendaciones por grupo sugerido
            grouped_recommendations = {}
            for rec in self.recommendations:
                group_key = (rec['Grupo Número'], tuple(sorted(rec['Grupo Sugerido'])))
                if group_key not in grouped_recommendations:
                    grouped_recommendations[group_key] = {
                        'Centros Únicos': [rec['Centro Único']],
                        'PLU Diferentes': set(rec['PLU Diferentes'])
                    }
                else:
                    grouped_recommendations[group_key]['Centros Únicos'].append(rec['Centro Único'])
                    grouped_recommendations[group_key]['PLU Diferentes'].update(rec['PLU Diferentes'])
            
            # Ordenar y clasificar las recomendaciones
            multiple_centers_recs = []
            single_center_recs = []
            
            for (group_number, group_centers), data in grouped_recommendations.items():
                if len(data['Centros Únicos']) > 1:
                    multiple_centers_recs.append((group_number, group_centers, data))
                else:
                    single_center_recs.append((group_number, group_centers, data))
            
            # Obtener el último número de grupo de portafolios idénticos
            last_group_number = len(self.identical_portfolios)
            
            # Posibles grupos (PG X)
            possible_groups = [(i, group) for i, group in enumerate(unique_groups, last_group_number + 1)]
            
            status_label.config(text="Mostrando recomendaciones...")
            loading_window.update()
            
            # Mostrar recomendaciones en el orden deseado
            row = 0
            col = 0
            
            # 1. Primero mostrar grupos de múltiples centros
            for group_number, group_centers, data in multiple_centers_recs:
                rec_frame = self.create_grouped_recommendation_frame(
                    self.recommendations_grid,
                    data['Centros Únicos'],
                    group_centers,
                    group_number,
                    sorted(list(data['PLU Diferentes']))
                )
                rec_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                self.recommendations_grid.grid_rowconfigure(row, weight=1)
                
                col += 1
                if col == 3:
                    col = 0
                    row += 1
            
            # 2. Luego mostrar recomendaciones de centro único
            for group_number, group_centers, data in single_center_recs:
                rec_frame = self.create_grouped_recommendation_frame(
                    self.recommendations_grid,
                    data['Centros Únicos'],
                    group_centers,
                    group_number,
                    sorted(list(data['PLU Diferentes']))
                )
                rec_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                self.recommendations_grid.grid_rowconfigure(row, weight=1)
                
                col += 1
                if col == 3:
                    col = 0
                    row += 1
            
            # 3. Finalmente mostrar los posibles grupos
            for group_num, group in possible_groups:
                rec_frame = self.create_unique_recommendation_frame(
                    self.recommendations_grid,
                    group['Centros'],
                    group['PLU Diferentes'],
                    group_num
                )
                rec_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                self.recommendations_grid.grid_rowconfigure(row, weight=1)
                
                col += 1
                if col == 3:
                    col = 0
                    row += 1
            
            # Mostrar recomendaciones de fusión
            status_label.config(text="Procesando recomendaciones de fusión...")
            loading_window.update()
            
            for i, rec in enumerate(group_recommendations, 1):
                rec_frame = self.create_group_merger_frame(
                    self.group_recommendations_grid,
                    rec['Groups'],
                    rec['PLU_Differences']
                )
                row = (i - 1) // 3
                col = (i - 1) % 3
                rec_frame.grid(row=row, column=col, padx=5, pady=5, sticky='ew')
                self.group_recommendations_grid.grid_rowconfigure(row, weight=1)
            
            # Actualizar pestaña de resumen
            status_label.config(text="Generando resumen...")
            loading_window.update()
            
            self.update_summary_tab()
            
            # Habilitar botón de reportes
            self.report_button.configure(state="normal")
            
            # Actualizar estado final
            self.status_var.set("Análisis completado")
            
            # Destruir ventana de carga
            loading_window.destroy()
            self.add_custom_grouping_tab()
                
        except Exception as e:
            if 'loading_window' in locals():
                loading_window.destroy()
            self.status_var.set("Error en el análisis")
            messagebox.showerror("Error", str(e))

def main():
    root = tk.Tk()
    app = ModernPortfolioAnalyzerApp(root)      
    # Verificar actualizaciones después de 3 segundos
    def delayed_update_check():
        root.after(3000, check_for_updates)
    
    # Iniciar verificación después de que la app esté corriendo
    root.after(1, delayed_update_check)    
    root.mainloop()

if __name__ == "__main__":
    main()

    